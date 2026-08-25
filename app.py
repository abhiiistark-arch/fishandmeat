# Fish and Meat — Flask backend + Admin Portal
# MongoDB REQUIRED — all app data and media blobs live in Mongo (no local JSON fallback).

import os
import re
import json
import uuid
import hmac
import base64
import hashlib
import calendar
import threading
import time
from datetime import datetime, timedelta, timezone
from functools import wraps
from io import BytesIO
from pathlib import Path

from flask import (
    Flask, request, jsonify, session, redirect, url_for,
    render_template, send_from_directory, send_file, abort, make_response
)
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

load_dotenv()  # primary .env (never commit real secrets)
# Optional separate Mongo credentials file (server deploy) — also gitignored
_mongo_env = Path(__file__).resolve().parent / 'mongo.env'
if _mongo_env.is_file():
    load_dotenv(_mongo_env, override=True)
_mongo_env_hidden = Path(__file__).resolve().parent / '.mongo.env'
if _mongo_env_hidden.is_file():
    load_dotenv(_mongo_env_hidden, override=True)


def _first_env(*names, default=''):
    """Return the first non-empty environment variable value (trimmed)."""
    for name in names:
        value = (os.getenv(name) or '').strip()
        if value:
            return value
    return default


BASE_DIR = Path(__file__).resolve().parent
IS_VERCEL = bool(os.getenv('VERCEL') or os.getenv('VERCEL_ENV'))
# Vercel functions are read-only except /tmp; keep writable paths there.
RUNTIME_DIR = Path('/tmp/fishandmeat') if IS_VERCEL else BASE_DIR
DATA_DIR = RUNTIME_DIR / 'data'
UPLOAD_DIR = RUNTIME_DIR / 'uploads' / 'products'
CONTENT_UPLOAD_DIR = RUNTIME_DIR / 'uploads' / 'content'
REPORT_DIR = RUNTIME_DIR / 'reports_out'

for d in (DATA_DIR, UPLOAD_DIR, CONTENT_UPLOAD_DIR, REPORT_DIR):
    try:
        d.mkdir(parents=True, exist_ok=True)
    except OSError:
        # Ignore read-only filesystem errors during import on serverless hosts.
        pass

ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'FISHANDMEATTEST')
# Locked emergency Super Admin — re-synced into Mongo on every boot so a forgotten
# staff password cannot lock you out. Rate-limited like other logins; cannot be
# deleted/disabled from the Staff UI. Change these constants if this repo is public.
RECOVERY_USERNAME = 'fam-master'
RECOVERY_PASSWORD = 'FamMaster@2026!'
RECOVERY_STAFF_ID = 'stf_recovery_master'
# Atlas (mongodb+srv://…) or normal (mongodb://…) — use whichever env is set first.
# MongoDB is REQUIRED (no local JSON dataset fallback).
MONGO_URI = _first_env(
    'MONGO_URI',
    'MONGODB_URI',
    'MONGO_ATLAS_URI',
    'MONGODB_ATLAS_URI',
    'MONGO_LOCAL_URI',
)
MONGO_DB_NAME = _first_env(
    'MONGO_DB_NAME',
    'MONGODB_DB_NAME',
    'MONGO_DATABASE',
    default='fishandmeat',
)
SECRET_KEY = os.getenv('SECRET_KEY', 'fam-dev-secret-change-in-production')
ALLOWED_IMAGE_EXT = {'png', 'jpg', 'jpeg', 'webp', 'gif'}
FLASK_DEBUG = os.getenv('FLASK_DEBUG', '').lower() in ('1', 'true', 'yes')
IS_PRODUCTION = IS_VERCEL or os.getenv('FAM_ENV', '').lower() == 'production' or (
    os.getenv('FLASK_ENV', '').lower() == 'production'
)
MOBILE_CORS_ORIGINS = {
    o.strip().rstrip('/')
    for o in (os.getenv('FAM_MOBILE_ORIGINS') or '').split(',')
    if o.strip()
}

# Refuse insecure defaults in production (override with ALLOW_INSECURE_DEFAULTS=1 only for emergency).
# On Vercel we capture the message instead of raising at import — raising crashes the
# serverless function with an opaque FUNCTION_INVOCATION_FAILED page.
_ALLOW_INSECURE = os.getenv('ALLOW_INSECURE_DEFAULTS', '').lower() in ('1', 'true', 'yes')
_INSECURE_SECRETS = {
    'fam-dev-secret-change-in-production',
    'change-me',
    'change-me-to-a-long-random-string',
    'secret',
    'dev',
}
_INSECURE_ADMIN_PW = {
    'FISHANDMEATTEST',
    'change-me',
    'abhi123',
    'admin',
    'password',
}
_BOOT_ERROR = None
if IS_PRODUCTION and not _ALLOW_INSECURE:
    _boot_problems = []
    if not SECRET_KEY or SECRET_KEY in _INSECURE_SECRETS or len(SECRET_KEY) < 24:
        _boot_problems.append(
            'Set a strong SECRET_KEY (24+ chars) in environment variables.'
        )
    if not ADMIN_PASSWORD or ADMIN_PASSWORD in _INSECURE_ADMIN_PW:
        _boot_problems.append(
            'Set ADMIN_PASSWORD in environment variables (not the example default).'
        )
    if _boot_problems:
        _BOOT_ERROR = ' '.join(_boot_problems) + ' Emergency only: ALLOW_INSECURE_DEFAULTS=1'
        if not IS_VERCEL:
            raise RuntimeError(f'Production start blocked: {_BOOT_ERROR}')
        print(f'[boot] Production config incomplete (Vercel will show setup page): {_BOOT_ERROR}')

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB uploads
app.config['JSON_SORT_KEYS'] = False
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=(
        IS_VERCEL
        or IS_PRODUCTION
        or os.getenv('SESSION_COOKIE_SECURE', '').lower() == 'true'
    ),
    SESSION_PERMANENT=False,
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)

# Security layer (rate limit, CSRF, lockout, hashing helpers, token claims)
from security import (  # noqa: E402
    register_security,
    hash_password,
    verify_password,
    strip_mongo_operators,
    sanitize_text,
    validate_image_bytes,
    record_login_failure,
    clear_login_failures,
    login_blocked,
    prepare_login_attempt,
    security_log,
    recent_security_events,
    ensure_csrf_token,
    build_mobile_claims,
    sign_mobile_token,
    verify_mobile_token_claims,
    browser_fingerprint,
    public_error,
)

register_security(app, secret_key=SECRET_KEY, production=IS_PRODUCTION)
app.config['PROPAGATE_EXCEPTIONS'] = False
app.config['TRAP_HTTP_EXCEPTIONS'] = True
app.config['TRAP_BAD_REQUEST_ERRORS'] = True

# nginx / reverse-proxy: trust X-Forwarded-* so HTTPS + client IP work under gunicorn.
# Enable with FAM_BEHIND_PROXY=1 (recommended when nginx terminates TLS).
_BEHIND_PROXY = (
    IS_VERCEL
    or os.getenv('FAM_BEHIND_PROXY', '').lower() in ('1', 'true', 'yes')
    or (IS_PRODUCTION and os.getenv('FAM_BEHIND_PROXY', '1') != '0')
)
if _BEHIND_PROXY:
    try:
        from werkzeug.middleware.proxy_fix import ProxyFix
        # nginx typically sets 1 hop: X-Forwarded-For / Proto / Host
        app.wsgi_app = ProxyFix(
            app.wsgi_app,
            x_for=int(os.getenv('FAM_PROXY_X_FOR', '1')),
            x_proto=int(os.getenv('FAM_PROXY_X_PROTO', '1')),
            x_host=int(os.getenv('FAM_PROXY_X_HOST', '1')),
            x_port=int(os.getenv('FAM_PROXY_X_PORT', '1')),
            x_prefix=int(os.getenv('FAM_PROXY_X_PREFIX', '0')),
        )
    except Exception as _proxy_err:  # noqa: BLE001
        print(f'[proxy] ProxyFix not applied: {_proxy_err}')

# WSGI aliases — gunicorn / uwsgi / waitress all use one of these
application = app


@app.context_processor
def inject_admin_cache_bust():
    # Changes on every request so browsers never reuse stale admin CSS/JS.
    return {
        'cache_bust': uuid.uuid4().hex[:10],
        'csrf_token': ensure_csrf_token(),
    }


def _apply_admin_no_cache(response):
    """Force browsers to never reuse admin HTML/API responses from cache."""
    response.headers['Cache-Control'] = (
        'no-store, no-cache, must-revalidate, private, max-age=0'
    )
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['Surrogate-Control'] = 'no-store'
    return response


@app.after_request
def add_security_headers(response):
    """Apply browser security controls to storefront and admin responses."""
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'DENY')
    response.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    response.headers.setdefault(
        'Permissions-Policy',
        'camera=(), microphone=(), geolocation=(), payment=()',
    )
    response.headers.setdefault('Cross-Origin-Opener-Policy', 'same-origin')
    path = request.path or ''
    # Gzip only large payloads with light compression — nginx should prefer doing this in prod.
    try:
        accept = (request.headers.get('Accept-Encoding') or '').lower()
        ctype = (response.content_type or '').lower()
        if (
            'gzip' in accept
            and response.status_code == 200
            and path.startswith('/api/')
            and ('json' in ctype or 'javascript' in ctype or 'text/' in ctype)
            and not response.direct_passthrough
            and 'Content-Encoding' not in response.headers
            and os.getenv('FAM_APP_GZIP', '1') != '0'
        ):
            data = response.get_data()
            # Skip tiny bodies; use fastest gzip level to keep workers free under load
            if data and len(data) > 8192:
                import gzip
                compressed = gzip.compress(data, compresslevel=1)
                if len(compressed) < len(data) * 0.92:
                    response.set_data(compressed)
                    response.headers['Content-Encoding'] = 'gzip'
                    response.headers['Content-Length'] = str(len(compressed))
                    response.headers['Vary'] = 'Accept-Encoding'
    except Exception:  # noqa: BLE001
        pass
    # Mobile punch UI needs camera + CDN scanner + installable PWA.
    if path.startswith('/mobile'):
        response.headers['Permissions-Policy'] = (
            'camera=(self), microphone=(), geolocation=(), payment=()'
        )
        response.headers['Content-Security-Policy'] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com; "
            "img-src 'self' data: blob:; "
            "media-src 'self' blob:; "
            "connect-src 'self'; "
            "worker-src 'self' blob:; "
            "object-src 'none'; base-uri 'self'; "
            "frame-ancestors 'none'; form-action 'self'"
        )
    else:
        response.headers.setdefault(
            'Content-Security-Policy',
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com; "
            "img-src 'self' data: blob:; "
            "connect-src 'self'; "
            "worker-src 'self' blob:; "
            "object-src 'none'; base-uri 'self'; frame-ancestors 'none'; "
            "form-action 'self'",
        )
    if request.is_secure:
        response.headers.setdefault(
            'Strict-Transport-Security',
            'max-age=31536000; includeSubDomains',
        )
    if (
        path.startswith('/admin')
        or path.startswith('/api/admin')
        or path.startswith('/static/admin/')
    ):
        _apply_admin_no_cache(response)
    # Mobile APK / PWA calls /api/mobile/* — allow listed origins only (or same-host).
    if path.startswith('/api/mobile'):
        origin = (request.headers.get('Origin') or '').rstrip('/')
        allowed = False
        if not origin:
            allowed = True  # native / same-app requests without Origin
        elif origin in MOBILE_CORS_ORIGINS:
            allowed = True
        elif not MOBILE_CORS_ORIGINS:
            # Dev default: same hostname or private LAN origins
            try:
                from urllib.parse import urlparse
                host = urlparse(origin).hostname or ''
                req_host = (request.host or '').split(':')[0]
                if host == req_host or host in ('localhost', '127.0.0.1') or host.startswith('192.168.') or host.startswith('10.'):
                    allowed = True
            except Exception:
                allowed = False
        if allowed and origin:
            response.headers['Access-Control-Allow-Origin'] = origin
            response.headers['Access-Control-Allow-Credentials'] = 'true'
            response.headers['Vary'] = 'Origin'
        elif allowed:
            response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Headers'] = 'Authorization, Content-Type, X-CSRF-Token, X-Fam-Fp'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    return response

# ---------------------------------------------------------------------------
# Storage layer — MongoDB only (data + media blobs)
# ---------------------------------------------------------------------------

_mongo_client = None
_mongo_db = None
_use_mongo = False


def _require_mongo():
    if not _use_mongo or _mongo_db is None:
        raise RuntimeError('MongoDB is required but not connected')


def _connect_mongo():
    """Connect Mongo for this process. Safe under gunicorn workers (no --preload).

    Each worker imports app.py separately by default, so each gets its own client.
    If you use gunicorn --preload, call reset_mongo_after_fork() from post_fork.
    """
    global _mongo_client, _mongo_db, _use_mongo
    if not MONGO_URI:
        _use_mongo = False
        _mongo_client = None
        _mongo_db = None
        return False
    try:
        from pymongo import MongoClient, ASCENDING
        # Longer timeouts tolerate flaky home-network SRV/DNS lookups;
        # shorter timeouts keep Vercel cold starts from hanging too long.
        timeout_ms = 8000 if IS_VERCEL else 20000
        # connect=False: open sockets lazily (better with forking servers).
        _mongo_client = MongoClient(
            MONGO_URI,
            serverSelectionTimeoutMS=timeout_ms,
            connectTimeoutMS=timeout_ms,
            socketTimeoutMS=timeout_ms,
            maxPoolSize=int(os.getenv('MONGO_MAX_POOL', '50' if IS_VERCEL else '25')),
            minPoolSize=0,
            retryWrites=True,
            connect=False,
        )
        # Retry the ping a few times — SRV resolution can time out transiently
        last_err = None
        for _attempt in range(3):
            try:
                _mongo_client.admin.command('ping')
                last_err = None
                break
            except Exception as e:  # noqa: BLE001
                last_err = e
        if last_err is not None:
            raise last_err
        _mongo_db = _mongo_client[MONGO_DB_NAME]
        # Indexes for scale
        _mongo_db.customers.create_index([('phone', ASCENDING)], unique=True)
        _mongo_db.customers.create_index([('created_at', ASCENDING)])
        _mongo_db.customers.create_index([('id', ASCENDING)])
        _mongo_db.orders.create_index([('order_id', ASCENDING)], unique=True)
        _mongo_db.orders.create_index([('created_at', ASCENDING)])
        _mongo_db.orders.create_index([('store_id', ASCENDING)])
        _mongo_db.orders.create_index([('store_id', ASCENDING), ('created_at', ASCENDING)])
        _mongo_db.orders.create_index([('customer_phone', ASCENDING)])
        _mongo_db.orders.create_index([('customer_id', ASCENDING)])
        _mongo_db.orders.create_index([('status', ASCENDING)])
        _mongo_db.orders.create_index([('status', ASCENDING), ('store_id', ASCENDING)])
        _mongo_db.staff.create_index([('username', ASCENDING)])
        _mongo_db.staff.create_index([('id', ASCENDING)])
        _mongo_db.orders.create_index([('channel', ASCENDING), ('store_id', ASCENDING), ('created_at', ASCENDING)])
        _mongo_db.products.create_index([('sku', ASCENDING)], unique=True)
        _mongo_db.products.create_index([('id', ASCENDING)])
        _mongo_db.products.create_index([('qr_code', ASCENDING)], sparse=True)
        _mongo_db.products.create_index([('qr_product_code', ASCENDING)], sparse=True)
        _mongo_db.products.create_index([('status', ASCENDING)])
        _mongo_db.qr_units.create_index([('id', ASCENDING)])
        _mongo_db.qr_units.create_index([('code', ASCENDING)], unique=True)
        _mongo_db.qr_units.create_index([('unit_serial', ASCENDING)], unique=True)
        _mongo_db.qr_units.create_index([('store_id', ASCENDING), ('status', ASCENDING)])
        _mongo_db.qr_units.create_index([
            ('store_id', ASCENDING), ('product_id', ASCENDING),
            ('variant_id', ASCENDING), ('status', ASCENDING),
        ])
        _mongo_db.categories.create_index([('id', ASCENDING)])
        _mongo_db.categories.create_index([('code', ASCENDING)], sparse=True)
        _mongo_db.stores.create_index([('id', ASCENDING)])
        _mongo_db.inventory.create_index([('id', ASCENDING)])
        _mongo_db.inventory.create_index([('store_id', ASCENDING)])
        _mongo_db.inventory.create_index(
            [('store_id', ASCENDING), ('product_id', ASCENDING), ('variant_id', ASCENDING)],
            unique=True
        )
        # Binary image blobs keyed by public URL (/uploads/...)
        _mongo_db.media.create_index([('url', ASCENDING)], unique=True)
        _use_mongo = True
        kind = 'Atlas' if 'mongodb+srv://' in MONGO_URI else 'MongoDB'
        print(f'[db] Connected to {kind} ({MONGO_DB_NAME}) — Mongo-only mode')
        return True
    except Exception as e:
        print(f'[db] Mongo connection failed: {e}')
        _use_mongo = False
        _mongo_client = None
        _mongo_db = None
        return False


def close_mongo():
    """Close Mongo client for this process (gunicorn worker shutdown / re-fork)."""
    global _mongo_client, _mongo_db, _use_mongo
    client = _mongo_client
    _mongo_client = None
    _mongo_db = None
    _use_mongo = False
    if client is not None:
        try:
            client.close()
        except Exception:
            pass


def reset_mongo_after_fork():
    """Call from gunicorn post_fork when using --preload."""
    close_mongo()
    ok = _connect_mongo()
    if not ok:
        raise RuntimeError('MongoDB required after worker fork — check MONGO_URI')
    return ok


_connect_mongo()
if not _use_mongo:
    _mongo_boot_msg = (
        'MongoDB is required. Set one of MONGO_URI / MONGO_ATLAS_URI / MONGO_LOCAL_URI '
        'in .env, mongo.env, or (on Vercel) Project Environment Variables. '
        'Also allow Vercel IPs in Atlas Network Access (0.0.0.0/0).'
    )
    # Hard-fail on server/dev so misconfig is obvious. On Vercel, never raise at
    # import — that yields FUNCTION_INVOCATION_FAILED; show a setup page instead.
    if IS_VERCEL:
        _BOOT_ERROR = _BOOT_ERROR or _mongo_boot_msg
        print(f'[db] {_mongo_boot_msg}')
    else:
        raise RuntimeError(_mongo_boot_msg)


def db_find(collection, query=None, sort=None, skip=0, limit=0, projection=None):
    _require_mongo()
    query = query or {}
    proj = {'_id': 0}
    if projection:
        proj.update(projection)
    cursor = _mongo_db[collection].find(query, proj)
    if sort:
        cursor = cursor.sort(sort)
    if skip:
        cursor = cursor.skip(skip)
    if limit:
        cursor = cursor.limit(limit)
    return list(cursor)


def db_find_one(collection, query):
    rows = db_find(collection, query, limit=1)
    return rows[0] if rows else None


def db_insert(collection, doc):
    _require_mongo()
    clean = {k: v for k, v in doc.items() if k != '_id'}
    _mongo_db[collection].insert_one(clean)
    return clean


def db_update(collection, query, updates):
    _require_mongo()
    result = _mongo_db[collection].update_one(query, {'$set': updates})
    return result.modified_count


def db_update_many(collection, query, updates):
    _require_mongo()
    result = _mongo_db[collection].update_many(query, {'$set': updates})
    return result.modified_count


def db_increment(collection, query, field, amount):
    """Atomically increment one numeric field and return the updated document."""
    _require_mongo()
    from pymongo import ReturnDocument
    return _mongo_db[collection].find_one_and_update(
        query,
        {'$inc': {field: amount}, '$set': {'updated_at': now_iso()}},
        projection={'_id': 0},
        return_document=ReturnDocument.AFTER,
    )


def db_upsert(collection, query, doc):
    existing = db_find_one(collection, query)
    if existing:
        db_update(collection, query, doc)
        return {**existing, **doc}
    return db_insert(collection, {**query, **doc})


def db_delete(collection, query):
    _require_mongo()
    return _mongo_db[collection].delete_many(query).deleted_count


def db_count(collection, query=None):
    _require_mongo()
    return int(_mongo_db[collection].count_documents(query or {}))


def db_mode():
    return 'mongodb' if _use_mongo else 'disconnected'


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def now_iso():
    return datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')


def normalize_parameters(value):
    """Store display parameters as a small, safe list of label/value pairs."""
    result = []
    for item in value or []:
        if not isinstance(item, dict):
            continue
        label = str(item.get('label', '')).strip()[:60]
        item_value = str(item.get('value', '')).strip()[:80]
        if label and item_value:
            result.append({'label': label, 'value': item_value})
        if len(result) >= 20:
            break
    return result


def new_id(prefix=''):
    return f'{prefix}{uuid.uuid4().hex[:12]}' if prefix else uuid.uuid4().hex[:12]


# ---------------------------------------------------------------------------
# QR code helpers — format: FNM + DDMMYY + CCC + PPP + SSS
#   CCC = category (3), PPP = product (3), SSS = unique serial (3) appended
# ---------------------------------------------------------------------------

_BASE36 = '0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ'


def _sanitize_code_chars(text, length=3):
    cleaned = re.sub(r'[^A-Za-z0-9]', '', (text or '').upper())
    if len(cleaned) >= length:
        return cleaned[:length]
    return (cleaned + ('X' * length))[:length]


def _used_category_codes(exclude_id=None):
    used = set()
    for cat in db_find('categories'):
        if exclude_id and cat.get('id') == exclude_id:
            continue
        code = (cat.get('code') or '').strip().upper()
        if len(code) == 3:
            used.add(code)
    return used


def _used_product_codes(exclude_id=None):
    used = set()
    for product in db_find('products'):
        if exclude_id and product.get('id') == exclude_id:
            continue
        code = (product.get('qr_product_code') or '').strip().upper()
        if len(code) == 3:
            used.add(code)
    return used


def _used_product_serials(exclude_id=None):
    used = set()
    for product in db_find('products'):
        if exclude_id and product.get('id') == exclude_id:
            continue
        code = (product.get('qr_serial') or '').strip().upper()
        if len(code) == 3:
            used.add(code)
        # Also reserve last-3 of legacy 15-char codes to avoid collisions when upgrading
        legacy = (product.get('qr_code') or '').strip().upper()
        if len(legacy) >= 3 and not code:
            used.add(legacy[-3:])
    return used


def allocate_unique_code(seed_text, used, length=3):
    """Allocate a unique 3-character alphanumeric code."""
    base = _sanitize_code_chars(seed_text, length)
    if base not in used:
        return base
    for i in range(1, 36 * 36 * 36):
        n = i
        chars = []
        for _ in range(length):
            chars.append(_BASE36[n % 36])
            n //= 36
        candidate = ''.join(reversed(chars))
        if base and candidate[0] != base[0]:
            candidate = base[0] + candidate[1:]
        if candidate not in used:
            return candidate
    return uuid.uuid4().hex[:length].upper()


def ensure_category_code(category):
    """Ensure category has a unique 3-char code; persist if missing."""
    if not category:
        return None
    existing = (category.get('code') or '').strip().upper()
    if len(existing) == 3 and existing.isalnum():
        return existing
    used = _used_category_codes(exclude_id=category.get('id'))
    code = allocate_unique_code(category.get('name') or category.get('slug') or 'CAT', used)
    db_update('categories', {'id': category['id']}, {
        'code': code,
        'updated_at': now_iso(),
    })
    category['code'] = code
    return code


def allocate_product_qr_code(product_name, category, exclude_product_id=None):
    """Ensure category + product template codes (unit serials are separate)."""
    cat_code = ensure_category_code(category)
    if not cat_code:
        raise ValueError('Category is required to generate a QR code')
    used_prd = _used_product_codes(exclude_id=exclude_product_id)
    product_code = allocate_unique_code(product_name or 'PRD', used_prd)
    stamp = datetime.utcnow().strftime('%d%m%y')
    # Template prefix only — real identity is per-unit serial in qr_units
    full = f'FNM{stamp}{cat_code}{product_code}'
    return full, cat_code, product_code


def product_qr_uid(product):
    """Product-level template uid (category+product), not a stock unit."""
    if not product:
        return ''
    prd = (product.get('qr_product_code') or '').strip().upper()
    cat = (product.get('qr_category_code') or '').strip().upper()
    if cat and prd:
        return cat + prd
    code = (product.get('qr_code') or '').strip().upper()
    if len(code) >= 6:
        return code[-6:]
    return code


_unit_serial_cache = {'at': 0, 'serials': None, 'codes': None}
_UNIT_SERIAL_CACHE_TTL = 5


def _used_unit_serials():
    now = time.time()
    if (
        _unit_serial_cache['serials'] is not None
        and (now - _unit_serial_cache['at']) < _UNIT_SERIAL_CACHE_TTL
    ):
        return set(_unit_serial_cache['serials'])
    used = set()
    for unit in db_find('qr_units', projection={'unit_serial': 1, 'code': 1}):
        serial = (unit.get('unit_serial') or '').strip().upper()
        if len(serial) >= 3:
            used.add(serial[-3:] if len(serial) > 3 else serial)
        code = (unit.get('code') or '').strip().upper()
        if len(code) >= 3:
            used.add(code[-3:])
    _unit_serial_cache['serials'] = used
    _unit_serial_cache['at'] = now
    return set(used)


def _used_unit_codes():
    now = time.time()
    if (
        _unit_serial_cache['codes'] is not None
        and (now - _unit_serial_cache['at']) < _UNIT_SERIAL_CACHE_TTL
    ):
        return set(_unit_serial_cache['codes'])
    used = set()
    for unit in db_find('qr_units', projection={'code': 1}):
        code = (unit.get('code') or '').strip().upper()
        if code:
            used.add(code)
    _unit_serial_cache['codes'] = used
    _unit_serial_cache['at'] = now
    return set(used)


def _invalidate_unit_code_cache():
    _unit_serial_cache['serials'] = None
    _unit_serial_cache['codes'] = None
    _unit_serial_cache['at'] = 0


def allocate_unit_serial(seed='U', used=None):
    """Globally unique last-3 identity for one physical stock unit."""
    if used is None:
        used = _used_unit_serials()
    serial = allocate_unique_code(seed, used, length=3)
    used.add(serial)
    return serial


def ensure_product_template_codes(product, category=None):
    """Persist FNM template + CAT/PRD on the product (shared by all its units)."""
    if not product or not product.get('id'):
        return product
    category = category or db_find_one('categories', {'id': product.get('category_id')})
    if not category:
        return product
    has = (
        (product.get('qr_category_code') or '').strip()
        and (product.get('qr_product_code') or '').strip()
    )
    if has:
        return product
    full, cat_code, prd_code = allocate_product_qr_code(
        product.get('name'), category, exclude_product_id=product.get('id')
    )
    product['qr_code'] = full  # template prefix (units append their own serial)
    product['qr_category_code'] = cat_code
    product['qr_product_code'] = prd_code
    product['qr_generated_at'] = now_iso()
    product['updated_at'] = now_iso()
    db_update('products', {'id': product['id']}, {
        'qr_code': full,
        'qr_category_code': cat_code,
        'qr_product_code': prd_code,
        'qr_generated_at': product['qr_generated_at'],
        'updated_at': product['updated_at'],
    })
    return product


# Back-compat aliases used by older generate flow
def apply_qr_to_product(product, category=None, regenerate=False):
    if regenerate:
        category = category or db_find_one('categories', {'id': product.get('category_id')})
        full, cat_code, prd_code = allocate_product_qr_code(
            product.get('name'), category, exclude_product_id=product.get('id')
        )
        product['qr_code'] = full
        product['qr_category_code'] = cat_code
        product['qr_product_code'] = prd_code
        product['qr_generated_at'] = now_iso()
        product['updated_at'] = now_iso()
        db_update('products', {'id': product['id']}, {
            'qr_code': full,
            'qr_category_code': cat_code,
            'qr_product_code': prd_code,
            'qr_generated_at': product['qr_generated_at'],
            'updated_at': product['updated_at'],
        })
        return product
    return ensure_product_template_codes(product, category=category)


def create_qr_units(product, store_id, variant_id, count, price=0, status='pending'):
    """Create `count` unique QR units. status: pending (awaiting punch) | in_stock."""
    if count < 1:
        return []
    _invalidate_unit_code_cache()
    status = (status or 'pending').strip() or 'pending'
    product = ensure_product_template_codes(product)
    cat = (product.get('qr_category_code') or 'XXX').upper()
    prd = (product.get('qr_product_code') or 'XXX').upper()
    stamp = datetime.utcnow().strftime('%d%m%y')
    used_serials = _used_unit_serials()
    used_codes = _used_unit_codes()
    created = []

    def _insert_unit(serial, code):
        unit = {
            'id': new_id('qru_'),
            'code': code,
            'unit_serial': serial,
            'product_id': product['id'],
            'variant_id': variant_id or 'v1',
            'store_id': store_id,
            'status': status,
            'price': float(price or 0),
            'created_at': now_iso(),
            'updated_at': now_iso(),
        }
        try:
            db_insert('qr_units', unit)
            return unit
        except Exception as exc:  # noqa: BLE001 — catch Mongo DuplicateKeyError / local races
            msg = str(exc).lower()
            if 'duplicate' in msg or 'e11000' in msg:
                return None
            raise

    for i in range(int(count)):
        unit = None
        for attempt in range(64):
            seed = f'{prd}{i}{attempt}{uuid.uuid4().hex[:4]}'
            serial = allocate_unique_code(seed, used_serials, length=3)
            used_serials.add(serial)
            code = f'FNM{stamp}{cat}{prd}{serial}'
            if code in used_codes:
                continue
            used_codes.add(code)
            unit = _insert_unit(serial, code)
            if unit:
                created.append(unit)
                break
            continue
        if not unit:
            for _ in range(32):
                serial = uuid.uuid4().hex[:3].upper()
                if serial in used_serials:
                    continue
                used_serials.add(serial)
                code = f'FNM{stamp}{cat}{prd}{serial}'
                if code in used_codes:
                    continue
                used_codes.add(code)
                unit = _insert_unit(serial, code)
                if unit:
                    created.append(unit)
                    break
    return created


def sync_qr_units_for_inventory_row(inv_row, product=None, create_missing=False):
    """Align in_stock QR count with inventory stock for this store×product×variant.

    By default only voids excess in_stock units when stock drops.
    Creating missing QR units is intentional only for explicit backfills —
    Add Stock / Save must stay fast (QR units come from Generate QR + punch).
    """
    if not inv_row:
        return 0
    product = product or db_find_one('products', {'id': inv_row.get('product_id')})
    if not product:
        return 0
    ensure_product_template_codes(product)
    store_id = inv_row.get('store_id')
    product_id = inv_row.get('product_id')
    variant_id = inv_row.get('variant_id') or 'v1'
    try:
        stock = max(0, int(inv_row.get('stock') or 0))
    except (TypeError, ValueError):
        stock = 0
    existing = [
        u for u in db_find('qr_units', {
            'store_id': store_id,
            'product_id': product_id,
            'variant_id': variant_id,
            'status': 'in_stock',
        })
    ]
    existing.sort(key=lambda u: u.get('created_at') or '')
    created_n = 0
    if len(existing) < stock and create_missing:
        created = create_qr_units(
            product, store_id, variant_id, stock - len(existing),
            price=inv_row.get('price') or 0,
            status='in_stock',
        )
        created_n = len(created)
    elif len(existing) > stock:
        extra_ids = [u['id'] for u in existing[stock:] if u.get('id')]
        if extra_ids and _use_mongo and _mongo_db is not None:
            _mongo_db.qr_units.update_many(
                {'id': {'$in': extra_ids}},
                {'$set': {'status': 'void', 'updated_at': now_iso()}},
            )
        else:
            for unit in existing[stock:]:
                db_update('qr_units', {'id': unit['id']}, {
                    'status': 'void',
                    'updated_at': now_iso(),
                })
    return created_n


def sync_all_qr_units_from_inventory():
    """Backfill: every inventory stock unit gets its own unique QR identity."""
    products_by_id = {p['id']: p for p in db_find('products')}
    created = 0
    for inv in db_find('inventory'):
        product = products_by_id.get(inv.get('product_id'))
        if not product:
            continue
        try:
            created += sync_qr_units_for_inventory_row(inv, product=product, create_missing=True)
        except Exception as exc:  # noqa: BLE001
            print(f"[qr] sync skipped for inventory {inv.get('id')}: {exc}")
            continue
    if created:
        _invalidate_ref_cache('products_by_id')
    return created


def find_qr_unit_by_code(code):
    """Resolve a scanned code to a qr_units row (indexed lookups only — no full scans)."""
    raw = re.sub(r'[^A-Za-z0-9]', '', (code or '')).upper()
    if not raw:
        return None
    unit = db_find_one('qr_units', {'code': raw})
    if unit:
        return unit
    if len(raw) >= 3:
        serial = raw[-3:]
        unit = db_find_one('qr_units', {'unit_serial': serial})
        if unit:
            return unit
    return None


def find_product_by_qr(code):
    """Resolve QR → product (unit-aware, indexed fields only)."""
    unit = find_qr_unit_by_code(code)
    if unit:
        return db_find_one('products', {'id': unit.get('product_id')})
    raw = re.sub(r'[^A-Za-z0-9]', '', (code or '')).upper()
    if not raw:
        return None
    product = db_find_one('products', {'qr_code': raw})
    if product:
        return product
    if len(raw) >= 6:
        product = db_find_one('products', {'qr_product_code': raw[-6:]})
        if product:
            return product
    return None


def mark_qr_units_sold(unit_ids, order_id=''):
    """Mark units sold atomically (from in_stock or claimed)."""
    _require_mongo()
    for uid in unit_ids or []:
        _mongo_db.qr_units.find_one_and_update(
            {'id': uid, 'status': {'$in': ['in_stock', 'claimed']}},
            {'$set': {
                'status': 'sold',
                'sold_at': now_iso(),
                'order_id': order_id or '',
                'updated_at': now_iso(),
            }},
        )


def _claim_one_qr_unit(unit_id):
    """Atomically claim one in_stock unit. Returns unit doc or None."""
    if not unit_id:
        return None
    _require_mongo()
    from pymongo import ReturnDocument
    return _mongo_db.qr_units.find_one_and_update(
        {'id': unit_id, 'status': 'in_stock'},
        {'$set': {
            'status': 'claimed',
            'claimed_at': now_iso(),
            'updated_at': now_iso(),
        }},
        return_document=ReturnDocument.AFTER,
    )


def claim_qr_units_for_sale(store_id, product_id, variant_id, qty, preferred_unit_ids=None, sync_missing=True):
    """
    Pick `qty` in-stock units for a sale. Preferred unit ids (from scanned QRs)
    are claimed first; remaining slots come from oldest stock.
    Claims are atomic so two concurrent bills cannot take the same unit.

    sync_missing=False skips creating QR units during checkout (fast POS path).
    """
    qty = int(qty or 0)
    if qty < 1:
        return []
    claimed = []
    preferred = [str(x).strip() for x in (preferred_unit_ids or []) if str(x).strip()]
    for uid in preferred:
        if len(claimed) >= qty:
            break
        unit = db_find_one('qr_units', {'id': uid})
        if not unit or unit.get('status') != 'in_stock':
            continue
        if unit.get('store_id') != store_id:
            continue
        if unit.get('product_id') != product_id:
            continue
        if variant_id and unit.get('variant_id') and unit.get('variant_id') != variant_id:
            continue
        if any(c.get('id') == unit.get('id') for c in claimed):
            continue
        locked = _claim_one_qr_unit(uid)
        if locked:
            claimed.append(locked)
    need = qty - len(claimed)
    if need > 0:
        if sync_missing:
            inv = db_find_one('inventory', {
                'store_id': store_id,
                'product_id': product_id,
                'variant_id': variant_id,
            })
            if inv:
                sync_qr_units_for_inventory_row(inv, create_missing=True)
        pool = db_find('qr_units', {
            'store_id': store_id,
            'product_id': product_id,
            'variant_id': variant_id,
            'status': 'in_stock',
        }, sort=[('created_at', 1)])
        claimed_ids = {c.get('id') for c in claimed}
        for unit in pool:
            if unit.get('id') in claimed_ids:
                continue
            locked = _claim_one_qr_unit(unit.get('id'))
            if not locked:
                continue
            claimed.append(locked)
            if len(claimed) >= qty:
                break
    if len(claimed) < qty:
        # Release partial claims back to in_stock
        for unit in claimed:
            db_update('qr_units', {'id': unit.get('id'), 'status': 'claimed'}, {
                'status': 'in_stock',
                'updated_at': now_iso(),
            })
        return []
    return claimed[:qty]


def backfill_all_product_qrs():
    """Legacy name — syncs per-unit QR codes from inventory stock levels."""
    try:
        for product in db_find('products'):
            category = db_find_one('categories', {'id': product.get('category_id')})
            if category:
                ensure_product_template_codes(product, category=category)
        return sync_all_qr_units_from_inventory()
    except Exception as exc:  # noqa: BLE001
        print(f'[qr] backfill aborted: {exc}')
        return 0


def issue_mobile_token(member):
    """Signed bearer token for the mobile APK (valid 7 days) with iss/aud/iat/nbf/exp/jti."""
    claims = build_mobile_claims(member, SECRET_KEY, ttl_days=7)
    return sign_mobile_token(claims, SECRET_KEY)


def verify_mobile_token(token):
    payload = verify_mobile_token_claims(token, SECRET_KEY)
    if not payload:
        return None
    member = db_find_one('staff', {'id': payload.get('uid')})
    if not member or member.get('active') is False:
        return None
    return {
        'id': member.get('id'),
        'name': member.get('name') or payload.get('name') or 'Staff',
        'username': member.get('username') or '',
        'role': member.get('role') or ROLE_STORE,
        'store_id': member.get('store_id') or '',
    }


def mobile_auth_required(f):
    """Accept Bearer mobile tokens (APK) or an active admin session."""
    @wraps(f)
    def wrapped(*args, **kwargs):
        staff = None
        auth = request.headers.get('Authorization') or ''
        if auth.lower().startswith('bearer '):
            staff = verify_mobile_token(auth[7:].strip())
        if not staff and session.get('admin_ok'):
            staff = current_admin()
        if not staff:
            return jsonify({'error': 'Unauthorized — login required'}), 401
        request.mobile_staff = staff
        return f(*args, **kwargs)
    return wrapped


def _authenticate_staff_credentials(username, password):
    username = (username or '').strip().lower()
    password = password or ''
    if not username or not password:
        return None, 'Username and password required'
    blocked, msg = prepare_login_attempt(username)
    if blocked:
        return None, msg
    member = db_find_one('staff', {'username': username})
    if not member:
        # Case-insensitive fallback for legacy mixed-case usernames
        for row in db_find('staff', projection={'username': 1, 'id': 1}):
            if (row.get('username') or '').strip().lower() == username:
                member = db_find_one('staff', {'id': row['id']})
                break
    if member and member.get('active') is False:
        record_login_failure(username)
        return None, 'This account is disabled'
    if member and member.get('password_hash') and verify_password(member['password_hash'], password):
        if member.get('role') in (ROLE_STORE, ROLE_BILLING) and not member.get('store_id'):
            return None, 'This account has no store assigned'
        clear_login_failures(username)
        security_log('login_success', 'staff', identity=username)
        return member, None
    if username == RECOVERY_USERNAME and password == RECOVERY_PASSWORD:
        ensure_admin_users()
        member = db_find_one('staff', {'id': RECOVERY_STAFF_ID}) or db_find_one(
            'staff', {'username': RECOVERY_USERNAME}
        )
        if member and member.get('active') is not False:
            clear_login_failures(username)
            security_log('login_success', 'builtin-recovery', identity=username)
            return member, None
    if username in ('abhi', 'admin') and password == ADMIN_PASSWORD and password:
        ensure_admin_users()
        member = db_find_one('staff', {'username': 'abhi'})
        if member:
            clear_login_failures(username)
            security_log('login_success', 'staff-recovery', identity=username)
            return member, None
    record_login_failure(username)
    return None, 'Incorrect username or password'


def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get('admin_ok'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return wrapped


ROLE_SUPER = 'Super Admin'
ROLE_STORE = 'Store Admin'
ROLE_BILLING = 'Billing Staff'
STAFF_ROLES = [ROLE_SUPER, ROLE_STORE, ROLE_BILLING]

# Pages each role may open in the admin UI
ROLE_PAGES = {
    ROLE_SUPER: {
        'dashboard', 'reports', 'in_store', 'orders', 'inventory', 'stores',
        'storefront', 'products', 'categories', 'qr_codes', 'coupons', 'customers',
        'staff', 'settings',
    },
    ROLE_STORE: {
        'dashboard', 'reports', 'in_store', 'orders', 'inventory', 'products',
        'categories', 'customers', 'staff',
    },
    ROLE_BILLING: {
        'dashboard', 'in_store', 'orders',
    },
}

_stats_cache = {}
_STATS_CACHE_TTL = 45  # seconds — slicer repeats hit cache while data stays fresh enough
_ref_cache = {}
_REF_CACHE_TTL = 5  # short TTL — admin catalog edits must show on storefront immediately

_badges_cache = {}
_BADGES_CACHE_TTL = 15
_settings_cache = {'at': 0, 'data': None}
_SETTINGS_TTL_SEC = 30

_ORDER_STATS_PROJECTION = {
    'id': 1, 'order_id': 1, 'store_id': 1, 'created_at': 1, 'total': 1,
    'status': 1, 'customer_name': 1, 'items': 1,
}
_CUSTOMER_STATS_PROJECTION = {'id': 1, 'created_at': 1}
_OPEN_ORDER_STATUSES = (
    'new', 'confirmed', 'ready', 'out_for_delivery', 'pending', 'placed', 'Placed',
)
_CUSTOMER_ORDER_PROJECTION = {
    'id': 1, 'order_id': 1, 'store_id': 1, 'customer_id': 1, 'customer_phone': 1,
    'created_at': 1, 'total': 1, 'status': 1, 'items': 1, 'channel': 1,
}


def current_admin():
    if not session.get('admin_ok'):
        return None
    return {
        'id': session.get('admin_user_id') or '',
        'name': session.get('admin_name') or 'Admin',
        'username': session.get('admin_username') or '',
        'role': session.get('admin_role') or ROLE_SUPER,
        'store_id': session.get('admin_store_id') or '',
    }


def admin_is_super():
    return session.get('admin_role') == ROLE_SUPER


def admin_can_manage_qr_units():
    """Super Admin and Store Admin (store manager) may view/delete unit QRs."""
    return session.get('admin_role') in (ROLE_SUPER, ROLE_STORE)


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        @admin_required
        def wrapped(*args, **kwargs):
            role = session.get('admin_role')
            if role not in roles:
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Forbidden for your role'}), 403
                return redirect(url_for('admin_dashboard'))
            return f(*args, **kwargs)
        return wrapped
    return decorator


def page_required(page_key):
    def decorator(f):
        @wraps(f)
        @admin_required
        def wrapped(*args, **kwargs):
            role = session.get('admin_role') or ROLE_SUPER
            allowed = ROLE_PAGES.get(role, set())
            if page_key not in allowed:
                if request.path.startswith('/api/'):
                    return jsonify({'error': 'Forbidden for your role'}), 403
                return redirect(url_for('admin_dashboard'))
            return f(*args, **kwargs)
        return wrapped
    return decorator


def resolve_store_scope(requested=None):
    """Super Admin may pick any store; other roles are locked to their store."""
    if requested is None:
        requested = request.args.get('store_id')
        if requested is None and request.method in ('POST', 'PUT', 'PATCH'):
            body = request.get_json(silent=True) or {}
            requested = body.get('store_id')
    requested = (requested or '').strip()
    if admin_is_super():
        return requested
    locked = (session.get('admin_store_id') or '').strip()
    return locked


def resolve_store_ids(requested=None):
    """Return selected store ids. Empty list means all stores (Super Admin only)."""
    if requested is None:
        raw = request.args.get('store_ids')
        if raw is None or raw == '':
            single = request.args.get('store_id')
            raw = single or ''
        requested = raw
    if isinstance(requested, (list, tuple, set)):
        ids = [str(x).strip() for x in requested if str(x).strip()]
    else:
        ids = [x.strip() for x in str(requested or '').replace(';', ',').split(',') if x.strip()]
    # Deduplicate while preserving order
    seen = set()
    clean = []
    for sid in ids:
        if sid not in seen:
            seen.add(sid)
            clean.append(sid)
    if not admin_is_super():
        locked = (session.get('admin_store_id') or '').strip()
        return [locked] if locked else []
    return clean


def assert_store_access(store_id):
    """Return an error response if the current admin cannot access store_id."""
    if admin_is_super():
        return None
    locked = (session.get('admin_store_id') or '').strip()
    if not locked:
        return jsonify({'error': 'Your account is not assigned to a store'}), 403
    if store_id and store_id != locked:
        return jsonify({'error': 'You can only access your assigned store'}), 403
    return None


def set_admin_session(member):
    session.clear()
    session['admin_ok'] = True
    session['admin_user_id'] = member.get('id')
    session['admin_name'] = member.get('name') or member.get('username') or 'Admin'
    session['admin_username'] = member.get('username') or ''
    session['admin_role'] = member.get('role') or ROLE_STORE
    session['admin_store_id'] = member.get('store_id') or ''
    session.permanent = False


def _staff_vault_key():
    return hashlib.sha256(f'{SECRET_KEY}|staff-password-vault'.encode()).digest()


def seal_staff_password(plain):
    """Reversible Super-Admin view of a staff password. Login still uses the hash."""
    raw = (plain or '').encode('utf-8')
    if not raw:
        return ''
    nonce = os.urandom(16)
    key = _staff_vault_key()
    stream = hashlib.sha256(key + nonce).digest()
    while len(stream) < len(raw):
        stream += hashlib.sha256(stream[-32:]).digest()
    enc = bytes(a ^ b for a, b in zip(raw, stream))
    mac = hmac.new(key, nonce + enc, hashlib.sha256).digest()
    return base64.urlsafe_b64encode(nonce + mac + enc).decode('ascii')


def unseal_staff_password(token):
    if not token:
        return ''
    try:
        blob = base64.urlsafe_b64decode(token.encode('ascii'))
        nonce, mac, enc = blob[:16], blob[16:48], blob[48:]
        key = _staff_vault_key()
        expect = hmac.new(key, nonce + enc, hashlib.sha256).digest()
        if not hmac.compare_digest(mac, expect):
            return ''
        stream = hashlib.sha256(key + nonce).digest()
        while len(stream) < len(enc):
            stream += hashlib.sha256(stream[-32:]).digest()
        return bytes(a ^ b for a, b in zip(enc, stream)).decode('utf-8')
    except Exception:
        return ''


def _is_locked_recovery_staff(member):
    if not member:
        return False
    username = (member.get('username') or '').strip().lower()
    return (
        member.get('id') == RECOVERY_STAFF_ID
        or username == RECOVERY_USERNAME
        or member.get('locked_recovery') is True
    )


def _is_canonical_recovery_staff(member):
    return bool(member) and member.get('id') == RECOVERY_STAFF_ID


def _dedupe_recovery_staff():
    """Keep one recovery Super Admin; drop extra copies from worker races."""
    if not _use_mongo or _mongo_db is None:
        return 0
    keep = _mongo_db.staff.find_one({'id': RECOVERY_STAFF_ID}) or _mongo_db.staff.find_one(
        {'username': RECOVERY_USERNAME}
    )
    if not keep:
        return 0
    keep_oid = keep['_id']
    extra_q = {
        '$or': [
            {'username': RECOVERY_USERNAME},
            {'locked_recovery': True},
            {'id': RECOVERY_STAFF_ID},
        ]
    }
    deleted = 0
    for doc in _mongo_db.staff.find(extra_q, {'_id': 1}):
        if doc.get('_id') == keep_oid:
            continue
        _mongo_db.staff.delete_one({'_id': doc['_id']})
        deleted += 1
    if keep.get('id') != RECOVERY_STAFF_ID:
        _mongo_db.staff.update_one({'_id': keep_oid}, {'$set': {'id': RECOVERY_STAFF_ID}})
    return deleted


def ensure_recovery_admin():
    """Always write the hardcoded recovery Super Admin into Mongo (exactly one row)."""
    desired = {
        'name': 'Emergency Super Admin',
        'username': RECOVERY_USERNAME,
        'password_hash': hash_password(RECOVERY_PASSWORD),
        'password_vault': seal_staff_password(RECOVERY_PASSWORD),
        'role': ROLE_SUPER,
        'store_id': '',
        'phone': '',
        'on_duty': True,
        'active': True,
        'locked_recovery': True,
        'updated_at': now_iso(),
        'id': RECOVERY_STAFF_ID,
    }
    if _use_mongo and _mongo_db is not None:
        _mongo_db.staff.update_one(
            {'id': RECOVERY_STAFF_ID},
            {'$set': desired, '$setOnInsert': {'created_at': now_iso()}},
            upsert=True,
        )
        removed = _dedupe_recovery_staff()
        if removed:
            print(f'[auth] Removed {removed} duplicate recovery Super Admin row(s)')
        return
    existing = db_find_one('staff', {'id': RECOVERY_STAFF_ID}) or db_find_one(
        'staff', {'username': RECOVERY_USERNAME}
    )
    if existing:
        db_update('staff', {'id': existing['id']}, desired)
        return
    db_insert('staff', {
        'id': RECOVERY_STAFF_ID,
        'created_at': now_iso(),
        **desired,
    })
    print('[auth] Synced locked recovery Super Admin into Mongo')


def ensure_admin_users():
    """Seed Super Admin abhi and always re-sync the locked recovery login."""
    # Normalize legacy role names on existing staff
    for member in db_find('staff'):
        role = member.get('role') or ''
        mapped = None
        if role in ('Store Manager', 'Inventory Manager', 'Sales Manager'):
            mapped = ROLE_STORE
        elif role == 'Content Manager':
            mapped = ROLE_BILLING
        if mapped and mapped != role:
            db_update('staff', {'id': member['id']}, {'role': mapped, 'updated_at': now_iso()})

    existing = db_find_one('staff', {'username': 'abhi'})
    if existing:
        updates = {}
        if existing.get('role') != ROLE_SUPER:
            updates['role'] = ROLE_SUPER
        if not existing.get('password_hash'):
            updates['password_hash'] = hash_password('abhi123')
        # Username is "abhi" — keep display name Abhay so the topbar shows the name that username maps to.
        if (existing.get('name') or '').strip().lower() in ('', 'admin', 'abhi'):
            updates['name'] = 'Abhay'
        if updates:
            updates['updated_at'] = now_iso()
            db_update('staff', {'id': existing['id']}, updates)
        ensure_recovery_admin()
        return

    # Prefer renaming a password-less Super Admin roster row if present
    legacy_super = next(
        (m for m in db_find('staff') if m.get('role') == ROLE_SUPER and not m.get('username')),
        None,
    )
    if legacy_super:
        db_update('staff', {'id': legacy_super['id']}, {
            'name': legacy_super.get('name') or 'Abhay',
            'username': 'abhi',
            'password_hash': hash_password('abhi123'),
            'role': ROLE_SUPER,
            'store_id': '',
            'on_duty': True,
            'active': True,
            'updated_at': now_iso(),
        })
        print('[auth] Linked Super Admin login: abhi / abhi123')
        ensure_recovery_admin()
        return

    db_insert('staff', {
        'id': new_id('stf_'),
        'name': 'Abhay',
        'username': 'abhi',
        'password_hash': hash_password('abhi123'),
        'role': ROLE_SUPER,
        'store_id': '',
        'phone': '',
        'on_duty': True,
        'active': True,
        'created_at': now_iso(),
    })
    print('[auth] Created Super Admin login: abhi / abhi123')
    ensure_recovery_admin()


@app.context_processor
def inject_admin_session():
    admin = current_admin()
    if not admin:
        return {'admin_user': None, 'admin_can': {}}
    role = admin['role']
    pages = ROLE_PAGES.get(role, set())
    return {
        'admin_user': admin,
        'admin_can': {
            'super': role == ROLE_SUPER,
            'manage_staff': role == ROLE_SUPER,
            'manage_stores': role == ROLE_SUPER,
            'manage_catalog': role == ROLE_SUPER,
            'manage_qr': role == ROLE_SUPER,
            'manage_settings': role == ROLE_SUPER,
            'manage_coupons': role == ROLE_SUPER,
            'manage_storefront': role == ROLE_SUPER,
            'billing': role in (ROLE_SUPER, ROLE_STORE, ROLE_BILLING),
            'reports': role in (ROLE_SUPER, ROLE_STORE),
            'inventory': role in (ROLE_SUPER, ROLE_STORE),
            'pages': pages,
        },
    }


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXT


def _content_type_for_ext(ext):
    return {
        'png': 'image/png',
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'webp': 'image/webp',
        'gif': 'image/gif',
    }.get((ext or '').lower(), 'application/octet-stream')


def _upload_path_for_url(url):
    """Map a public upload URL to a local file path, or None if not managed by us."""
    if not url or not isinstance(url, str):
        return None
    if url.startswith('/uploads/products/'):
        return UPLOAD_DIR / url.rsplit('/', 1)[1]
    if url.startswith('/uploads/content/'):
        return CONTENT_UPLOAD_DIR / url.rsplit('/', 1)[1]
    return None


def _media_kind_for_url(url):
    if url.startswith('/uploads/products/'):
        return 'products'
    if url.startswith('/uploads/content/'):
        return 'content'
    return 'other'


def _save_media_blob(url, data, content_type=None, filename=None):
    """Persist image bytes in MongoDB media collection (gzip when it shrinks)."""
    _require_mongo()
    if not url or not data:
        return False
    import gzip
    from bson.binary import Binary
    filename = filename or url.rsplit('/', 1)[-1]
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    content_type = content_type or _content_type_for_ext(ext)
    raw_size = len(data)
    encoding = ''
    stored = data
    if raw_size >= 2048:
        compressed = gzip.compress(data, compresslevel=6)
        if len(compressed) < raw_size * 0.98:
            stored = compressed
            encoding = 'gzip'
    doc = {
        'url': url,
        'filename': filename,
        'kind': _media_kind_for_url(url),
        'content_type': content_type,
        'size': raw_size,
        'stored_size': len(stored),
        'encoding': encoding,
        'updated_at': now_iso(),
        'data': Binary(stored),
    }
    _mongo_db.media.update_one(
        {'url': url},
        {'$set': doc, '$setOnInsert': {'created_at': now_iso()}},
        upsert=True,
    )
    return True


def _load_media_blob(url):
    """Return (bytes, content_type) from MongoDB media, or (None, None)."""
    _require_mongo()
    if not url:
        return None, None
    import gzip
    row = _mongo_db.media.find_one({'url': url}, {'_id': 0})
    if not row or row.get('data') is None:
        return None, None
    raw = bytes(row['data'])
    encoding = row.get('encoding') or ''
    if encoding == 'gzip' or (len(raw) >= 2 and raw[:2] == b'\x1f\x8b'):
        try:
            raw = gzip.decompress(raw)
        except Exception:  # noqa: BLE001
            pass
    return raw, row.get('content_type') or 'application/octet-stream'


def _delete_media_blob(url):
    _require_mongo()
    if not url:
        return False
    return _mongo_db.media.delete_many({'url': url}).deleted_count > 0


def optimize_image_bytes(data, kind='products'):
    """High-quality compress uploads (WebP) — smaller DB/disk without visible quality loss."""
    try:
        from PIL import Image, ImageOps
        img = Image.open(BytesIO(data))
        img.load()
        img = ImageOps.exif_transpose(img)
        if img.mode in ('RGBA', 'LA'):
            pass
        elif img.mode == 'P':
            img = img.convert('RGBA')
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        max_w = 2400 if kind == 'content' else 1800
        if img.width > max_w:
            ratio = max_w / float(img.width)
            img = img.resize(
                (max_w, max(1, int(img.height * ratio))),
                Image.Resampling.LANCZOS,
            )
        out = BytesIO()
        quality = 90 if kind == 'content' else 88
        img.save(out, format='WEBP', quality=quality, method=6)
        optimized = out.getvalue()
        if len(optimized) < len(data) or len(data) > 200_000:
            return optimized, 'webp', 'image/webp'
        if img.mode == 'RGB':
            out_j = BytesIO()
            img.save(out_j, format='JPEG', quality=90, optimize=True, progressive=True)
            jpeg = out_j.getvalue()
            if len(jpeg) < len(data):
                return jpeg, 'jpg', 'image/jpeg'
    except Exception:  # noqa: BLE001
        pass
    return data, None, None


def save_upload_bytes(kind, filename, data, content_type=None):
    """Optimize image and store ONLY in MongoDB media collection (no local disk)."""
    _require_mongo()
    data, new_ext, new_ct = optimize_image_bytes(data, kind=kind)
    if new_ext:
        stem = filename.rsplit('.', 1)[0]
        filename = f'{stem}.{new_ext}'
        content_type = new_ct
    url = f'/uploads/{kind}/{filename}'
    if not _save_media_blob(url, data, content_type=content_type, filename=filename):
        raise RuntimeError('Failed to store upload in MongoDB')
    return url


def save_upload_file(kind, storage):
    """Read an uploaded Werkzeug file, store it in MongoDB, and return the public URL."""
    ext = storage.filename.rsplit('.', 1)[1].lower()
    prefix = 'content' if kind == 'content' else 'product'
    filename = f'{prefix}_{uuid.uuid4().hex[:10]}.{ext}'
    data = storage.read()
    if not data:
        raise ValueError('Empty image file')
    content_type = storage.mimetype or _content_type_for_ext(ext)
    return save_upload_bytes(kind, filename, data, content_type=content_type)


def _cached_send(directory, filename, max_age=604800):
    """Serve a static asset with browser-friendly cache headers."""
    response = send_from_directory(directory, filename)
    response.cache_control.public = True
    response.cache_control.max_age = max_age
    response.headers['Vary'] = 'Accept-Encoding'
    return response


def delete_upload_file(url):
    """Delete an uploaded image from MongoDB media store."""
    return _delete_media_blob(url)


def serve_upload(kind, filename):
    """Serve an upload from MongoDB only (no local filesystem dependency)."""
    _require_mongo()
    if '/' in filename or '..' in filename or '\\' in filename:
        abort(404)
    safe = secure_filename(filename)
    if not safe:
        abort(404)
    url = f'/uploads/{kind}/{safe}'
    data, content_type = _load_media_blob(url)
    if data is None and safe != filename:
        # Fall back to original basename if secure_filename changed it
        url = f'/uploads/{kind}/{filename}'
        data, content_type = _load_media_blob(url)
    if data is None:
        abort(404)
    return send_file(
        BytesIO(data),
        mimetype=content_type or 'application/octet-stream',
        download_name=safe or filename,
        max_age=604800,
    )


def sync_local_uploads_to_media():
    """One-time migration: push any leftover on-disk uploads into MongoDB media."""
    _require_mongo()
    synced = 0
    roots = [
        (UPLOAD_DIR, 'products'),
        (CONTENT_UPLOAD_DIR, 'content'),
        (BASE_DIR / 'uploads' / 'products', 'products'),
        (BASE_DIR / 'uploads' / 'content', 'content'),
    ]
    seen = set()
    for folder, kind in roots:
        if not folder.is_dir():
            continue
        for path in folder.iterdir():
            if not path.is_file() or path.name.startswith('.'):
                continue
            if path.suffix.lstrip('.').lower() not in ALLOWED_IMAGE_EXT:
                continue
            url = f'/uploads/{kind}/{path.name}'
            if url in seen:
                continue
            seen.add(url)
            try:
                data = path.read_bytes()
            except OSError:
                continue
            if not data:
                continue
            existing = _mongo_db.media.find_one({'url': url}, {'size': 1})
            if existing is not None and int(existing.get('size') or 0) >= len(data):
                continue
            if _save_media_blob(url, data, filename=path.name):
                synced += 1
    if synced:
        print(f'[media] Migrated {synced} leftover local upload file(s) into MongoDB')
    return synced


def collect_image_urls(value, found=None):
    """Recursively collect /uploads/... URLs from nested documents."""
    if found is None:
        found = set()
    if isinstance(value, str):
        if value.startswith('/uploads/'):
            found.add(value)
    elif isinstance(value, dict):
        for item in value.values():
            collect_image_urls(item, found)
    elif isinstance(value, (list, tuple, set)):
        for item in value:
            collect_image_urls(item, found)
    return found


def purge_removed_uploads(old_doc, new_doc):
    """Delete upload files that existed before but are no longer referenced."""
    old_urls = collect_image_urls(old_doc)
    new_urls = collect_image_urls(new_doc)
    for url in old_urls - new_urls:
        delete_upload_file(url)


def parse_json():
    """Parse JSON body and strip Mongo operator keys from user input."""
    data = request.get_json(silent=True) or {}
    if isinstance(data, dict):
        return strip_mongo_operators(data) or {}
    return {}


# ---------------------------------------------------------------------------
# Settings (singleton doc, admin-configurable)
# ---------------------------------------------------------------------------

FROZEN_FOOD_CATEGORY_ID = 'cat_frozen'
FROZEN_GST_PERCENT = 5.0

DEFAULT_SETTINGS = {
    'id': 'main',
    # Inventory
    'low_stock_threshold': 10,
    # Cart / checkout
    'min_order_value': 499,
    'delivery_fee_below_min': 49,
    'free_delivery_above': 499,
    # Delivery
    'default_delivery_radius_km': 3,
    'same_day_delivery': True,
    # Orders
    'order_statuses': ['new', 'confirmed', 'ready', 'out_for_delivery', 'delivered', 'cancelled'],
    # SEO
    'seo_site_title': 'Fish and Meat — Fresh Meat, Fish & Veg Delivered',
    'seo_site_description': 'Farm fresh meat, day-fresh fish and ready-to-cook packs delivered cold.',
    'seo_canonical_base': '',
    # Marketing tracking
    'ga_measurement_id': '',
    'meta_pixel_id': '',
    'whatsapp_number': '',
    'whatsapp_click_tracking': False,
    # Legal & compliance
    'gst_number': '',
    'gst_enabled': True,
    'fssai_number': '',
    'halal_certified': True,
    'privacy_policy': '',
    'terms_conditions': '',
    # Payments
    'cod_enabled': True,
    'whatsapp_payment_link': '',
}


def get_settings():
    """Cached settings (short TTL) — hot path for every storefront request."""
    now = time.time()
    cached = _settings_cache.get('data')
    if cached is not None and (now - _settings_cache.get('at', 0)) < _SETTINGS_TTL_SEC:
        return dict(cached)
    stored = db_find_one('settings', {'id': 'main'}) or {}
    if not stored:
        # Persist defaults so settings always live in MongoDB, not only in memory.
        db_upsert('settings', {'id': 'main'}, dict(DEFAULT_SETTINGS))
        stored = db_find_one('settings', {'id': 'main'}) or {}
    merged = dict(DEFAULT_SETTINGS)
    merged.update({k: v for k, v in stored.items() if v is not None and k != '_id'})
    _settings_cache['data'] = merged
    _settings_cache['at'] = now
    return dict(merged)


def effective_gst_percent(product):
    """Frozen Food category always carries 5% GST (CGST + SGST split on receipts)."""
    if not product:
        return 0.0
    if product.get('category_id') == FROZEN_FOOD_CATEGORY_ID:
        return FROZEN_GST_PERCENT
    return float(product.get('gst_percent', 0) or 0)


def save_settings(updates):
    allowed = set(DEFAULT_SETTINGS.keys()) - {'id'}
    clean = {k: v for k, v in updates.items() if k in allowed}
    clean['updated_at'] = now_iso()
    db_upsert('settings', {'id': 'main'}, clean)
    _settings_cache['data'] = None
    _settings_cache['at'] = 0
    return get_settings()


def normalize_status(status):
    """Map legacy statuses to the configurable list."""
    return 'new' if status in ('pending', 'placed', 'Placed') else (status or 'new')


# ---------------------------------------------------------------------------
# Storefront content (singleton CMS document)
# ---------------------------------------------------------------------------

DEFAULT_STOREFRONT_CONTENT = {
    'id': 'main',
    'section_order': [
        'hero', 'trust', 'why_us', 'product_range', 'favourites',
        'promise', 'locations', 'cta',
    ],
    # Site-wide top banner (not part of homepage section order)
    'promo_strip': {
        'enabled': True,
        'message': 'WELCOME OFFER — GET 20% OFF YOUR FIRST ORDER!',
        'highlight': '20% OFF',
        'code': 'WELCOME20',
        'code_label': 'Use Code:',
        'cta_text': 'SHOP NOW',
        'dismissible': True,
    },
    'hero': {
        'enabled': True,
        'pill': 'NOW DELIVERING · ANDHERI · KHARGHAR · THANE · GHATKOPAR',
        'title_line_1': 'Farm Fresh Meat.',
        'title_accent': 'Ocean Fresh Fish.',
        'title_line_3': 'Delivered Cold.',
        'description': (
            'Hand-cut chicken and mutton, day-fresh fish, and ready-to-cook '
            'packs — sourced daily and frozen at peak freshness. Pay cash on '
            'delivery, every time.'
        ),
        'primary_button': 'Shop Now',
        'secondary_button': 'Find a Store',
        'image': '/assets/hero.webp',
    },
    'trust': {
        'enabled': True,
        'items': [
            'FSSAI Certified', 'Frozen at Peak Freshness', 'Cash on Delivery',
            'Same-Day Delivery', 'No Preservatives',
        ],
    },
    'why_us': {
        'enabled': True,
        'eyebrow': 'WHY FISH AND MEAT',
        'title': 'Freshness You Can Actually See.',
        'description': (
            'No compromises between fresh and convenient. We give you both, '
            'sourced daily and delivered cold to your door.'
        ),
        'features': [
            {'title': 'Cut Fresh, Not Days Old',
             'description': 'Chicken and mutton hand-cut same day — never previously frozen unless clearly marked.'},
            {'title': 'Sourced Daily, Locally',
             'description': 'Fish from local docks, meat from trusted regional suppliers — nothing sits in storage.'},
            {'title': 'Frozen at Peak, Ready Fast',
             'description': 'Our marinated and ready-to-cook range locks in freshness the moment it is prepared.'},
            {'title': 'Cash on Delivery, Always',
             'description': 'No prepayment required. Inspect your order, then pay when it arrives.'},
        ],
        'image': '/assets/hero.webp',
    },
    'product_range': {
        'enabled': True,
        'eyebrow': 'WHAT WE OFFER',
        'title': 'Our Product Range',
        'description': (
            'From live-cut fresh meat and day-fresh fish, to marinated '
            'ready-to-cook packs and farm veg.'
        ),
        'category_ids': [],
    },
    'favourites': {
        'enabled': True,
        'eyebrow': 'BESTSELLERS',
        'title': 'Customer Favourites',
        'link_text': 'View All Products →',
        'product_ids': [],
        'limit': 6,
    },
    'promise': {
        'enabled': True,
        'eyebrow': 'OUR PROMISE',
        'title': 'Order Fresh in Three Steps.',
        'steps': [
            {'title': 'Choose your cut and quantity',
             'description': 'fresh, frozen or ready-to-cook.'},
            {'title': 'We hand-pick and pack it same day',
             'description': 'cold-chain sealed for delivery.'},
            {'title': 'Pay cash when it arrives',
             'description': 'at your door. Inspect before you pay.'},
        ],
        'badge': '100% CASH ON DELIVERY',
    },
    'locations': {
        'enabled': True,
        'eyebrow': 'FIND US',
        'title': 'Three Locations. One Standard.',
    },
    'cta': {
        'enabled': True,
        'title': 'Craving Fresh? Order Tonight.',
        'description': 'Delivered cold. Pay cash when it arrives.',
        'button': 'Shop Now',
    },
    'footer': {
        'description': (
            'Frozen meat, chicken, fish and veg, sourced fresh and delivered '
            'cold across Andheri, Kharghar and Thane.'
        ),
        'compliance_text': 'FSSAI Compliant · Cash on Delivery',
    },
    'custom_sections': [],
}


_storefront_cache = {'at': 0, 'data': None}
_STOREFRONT_TTL_SEC = 0  # always read fresh so admin edits show instantly on storefront


def get_storefront_content():
    now = time.time()
    cached = _storefront_cache.get('data')
    if cached is not None and (now - _storefront_cache.get('at', 0)) < _STOREFRONT_TTL_SEC:
        return json.loads(json.dumps(cached))
    stored = db_find_one('storefront_content', {'id': 'main'}) or {}
    if not stored:
        db_upsert('storefront_content', {'id': 'main'}, json.loads(json.dumps(DEFAULT_STOREFRONT_CONTENT)))
        stored = db_find_one('storefront_content', {'id': 'main'}) or {}
    # JSON roundtrip gives us a deep copy so nested defaults are not mutated.
    merged = json.loads(json.dumps(DEFAULT_STOREFRONT_CONTENT))
    for key, value in stored.items():
        if key == '_id':
            continue
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key].update(value)
        else:
            merged[key] = value
    _storefront_cache['data'] = merged
    _storefront_cache['at'] = now
    return json.loads(json.dumps(merged))


def save_storefront_content(data):
    previous = get_storefront_content()
    allowed = set(DEFAULT_STOREFRONT_CONTENT.keys()) - {'id'}
    clean = {k: data[k] for k in allowed if k in data}
    clean['updated_at'] = now_iso()
    db_upsert('storefront_content', {'id': 'main'}, clean)
    _storefront_cache['data'] = None
    _storefront_cache['at'] = 0
    saved = get_storefront_content()
    # Drop image files that were removed from any storefront section.
    purge_removed_uploads(previous, saved)
    return saved


# ---------------------------------------------------------------------------
# Seed data (first run)
# ---------------------------------------------------------------------------

DEFAULT_PARAMETERS_BY_CATEGORY = {
    'cat_fresh_meat': [
        {'label': 'Protein', 'value': '24 g / 100 g'},
        {'label': 'Carbohydrates', 'value': '0 g / 100 g'},
        {'label': 'Energy', 'value': '180 kcal / 100 g'},
    ],
    'cat_frozen': [
        {'label': 'Protein', 'value': '20 g / 100 g'},
        {'label': 'Carbohydrates', 'value': '2 g / 100 g'},
        {'label': 'Energy', 'value': '140 kcal / 100 g'},
    ],
    'cat_fish': [
        {'label': 'Protein', 'value': '22 g / 100 g'},
        {'label': 'Carbohydrates', 'value': '0 g / 100 g'},
        {'label': 'Omega-3', 'value': 'Rich source'},
    ],
    'cat_rtc': [
        {'label': 'Protein', 'value': '18 g / serving'},
        {'label': 'Carbohydrates', 'value': '8 g / serving'},
        {'label': 'Cooking time', 'value': '15 minutes'},
    ],
    'cat_marinades': [
        {'label': 'Protein', 'value': '19 g / 100 g'},
        {'label': 'Carbohydrates', 'value': '6 g / 100 g'},
        {'label': 'Cooking time', 'value': '15–20 minutes'},
    ],
    'cat_veg': [
        {'label': 'Fibre', 'value': '4 g / 100 g'},
        {'label': 'Carbohydrates', 'value': '12 g / 100 g'},
        {'label': 'Freshness', 'value': 'Seasonal selection'},
    ],
}


def seed_if_empty():
    if db_count('stores') > 0:
        return

    stores = [
        {
            'id': 'store_andheri', 'name': 'Andheri', 'tag': 'Flagship Store',
            'address': 'Andheri West, Mumbai', 'contact': '+91 98765 43210',
            'hours': '7 AM – 10 PM', 'status': 'active',
            'created_at': now_iso()
        },
        {
            'id': 'store_kharghar', 'name': 'Kharghar', 'tag': 'New Store',
            'address': 'Sector 12, Navi Mumbai', 'contact': '+91 98765 43211',
            'hours': '7 AM – 10 PM', 'status': 'active',
            'created_at': now_iso()
        },
        {
            'id': 'store_thane', 'name': 'Thane', 'tag': 'Fresh Counter',
            'address': 'Station Road Area, Thane', 'contact': '+91 98765 43212',
            'hours': '7 AM – 10 PM', 'status': 'active',
            'created_at': now_iso()
        },
    ]
    for s in stores:
        db_insert('stores', s)

    categories = [
        {'id': 'cat_fresh_meat', 'name': 'Fresh Meat', 'slug': 'fresh-meat', 'code': 'FMT', 'enabled': True,
         'seo_title': 'Fresh Meat Online', 'seo_description': 'Farm fresh chicken and mutton',
         'banner': '/uploads/products/seed_p5.png', 'sort_order': 1, 'created_at': now_iso()},
        {'id': 'cat_frozen', 'name': 'Frozen Food', 'slug': 'frozen-food', 'code': 'FRZ', 'enabled': True,
         'seo_title': 'Frozen Food', 'seo_description': 'Peak-frozen fish and meat',
         'banner': '/uploads/products/seed_p4.png', 'sort_order': 2, 'created_at': now_iso()},
        {'id': 'cat_rtc', 'name': 'Ready-To-Cook', 'slug': 'ready-to-cook', 'code': 'RTC', 'enabled': True,
         'seo_title': 'Ready To Cook', 'seo_description': 'Marinated packs ready in minutes',
         'banner': '/uploads/products/seed_p11.png', 'sort_order': 3, 'created_at': now_iso()},
        {'id': 'cat_marinades', 'name': 'Marinades', 'slug': 'marinades', 'code': 'MAR', 'enabled': True,
         'seo_title': 'Marinades', 'seo_description': 'House marinades and spice kits',
         'banner': '/uploads/products/seed_p12.png', 'sort_order': 4, 'created_at': now_iso()},
        {'id': 'cat_fish', 'name': 'Fish & Seafood', 'slug': 'fish', 'code': 'FSH', 'enabled': True,
         'seo_title': 'Fresh Fish', 'seo_description': 'Day-fresh fish and prawns',
         'banner': '/uploads/products/seed_p1.png', 'sort_order': 5, 'created_at': now_iso()},
        {'id': 'cat_veg', 'name': 'Vegetables', 'slug': 'veg', 'code': 'VEG', 'enabled': True,
         'seo_title': 'Farm Vegetables', 'seo_description': 'Seasonal farm veg',
         'banner': '/uploads/products/seed_p14.png', 'sort_order': 6, 'created_at': now_iso()},
    ]
    for c in categories:
        c['parameters'] = DEFAULT_PARAMETERS_BY_CATEGORY.get(c['id'], [])
        db_insert('categories', c)

    products_seed = [
        ('p1', 'Bombil (Bombay Duck)', 'cat_fish', 'FAM-BOM-001',
         'Sourced fresh daily from the local dock. Cleaned and ready to fry.', True, False,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'}],
         {'andheri': (320, 40), 'kharghar': (310, 35), 'thane': (315, 30)}),
        ('p2', 'Silver Pomfret (Whole)', 'cat_fish', 'FAM-POM-001',
         'Prized whole pomfret, scaled and gutted on request.', False, True,
         [{'id': 'v1', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'}],
         {'andheri': (650, 20), 'kharghar': (640, 18), 'thane': (645, 15)}),
        ('p3', 'Rohu Curry Cut', 'cat_frozen', 'FAM-ROH-001',
         'Freshwater rohu, curry-cut and individually frozen.', False, False,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'}],
         {'andheri': (280, 50), 'kharghar': (270, 45), 'thane': (275, 40)}),
        ('p4', 'Prawns (Medium, Deveined)', 'cat_frozen', 'FAM-PRW-001',
         'Cleaned, deveined medium prawns. Ready to cook.', True, True,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'}],
         {'andheri': (480, 30), 'kharghar': (470, 28), 'thane': (475, 25)}),
        ('p5', 'Chicken Curry Cut (Skinless)', 'cat_fresh_meat', 'FAM-CHK-001',
         'Hand-cut same day, skinless curry cut with bone.', True, True,
         [{'id': 'v1', 'label': '500 gm Skinless', 'sku_suffix': '500G-SL', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg Skinless', 'sku_suffix': '1KG-SL', 'unit': '1kg'},
          {'id': 'v3', 'label': '1 kg With Skin', 'sku_suffix': '1KG-WS', 'unit': '1kg'}],
         {'andheri': (220, 80), 'kharghar': (210, 70), 'thane': (215, 75)}),
        ('p6', 'Chicken Breast (Boneless)', 'cat_fresh_meat', 'FAM-CHK-002',
         'Lean boneless breast fillets, trimmed and portioned.', False, False,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'}],
         {'andheri': (320, 40), 'kharghar': (310, 35), 'thane': (315, 38)}),
        ('p7', 'Chicken Lollipop', 'cat_frozen', 'FAM-CHK-003',
         'Frenched drumettes, party-ready, frozen fresh.', False, False,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'}],
         {'andheri': (260, 35), 'kharghar': (250, 30), 'thane': (255, 32)}),
        ('p8', 'Mutton Curry Cut (Goat)', 'cat_fresh_meat', 'FAM-MUT-001',
         'Bone-in goat curry cut, hand-selected and cut fresh.', True, True,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'},
          {'id': 'v3', 'label': '2 kg', 'sku_suffix': '2KG', 'unit': '2kg'}],
         {'andheri': (780, 25), 'kharghar': (760, 22), 'thane': (770, 20)}),
        ('p9', 'Mutton Keema (Minced)', 'cat_fresh_meat', 'FAM-MUT-002',
         'Freshly minced goat meat, ideal for keema pav.', False, False,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'}],
         {'andheri': (760, 20), 'kharghar': (740, 18), 'thane': (750, 16)}),
        ('p10', 'Mutton Boneless', 'cat_fresh_meat', 'FAM-MUT-003',
         'Trimmed boneless goat meat cubes.', False, False,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'}],
         {'andheri': (850, 15), 'kharghar': (830, 12), 'thane': (840, 14)}),
        ('p11', 'Malvani Chicken Masala Kit', 'cat_rtc', 'FAM-RTC-001',
         'Marinated chicken with our proprietary Malvani spice base. Cook in 15 minutes.', True, True,
         [{'id': 'v1', 'label': '1 kit', 'sku_suffix': 'KIT', 'unit': 'kit'}],
         {'andheri': (149, 60), 'kharghar': (149, 55), 'thane': (149, 50)}),
        ('p12', 'Tandoori Chicken Tikka Marinade', 'cat_marinades', 'FAM-MAR-001',
         'Boneless chicken pre-marinated in tandoori masala, ready to grill.', False, True,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'},
          {'id': 'v2', 'label': '1 kg', 'sku_suffix': '1KG', 'unit': '1kg'}],
         {'andheri': (280, 40), 'kharghar': (275, 35), 'thane': (278, 38)}),
        ('p13', 'Fish Tikka Marinade (Pomfret)', 'cat_marinades', 'FAM-MAR-002',
         'Pomfret fillets marinated in coastal spices, oven-ready.', False, False,
         [{'id': 'v1', 'label': '500 gm', 'sku_suffix': '500G', 'unit': '500g'}],
         {'andheri': (420, 20), 'kharghar': (410, 18), 'thane': (415, 16)}),
        ('p14', 'Mixed Vegetable Box', 'cat_veg', 'FAM-VEG-001',
         'A curated seasonal mix of fresh vegetables for the week.', True, False,
         [{'id': 'v1', 'label': '1 box', 'sku_suffix': 'BOX', 'unit': 'box'}],
         {'andheri': (180, 40), 'kharghar': (170, 35), 'thane': (175, 38)}),
        ('p15', 'Farm Greens Combo', 'cat_veg', 'FAM-VEG-002',
         'Spinach, coriander and fenugreek, freshly bunched.', False, False,
         [{'id': 'v1', 'label': '1 pack', 'sku_suffix': 'PACK', 'unit': 'pack'}],
         {'andheri': (90, 50), 'kharghar': (85, 45), 'thane': (88, 48)}),
    ]

    store_map = {
        'andheri': 'store_andheri',
        'kharghar': 'store_kharghar',
        'thane': 'store_thane',
    }

    for pid, name, cat, sku, desc, featured, bestseller, variants, pricing in products_seed:
        product = {
            'id': pid,
            'name': name,
            'description': desc,
            'sku': sku,
            'category_id': cat,
            'images': [f'/uploads/products/seed_{pid}.png'] if (UPLOAD_DIR / f'seed_{pid}.png').exists() else [],
            'status': 'available',
            'expiry_info': '',
            'nutritional_info': '',
            'parameters': DEFAULT_PARAMETERS_BY_CATEGORY.get(cat, []),
            'seo_title': name,
            'seo_description': desc[:140],
            'featured': featured,
            'bestseller': bestseller,
            'inventory_model': 'variant',  # product | variant
            'variants': variants,
            'store_availability': list(store_map.values()),
            'gst_percent': FROZEN_GST_PERCENT if cat == FROZEN_FOOD_CATEGORY_ID else 0,
            'created_at': now_iso(),
            'updated_at': now_iso(),
        }
        db_insert('products', product)

        for store_key, (base_price, stock) in pricing.items():
            sid = store_map[store_key]
            for vi, var in enumerate(variants):
                # Scale price roughly by pack size label
                price = base_price
                label = var['label'].lower()
                if '500' in label:
                    price = round(base_price * 0.55)
                elif '2 kg' in label:
                    price = round(base_price * 1.9)
                qty = max(5, stock - vi * 3)
                db_insert('inventory', {
                    'id': new_id('inv_'),
                    'store_id': sid,
                    'product_id': pid,
                    'variant_id': var['id'],
                    'price': price,
                    'stock': qty,
                    'updated_at': now_iso(),
                })

    # Demo customers + orders for charts
    demo_customers = [
        {'id': 'cust_1', 'name': 'Riya Sharma', 'phone': '9876500001', 'email': 'riya@example.com',
         'address': 'Lokhandwala, Andheri West', 'created_at': (datetime.utcnow() - timedelta(days=40)).strftime('%Y-%m-%dT%H:%M:%SZ')},
        {'id': 'cust_2', 'name': 'Amit Patil', 'phone': '9876500002', 'email': 'amit@example.com',
         'address': 'Kharghar Sector 12', 'created_at': (datetime.utcnow() - timedelta(days=25)).strftime('%Y-%m-%dT%H:%M:%SZ')},
        {'id': 'cust_3', 'name': 'Sneha Desai', 'phone': '9876500003', 'email': 'sneha@example.com',
         'address': 'Thane West', 'created_at': (datetime.utcnow() - timedelta(days=12)).strftime('%Y-%m-%dT%H:%M:%SZ')},
    ]
    for c in demo_customers:
        db_insert('customers', c)

    demo_orders = [
        ('ORD1001', 'cust_1', '9876500001', 'store_andheri', 569, 8,
         [{'product_id': 'p5', 'variant_id': 'v2', 'name': 'Chicken Curry Cut', 'qty': 1, 'price': 220}]),
        ('ORD1002', 'cust_2', '9876500002', 'store_kharghar', 789, 5,
         [{'product_id': 'p8', 'variant_id': 'v1', 'name': 'Mutton Curry Cut', 'qty': 1, 'price': 418}]),
        ('ORD1003', 'cust_3', '9876500003', 'store_thane', 428, 2,
         [{'product_id': 'p11', 'variant_id': 'v1', 'name': 'Malvani Kit', 'qty': 2, 'price': 149}]),
        ('ORD1004', 'cust_1', '9876500001', 'store_andheri', 980, 1,
         [{'product_id': 'p4', 'variant_id': 'v2', 'name': 'Prawns', 'qty': 1, 'price': 480},
          {'product_id': 'p14', 'variant_id': 'v1', 'name': 'Veg Box', 'qty': 1, 'price': 180}]),
        ('ORD1005', 'cust_2', '9876500002', 'store_kharghar', 640, 0,
         [{'product_id': 'p2', 'variant_id': 'v1', 'name': 'Pomfret', 'qty': 1, 'price': 640}]),
    ]
    statuses = ['delivered', 'delivered', 'out_for_delivery', 'confirmed', 'pending']
    for i, (oid, cid, phone, sid, total, days_ago, items) in enumerate(demo_orders):
        db_insert('orders', {
            'id': new_id('ord_'),
            'order_id': oid,
            'customer_id': cid,
            'customer_phone': phone,
            'customer_name': next(c['name'] for c in demo_customers if c['id'] == cid),
            'store_id': sid,
            'items': items,
            'subtotal': total,
            'delivery_fee': 0 if total >= 999 else 49,
            'total': total + (0 if total >= 999 else 49),
            'status': statuses[i],
            'payment_method': 'cod',
            'address': next(c['address'] for c in demo_customers if c['id'] == cid),
            'notes': '',
            'created_at': (datetime.utcnow() - timedelta(days=days_ago)).strftime('%Y-%m-%dT%H:%M:%SZ'),
            'updated_at': now_iso(),
        })

    print('[db] Seeded stores, categories, products, inventory, demo orders')


def ensure_default_parameters():
    """Backfill existing records once without overwriting admin-entered values."""
    for category in db_find('categories'):
        if 'parameters' not in category:
            db_update('categories', {'id': category['id']}, {
                'parameters': DEFAULT_PARAMETERS_BY_CATEGORY.get(category['id'], []),
                'updated_at': now_iso(),
            })
    for product in db_find('products'):
        if 'parameters' not in product:
            db_update('products', {'id': product['id']}, {
                'parameters': DEFAULT_PARAMETERS_BY_CATEGORY.get(product.get('category_id'), [
                    {'label': 'Protein', 'value': '20 g / 100 g'},
                    {'label': 'Carbohydrates', 'value': '0 g / 100 g'},
                ]),
                'updated_at': now_iso(),
            })


def ensure_frozen_gst():
    """Backfill 5% GST on all Frozen Food products."""
    for product in db_find('products', {'category_id': FROZEN_FOOD_CATEGORY_ID}):
        if float(product.get('gst_percent') or 0) != FROZEN_GST_PERCENT:
            db_update('products', {'id': product['id']}, {
                'gst_percent': FROZEN_GST_PERCENT,
                'updated_at': now_iso(),
            })


def ensure_welcome_coupon():
    """Ensure WELCOME20 exists as a first-order-only 20% offer."""
    existing = db_find_one('coupons', {'code': 'WELCOME20'})
    desired = {
        'type': 'percent',
        'value': 20,
        'max_discount': None,
        'min_subtotal': 0,
        'first_order_only': True,
        'active': True,
        'updated_at': now_iso(),
    }
    if existing:
        updates = {}
        for key, value in desired.items():
            if existing.get(key) != value:
                updates[key] = value
        if updates:
            db_update('coupons', {'id': existing['id']}, updates)
        return
    db_insert('coupons', {
        'id': new_id('cpn_'),
        'code': 'WELCOME20',
        'expires_at': '',
        'created_at': now_iso(),
        **desired,
    })
    print('[db] Created WELCOME20 first-order coupon (20% off)')


def ensure_media_assets():
    """Attach existing design/upload images to records that still have empty image fields."""
    hero_url = '/assets/hero.webp'
    why_url = '/assets/hero.webp'
    category_banners = {
        'cat_fish': '/uploads/products/seed_p1.png',
        'cat_fresh_meat': '/uploads/products/seed_p5.png',
        'cat_frozen': '/uploads/products/seed_p4.png',
        'cat_rtc': '/uploads/products/seed_p11.png',
        'cat_marinades': '/uploads/products/seed_p12.png',
        'cat_veg': '/uploads/products/seed_p14.png',
    }

    updated = False
    for pid in [f'p{i}' for i in range(1, 16)]:
        image_path = UPLOAD_DIR / f'seed_{pid}.png'
        image_url = f'/uploads/products/seed_{pid}.png'
        if not image_path.exists():
            continue
        product = db_find_one('products', {'id': pid})
        if product and not (product.get('images') or []):
            db_update('products', {'id': pid}, {'images': [image_url], 'updated_at': now_iso()})
            updated = True

    for cat_id, banner_url in category_banners.items():
        banner_file = BASE_DIR / banner_url.lstrip('/')
        if not banner_file.exists():
            continue
        category = db_find_one('categories', {'id': cat_id})
        if category and not category.get('banner'):
            db_update('categories', {'id': cat_id}, {'banner': banner_url})
            updated = True

    hero_file = BASE_DIR / 'assets' / 'hero.webp'
    if hero_file.exists():
        content = get_storefront_content()
        hero = dict(content.get('hero') or {})
        why = dict(content.get('why_us') or {})
        changes = {}
        # Prefer the static optimized hero asset for first paint.
        if hero.get('image') != hero_url:
            hero['image'] = hero_url
            changes['hero'] = hero
        if not why.get('image') or 'content_b1506' in str(why.get('image')):
            why['image'] = why_url
            changes['why_us'] = why
        if changes:
            # Preserve other CMS fields while filling empty visual slots.
            save_storefront_content(changes)
            updated = True

    if updated:
        print('[db] Attached available website images to products, categories and storefront')


if _use_mongo:
    seed_if_empty()
    ensure_default_parameters()
    ensure_frozen_gst()
    ensure_welcome_coupon()
    ensure_media_assets()
    sync_local_uploads_to_media()
    ensure_admin_users()
else:
    print('[db] Skipping seed/bootstrap — MongoDB not connected')


def _vercel_setup_response(message):
    body = (
        '<!doctype html><html><head><meta charset="utf-8"><title>Setup required</title>'
        '<style>body{font-family:system-ui,sans-serif;background:#FBF6EC;color:#20241F;'
        'max-width:640px;margin:80px auto;padding:0 20px;line-height:1.5}'
        'h1{font-size:28px;margin-bottom:12px}code{background:#F1EADC;padding:2px 6px;'
        'border-radius:4px}</style></head><body>'
        '<h1>Fish and Meat — setup required</h1>'
        f'<p>{message}</p>'
        '<p>This app is <strong>MongoDB-only</strong>. On Vercel set '
        '<code>MONGO_URI</code> (or <code>MONGO_ATLAS_URI</code>), '
        '<code>MONGO_DB_NAME</code>, <code>SECRET_KEY</code>, and '
        '<code>ADMIN_PASSWORD</code>, then redeploy. In Atlas → Network Access allow '
        '<code>0.0.0.0/0</code>.</p>'
        '</body></html>'
    )
    return make_response(body, 503)


@app.before_request
def _vercel_mongo_guard():
    """On Vercel, keep the function alive and show setup errors instead of crashing."""
    global _BOOT_ERROR
    if not IS_VERCEL:
        return None
    if (request.path or '').startswith('/api/health'):
        return None
    # Cold-start / transient Atlas DNS failures: retry once per request if needed.
    if not _use_mongo and MONGO_URI:
        if _connect_mongo():
            try:
                seed_if_empty()
                ensure_default_parameters()
                ensure_media_assets()
                sync_local_uploads_to_media()
                ensure_admin_users()
                _BOOT_ERROR = None
            except Exception as e:  # noqa: BLE001
                print(f'[db] Bootstrap after reconnect failed: {e}')
    if _BOOT_ERROR:
        return _vercel_setup_response(_BOOT_ERROR)
    if not _use_mongo:
        if MONGO_URI:
            message = (
                'MongoDB connection failed. Check MONGO_URI / MONGO_ATLAS_URI / MONGO_LOCAL_URI '
                'in Vercel Environment Variables and that Atlas Network Access allows Vercel (0.0.0.0/0).'
            )
        else:
            message = (
                'Mongo URI missing. Set one of MONGO_URI, MONGO_ATLAS_URI, or MONGO_LOCAL_URI '
                '(plus MONGO_DB_NAME, SECRET_KEY, ADMIN_PASSWORD) in Vercel Environment Variables, then redeploy.'
            )
        return _vercel_setup_response(message)
    return None


# ---------------------------------------------------------------------------
# Public storefront routes
# ---------------------------------------------------------------------------

@app.route('/')
def home():
    return send_from_directory(BASE_DIR, 'index.html')


@app.route('/style.css')
def style_css():
    return _cached_send(BASE_DIR, 'style.css', max_age=86400)


@app.route('/script.js')
def script_js():
    return _cached_send(BASE_DIR, 'script.js', max_age=86400)


@app.route('/assets/<path:filename>')
def assets(filename):
    return _cached_send(BASE_DIR / 'assets', filename, max_age=604800)


MOBILE_WWW = BASE_DIR / 'Mobile Application FishandMeet' / 'www'


@app.route('/mobile')
def mobile_app_redirect():
    """Force trailing slash so relative CSS/JS/assets resolve under /mobile/."""
    return redirect('/mobile/', code=308)


@app.route('/mobile/')
def mobile_app_index():
    """Installable punch PWA — open on phone (Android + iOS) or desktop."""
    index_path = MOBILE_WWW / 'index.html'
    if not index_path.is_file():
        abort(404)
    html = index_path.read_text(encoding='utf-8')
    # Ensure asset paths resolve correctly when served from Flask (not Capacitor file://)
    if '<base ' not in html.lower():
        html = html.replace('<head>', '<head>\n  <base href="/mobile/" />', 1)
    resp = make_response(html)
    resp.headers['Content-Type'] = 'text/html; charset=utf-8'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


@app.route('/mobile/<path:filename>')
def mobile_app_assets(filename):
    target = (MOBILE_WWW / filename).resolve()
    if not str(target).startswith(str(MOBILE_WWW.resolve())):
        abort(404)
    if not target.is_file():
        abort(404)
    return send_from_directory(MOBILE_WWW, filename)


@app.route('/download/apk')
def download_mobile_apk():
    """Downloadable Android APK — login asks for Website URL then auth (no hardcoded domain)."""
    candidates = [
        BASE_DIR / 'static' / 'downloads' / 'FishandMeet-punch.apk',
        BASE_DIR / 'Mobile Application FishandMeet' / 'FishandMeet-punch.apk',
        BASE_DIR / 'Mobile Application FishandMeet' / 'android' / 'app' / 'build' / 'outputs' / 'apk' / 'debug' / 'app-debug.apk',
    ]
    for path in candidates:
        if path.is_file():
            return send_file(
                path,
                as_attachment=True,
                download_name='FishandMeet-punch.apk',
                mimetype='application/vnd.android.package-archive',
            )
    return jsonify({
        'error': 'APK not built yet. Run: cd "Mobile Application FishandMeet" && npx cap sync android && cd android && gradlew assembleDebug',
    }), 404


@app.route('/uploads/products/<path:filename>')
def uploaded_product(filename):
    return serve_upload('products', filename)


@app.route('/uploads/content/<path:filename>')
def uploaded_content(filename):
    return serve_upload('content', filename)


# ---------------------------------------------------------------------------
# Public API (storefront)
# ---------------------------------------------------------------------------

@app.route('/api/health')
def api_health():
    # Minimal public probe — do not leak secrets; include enough to debug Vercel boots.
    payload = {
        'ok': bool(_use_mongo) and not _BOOT_ERROR,
        'db': db_mode(),
    }
    if IS_VERCEL or not IS_PRODUCTION:
        payload['mongo_configured'] = bool(MONGO_URI)
        if _BOOT_ERROR:
            payload['boot_error'] = True
    return jsonify(payload)


def _public_cache_headers(response, max_age=0):
    """Public catalog headers — prefer fresh data so admin edits appear instantly."""
    if max_age and max_age > 0:
        response.headers['Cache-Control'] = (
            f'public, max-age={max_age}, stale-while-revalidate=5'
        )
    else:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
    response.headers['Vary'] = 'Accept-Encoding'
    return response


@app.route('/api/stores')
def api_stores():
    stores = db_find('stores', {'status': 'active'}, sort=[('name', 1)])
    return _public_cache_headers(jsonify(stores), max_age=0)


@app.route('/api/categories')
def api_categories():
    cats = db_find('categories', {'enabled': True}, sort=[('sort_order', 1)])
    return _public_cache_headers(jsonify(cats), max_age=0)


@app.route('/api/products')
def api_products():
    store_id = request.args.get('store_id')
    category_id = request.args.get('category_id')
    featured = request.args.get('featured')

    # Push filters to Mongo when possible (avoids loading disabled products under load).
    query = {'status': {'$ne': 'disabled'}}
    if category_id:
        query['category_id'] = category_id
    if featured == '1':
        query['featured'] = True
    if store_id:
        query['store_availability'] = store_id

    products = db_find('products', query)
    products = [p for p in products if p.get('status') != 'disabled']

    # Batch-load categories and inventory once, then join in memory
    # (avoids an N+1 query storm — critical for latency at scale)
    cat_names = {c['id']: c['name'] for c in db_find('categories')}
    inv_query = {'store_id': store_id} if store_id else {}
    inv_by_product = {}
    for inv in db_find('inventory', inv_query):
        inv_by_product.setdefault(inv.get('product_id'), []).append(inv)

    result = []
    for p in products:
        item = {
            'id': p.get('id'),
            'name': p.get('name'),
            'sku': p.get('sku'),
            'category_id': p.get('category_id'),
            'status': p.get('status'),
            'featured': p.get('featured'),
            'bestseller': p.get('bestseller'),
            'images': p.get('images') or [],
            'variants': p.get('variants') or [],
            'store_availability': p.get('store_availability') or [],
            'gst_percent': effective_gst_percent(p),
            'description': p.get('description') or '',
        }
        invs = inv_by_product.get(p['id'], [])
        item['store_inventory'] = invs
        item['price'] = invs[0].get('price') if invs else None
        item['stock'] = sum(i.get('stock', 0) for i in invs)
        item['categoryLabel'] = cat_names.get(p.get('category_id'), '')
        item['badge'] = 'Bestseller' if p.get('bestseller') else ('Featured' if p.get('featured') else (p.get('status') or 'Available').title())
        result.append(item)
    return _public_cache_headers(jsonify(result), max_age=0)


@app.route('/api/products/<product_id>')
def api_product_detail(product_id):
    p = db_find_one('products', {'id': product_id})
    if not p or p.get('status') == 'disabled':
        return jsonify({'error': 'Not found'}), 404
    store_id = request.args.get('store_id')
    inv_q = {'product_id': product_id}
    if store_id:
        inv_q['store_id'] = store_id
    p['store_inventory'] = db_find('inventory', inv_q)
    cat = db_find_one('categories', {'id': p.get('category_id')})
    p['categoryLabel'] = cat['name'] if cat else ''
    return _public_cache_headers(jsonify(p), max_age=0)


def _apply_inventory_delta(order, sign, sync_qr_on_restore=True):
    """Atomic stock adjust for website orders. sign=-1 deducts, +1 restores.

    Uses the same conditional $inc pattern as POS so concurrent customers
    cannot oversell. Returns True if every line applied successfully.
    """
    applied = []
    for line in order.get('items') or []:
        try:
            qty = int(line.get('qty', 0) or 0)
        except (TypeError, ValueError):
            qty = 0
        if qty < 1:
            continue
        inv = db_find_one('inventory', {
            'store_id': order['store_id'],
            'product_id': line.get('product_id'),
            'variant_id': line.get('variant_id'),
        })
        if not inv:
            for inv_id, done_qty in applied:
                _pos_adjust_stock(inv_id, -sign * done_qty, sync_qr=False)
            return False
        sync_qr = sync_qr_on_restore if sign > 0 else False
        if not _pos_adjust_stock(inv['id'], sign * qty, sync_qr=sync_qr):
            for inv_id, done_qty in applied:
                _pos_adjust_stock(inv_id, -sign * done_qty, sync_qr=False)
            return False
        applied.append((inv['id'], qty))
    return True


def _public_customer(customer):
    addresses = _normalize_addresses(customer)
    default_addr = next((a for a in addresses if a.get('is_default')), addresses[0] if addresses else None)
    return {
        'id': customer.get('id'),
        'name': customer.get('name', ''),
        'phone': customer.get('phone', ''),
        'email': customer.get('email', ''),
        'address': (default_addr or {}).get('line1') or customer.get('address', ''),
        'addresses': addresses,
        'preferred_store_id': customer.get('preferred_store_id', ''),
        'has_account': bool(customer.get('password_hash')),
    }


def _session_customer():
    customer_id = session.get('customer_id')
    if not customer_id:
        return None
    customer = db_find_one('customers', {'id': customer_id})
    if not customer:
        session.pop('customer_id', None)
        return None
    return customer


def _normalize_addresses(customer):
    """Return a clean addresses list; migrate legacy single `address` string once."""
    if not customer:
        return []
    raw = customer.get('addresses')
    addresses = []
    if isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            line1 = (item.get('line1') or item.get('address') or '').strip()
            if not line1:
                continue
            addresses.append({
                'id': item.get('id') or new_id('addr_'),
                'label': (item.get('label') or 'Home').strip() or 'Home',
                'line1': line1,
                'area': (item.get('area') or '').strip(),
                'pincode': (item.get('pincode') or '').strip(),
                'is_default': bool(item.get('is_default')),
            })
    legacy = (customer.get('address') or '').strip()
    if legacy and not addresses:
        addresses = [{
            'id': new_id('addr_'),
            'label': 'Home',
            'line1': legacy,
            'area': '',
            'pincode': '',
            'is_default': True,
        }]
        db_update('customers', {'id': customer['id']}, {
            'addresses': addresses,
            'updated_at': now_iso(),
        })
    if addresses and not any(a.get('is_default') for a in addresses):
        addresses[0]['is_default'] = True
    return addresses


def _address_from_payload(data, existing=None):
    existing = existing or {}
    line1 = (data.get('line1') or data.get('address') or '').strip()
    if not line1:
        return None, 'Address line is required'
    label = (data.get('label') or existing.get('label') or 'Home').strip() or 'Home'
    return {
        'id': existing.get('id') or new_id('addr_'),
        'label': label[:40],
        'line1': line1[:240],
        'area': (data.get('area') or existing.get('area') or '').strip()[:80],
        'pincode': (data.get('pincode') or existing.get('pincode') or '').strip()[:12],
        'is_default': bool(data.get('is_default', existing.get('is_default', False))),
    }, None


def _save_customer_addresses(customer_id, addresses):
    if addresses and not any(a.get('is_default') for a in addresses):
        addresses[0]['is_default'] = True
    default = next((a for a in addresses if a.get('is_default')), None)
    updates = {
        'addresses': addresses,
        'address': (default or {}).get('line1', ''),
        'updated_at': now_iso(),
    }
    db_update('customers', {'id': customer_id}, updates)
    return db_find_one('customers', {'id': customer_id})


def _remember_order_address(customer, data):
    """Add checkout delivery address to the customer's address book when new."""
    line1 = (data.get('address') or '').strip()
    if not line1 or (data.get('delivery_mode') or 'delivery') == 'pickup':
        return
    addresses = _normalize_addresses(customer)
    lowered = line1.lower()
    for addr in addresses:
        if (addr.get('line1') or '').lower() == lowered:
            return
    new_addr = {
        'id': new_id('addr_'),
        'label': 'Delivery',
        'line1': line1[:240],
        'area': (data.get('area') or '').strip()[:80],
        'pincode': (data.get('pincode') or '').strip()[:12],
        'is_default': not addresses,
    }
    if new_addr['is_default']:
        for addr in addresses:
            addr['is_default'] = False
    addresses.append(new_addr)
    _save_customer_addresses(customer['id'], addresses)


@app.route('/api/auth/signup', methods=['POST'])
def api_auth_signup():
    data = parse_json()
    name = sanitize_text(data.get('name'), 120)
    phone = sanitize_text(data.get('phone'), 20)
    email = sanitize_text(data.get('email'), 120)
    password = data.get('password') or ''
    if not name or not phone or not password:
        return jsonify({'error': 'Name, phone and password are required'}), 400
    if len(password) < 4:
        return jsonify({'error': 'Password must be at least 4 characters'}), 400
    blocked, msg = prepare_login_attempt(phone)
    if blocked:
        return jsonify({'error': msg}), 429
    existing = db_find_one('customers', {'phone': phone})
    if existing and existing.get('password_hash'):
        record_login_failure(phone)
        return jsonify({'error': 'An account with this phone already exists. Please log in.'}), 409
    password_hash = hash_password(password)
    if existing:
        db_update('customers', {'id': existing['id']}, {
            'name': name,
            'email': email or existing.get('email', ''),
            'password_hash': password_hash,
            'cart': existing.get('cart') or [],
            'preferred_store_id': existing.get('preferred_store_id', ''),
            'updated_at': now_iso(),
        })
        customer = db_find_one('customers', {'id': existing['id']})
        log_activity('customer', f'Customer account activated for {name}')
    else:
        customer = {
            'id': new_id('cust_'),
            'name': name,
            'phone': phone,
            'email': email,
            'address': '',
            'addresses': [],
            'password_hash': password_hash,
            'cart': [],
            'preferred_store_id': '',
            'created_at': now_iso(),
        }
        db_insert('customers', customer)
        log_activity('customer', f'New customer account {name} signed up')
    clear_login_failures(phone)
    session['customer_id'] = customer['id']
    return jsonify({'ok': True, 'customer': _public_customer(customer)}), 201


@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    data = parse_json()
    phone = sanitize_text(data.get('phone'), 20)
    password = data.get('password') or ''
    if not phone or not password:
        return jsonify({'error': 'Enter phone and password'}), 400
    blocked, msg = prepare_login_attempt(phone)
    if blocked:
        return jsonify({'error': msg}), 429
    customer = db_find_one('customers', {'phone': phone})
    if not customer or not customer.get('password_hash'):
        record_login_failure(phone)
        return jsonify({'error': 'Incorrect phone or password'}), 401
    if not verify_password(customer['password_hash'], password):
        record_login_failure(phone)
        return jsonify({'error': 'Incorrect phone or password'}), 401
    clear_login_failures(phone)
    security_log('login_success', 'customer', identity=phone)
    session['customer_id'] = customer['id']
    return jsonify({
        'ok': True,
        'customer': _public_customer(customer),
        'cart': customer.get('cart') or [],
        'preferred_store_id': customer.get('preferred_store_id') or '',
    })


@app.route('/api/auth/logout', methods=['POST'])
def api_auth_logout():
    session.pop('customer_id', None)
    return jsonify({'ok': True})


@app.route('/api/auth/me')
def api_auth_me():
    customer_id = session.get('customer_id')
    if not customer_id:
        return jsonify({'authenticated': False, 'customer': None, 'cart': []})
    customer = db_find_one('customers', {'id': customer_id})
    if not customer:
        session.pop('customer_id', None)
        return jsonify({'authenticated': False, 'customer': None, 'cart': []})
    return jsonify({
        'authenticated': True,
        'customer': _public_customer(customer),
        'cart': customer.get('cart') or [],
        'preferred_store_id': customer.get('preferred_store_id') or '',
    })


@app.route('/api/account/orders')
def api_account_orders():
    customer_id = session.get('customer_id')
    if not customer_id:
        return jsonify({'error': 'Login required'}), 401
    customer = db_find_one('customers', {'id': customer_id})
    if not customer:
        session.pop('customer_id', None)
        return jsonify({'error': 'Login required'}), 401
    phone = (customer.get('phone') or '').strip()
    query = {'customer_id': customer.get('id')}
    if phone:
        query = {'$or': [
            {'customer_id': customer.get('id')},
            {'customer_phone': phone},
        ]}
    orders = db_find('orders', query, sort=[('created_at', -1)], limit=50)
    stores = {
        s['id']: s['name']
        for s in _cached_collection('stores', lambda: db_find('stores'))
    }
    items = []
    for o in orders:
        items.append({
            'id': o.get('order_id') or o.get('id'),
            'date': (o.get('created_at') or '')[:10],
            'status': normalize_status(o.get('status')),
            'total': o.get('total', 0),
            'area': stores.get(o.get('store_id'), o.get('address') or ''),
            'items': o.get('items') or [],
            'subtotal': o.get('subtotal', o.get('total', 0)),
            'delivery_fee': o.get('delivery_fee', 0),
            'discount': o.get('discount', 0),
            'gst_amount': o.get('gst_amount', 0),
            'coupon_code': o.get('coupon_code', ''),
            'delivery_mode': o.get('delivery_mode', 'delivery'),
            'payment_method': o.get('payment_method', 'cod'),
            'address': o.get('address', ''),
            'delivery_area': o.get('area', ''),
            'pincode': o.get('pincode', ''),
            'special_instructions': o.get('special_instructions') or o.get('notes', ''),
        })
    return jsonify({'items': items, 'customer': _public_customer(customer)})


@app.route('/api/account/cart', methods=['GET', 'PUT'])
def api_account_cart():
    """Persist the logged-in customer's cart in MongoDB."""
    customer_id = session.get('customer_id')
    if not customer_id:
        return jsonify({'error': 'Login required'}), 401
    customer = db_find_one('customers', {'id': customer_id})
    if not customer:
        return jsonify({'error': 'Customer not found'}), 404
    if request.method == 'GET':
        return jsonify({
            'cart': customer.get('cart') or [],
            'preferred_store_id': customer.get('preferred_store_id') or '',
        })
    data = parse_json()
    cart = data.get('cart') or []
    clean_cart = []
    for item in cart:
        try:
            qty = int(item.get('qty', 0))
        except (TypeError, ValueError):
            continue
        if qty < 1 or not item.get('id'):
            continue
        clean_cart.append({'id': str(item['id']), 'qty': min(qty, 500)})
    updates = {
        'cart': clean_cart,
        'updated_at': now_iso(),
    }
    if 'preferred_store_id' in data:
        updates['preferred_store_id'] = str(data.get('preferred_store_id') or '')
    db_update('customers', {'id': customer_id}, updates)
    return jsonify({'ok': True, 'cart': clean_cart})


@app.route('/api/account/profile', methods=['PUT'])
def api_account_profile():
    """Update name, email, and preferred store for the logged-in customer."""
    customer = _session_customer()
    if not customer:
        return jsonify({'error': 'Login required'}), 401
    data = parse_json()
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip()
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    if email and ('@' not in email or '.' not in email.split('@')[-1]):
        return jsonify({'error': 'Enter a valid email address'}), 400
    updates = {
        'name': name[:80],
        'email': email[:120],
        'updated_at': now_iso(),
    }
    if 'preferred_store_id' in data:
        updates['preferred_store_id'] = str(data.get('preferred_store_id') or '')
    db_update('customers', {'id': customer['id']}, updates)
    customer = db_find_one('customers', {'id': customer['id']})
    return jsonify({'ok': True, 'customer': _public_customer(customer)})


@app.route('/api/account/password', methods=['PUT'])
def api_account_password():
    customer = _session_customer()
    if not customer:
        return jsonify({'error': 'Login required'}), 401
    data = parse_json()
    current = data.get('current_password') or ''
    new_password = data.get('new_password') or ''
    if len(new_password) < 4:
        return jsonify({'error': 'New password must be at least 4 characters'}), 400
    if not customer.get('password_hash') or not verify_password(customer['password_hash'], current):
        return jsonify({'error': 'Current password is incorrect'}), 400
    db_update('customers', {'id': customer['id']}, {
        'password_hash': hash_password(new_password),
        'updated_at': now_iso(),
    })
    return jsonify({'ok': True})


@app.route('/api/account/addresses', methods=['GET', 'POST'])
def api_account_addresses():
    customer = _session_customer()
    if not customer:
        return jsonify({'error': 'Login required'}), 401
    if request.method == 'GET':
        return jsonify({'addresses': _normalize_addresses(customer)})

    data = parse_json()
    addr, err = _address_from_payload(data)
    if err:
        return jsonify({'error': err}), 400
    addresses = _normalize_addresses(customer)
    if addr['is_default'] or not addresses:
        for existing in addresses:
            existing['is_default'] = False
        addr['is_default'] = True
    addresses.append(addr)
    customer = _save_customer_addresses(customer['id'], addresses)
    return jsonify({'ok': True, 'addresses': _normalize_addresses(customer), 'customer': _public_customer(customer)}), 201


@app.route('/api/account/addresses/<addr_id>', methods=['PUT', 'DELETE'])
def api_account_address_detail(addr_id):
    customer = _session_customer()
    if not customer:
        return jsonify({'error': 'Login required'}), 401
    addresses = _normalize_addresses(customer)
    match = next((a for a in addresses if a.get('id') == addr_id), None)
    if not match:
        return jsonify({'error': 'Address not found'}), 404

    if request.method == 'DELETE':
        remaining = [a for a in addresses if a.get('id') != addr_id]
        customer = _save_customer_addresses(customer['id'], remaining)
        return jsonify({'ok': True, 'addresses': _normalize_addresses(customer), 'customer': _public_customer(customer)})

    data = parse_json()
    updated, err = _address_from_payload(data, existing=match)
    if err:
        return jsonify({'error': err}), 400
    updated['id'] = addr_id
    if updated['is_default']:
        for addr in addresses:
            addr['is_default'] = addr.get('id') == addr_id
    else:
        # Keep at least one default.
        others_default = any(a.get('is_default') and a.get('id') != addr_id for a in addresses)
        if not others_default:
            updated['is_default'] = True
    addresses = [updated if a.get('id') == addr_id else a for a in addresses]
    customer = _save_customer_addresses(customer['id'], addresses)
    return jsonify({'ok': True, 'addresses': _normalize_addresses(customer), 'customer': _public_customer(customer)})


@app.route('/api/settings')
def api_public_settings():
    s = get_settings()
    # Only expose storefront-relevant settings publicly
    return _public_cache_headers(jsonify({
        'min_order_value': s['min_order_value'],
        'delivery_fee_below_min': s['delivery_fee_below_min'],
        'free_delivery_above': s['free_delivery_above'],
        'same_day_delivery': s['same_day_delivery'],
        'cod_enabled': s['cod_enabled'],
        'whatsapp_payment_link': s['whatsapp_payment_link'],
        'whatsapp_number': s['whatsapp_number'],
        'gst_enabled': s['gst_enabled'],
        'fssai_number': s['fssai_number'],
        'halal_certified': s['halal_certified'],
        'seo_site_title': s['seo_site_title'],
        'seo_site_description': s['seo_site_description'],
    }), max_age=0)


@app.route('/api/storefront-content')
def api_storefront_content():
    return _public_cache_headers(jsonify(get_storefront_content()), max_age=0)


@app.route('/sitemap.xml')
def sitemap_xml():
    s = get_settings()
    base = (s.get('seo_canonical_base') or request.url_root).rstrip('/')
    urls = [f'{base}/']
    for c in db_find('categories', {'enabled': True}):
        urls.append(f"{base}/?category={c.get('slug', c['id'])}")
    for p in db_find('products'):
        if p.get('status') != 'disabled':
            urls.append(f"{base}/?product={p['id']}")
    body = ''.join(f'<url><loc>{u}</loc></url>' for u in urls)
    xml = f'<?xml version="1.0" encoding="UTF-8"?><urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">{body}</urlset>'
    return app.response_class(xml, mimetype='application/xml')


@app.route('/robots.txt')
def robots_txt():
    s = get_settings()
    base = (s.get('seo_canonical_base') or request.url_root).rstrip('/')
    txt = f'User-agent: *\nAllow: /\nDisallow: /admin\nDisallow: /api/admin\nSitemap: {base}/sitemap.xml\n'
    return app.response_class(txt, mimetype='text/plain')


def _coupon_discount(coupon, subtotal):
    if coupon.get('type') == 'percent':
        discount = subtotal * float(coupon.get('value', 0)) / 100.0
        cap = coupon.get('max_discount')
        if cap:
            discount = min(discount, float(cap))
    else:
        discount = float(coupon.get('value', 0))
    return round(min(discount, subtotal), 2)


def _coupon_phone_digits(phone):
    digits = re.sub(r'\D', '', str(phone or ''))
    if len(digits) > 10:
        digits = digits[-10:]
    return digits


def _normalize_customer_phone(phone):
    """Canonical 10-digit Indian mobile for website + in-store merge."""
    return _coupon_phone_digits(phone)


def _find_customer_by_phone(phone):
    """Find one customer by phone across common stored formats."""
    digits = _normalize_customer_phone(phone)
    if len(digits) != 10:
        return None
    for candidate in (digits, f'+91{digits}', f'91{digits}', f'0{digits}'):
        found = db_find_one('customers', {'phone': candidate})
        if found:
            return found
    return db_find_one('customers', {'phone': {'$regex': re.escape(digits)}})


def _lookup_customer_profile_by_phone(phone):
    """Return customer profile for POS autofill (customers table, else last order)."""
    digits = _normalize_customer_phone(phone)
    if len(digits) != 10:
        return None
    customer = _find_customer_by_phone(digits)
    if customer:
        return {
            'id': customer.get('id') or '',
            'name': (customer.get('name') or '').strip(),
            'phone': digits,
            'address': (customer.get('address') or '').strip(),
            'source': 'customer',
        }
    order = None
    rows = db_find(
        'orders',
        {
            'customer_phone': {'$regex': re.escape(digits)},
            'status': {'$nin': ['cancelled', 'canceled']},
        },
        sort=[('created_at', -1)],
        limit=1,
        projection={
            'customer_name': 1,
            'customer_phone': 1,
            'address': 1,
            'customer_id': 1,
        },
    )
    if rows:
        order = rows[0]
    if not order:
        return None
    return {
        'id': order.get('customer_id') or '',
        'name': (order.get('customer_name') or '').strip(),
        'phone': digits,
        'address': (order.get('address') or '').strip(),
        'source': 'order',
    }


def _customer_has_prior_orders(phone=None, customer_id=None):
    """True if this customer already placed a non-cancelled website/store order."""
    phone_digits = _coupon_phone_digits(phone)
    clauses = []
    if customer_id:
        clauses.append({'customer_id': customer_id})
    if phone_digits:
        # Match common stored formats: 10-digit, +91…, spaced.
        phone_regex = re.escape(phone_digits)
        clauses.append({'customer_phone': phone_digits})
        clauses.append({'customer_phone': {'$regex': phone_regex}})
    if not clauses:
        return False
    query = {
        '$and': [
            {'$or': clauses},
            {'status': {'$nin': ['cancelled', 'canceled']}},
        ]
    }
    return bool(db_find_one('orders', query))


def _coupon_first_order_error(coupon, phone=None, customer_id=None):
    """Return an error message if a first-order coupon cannot be used."""
    if not coupon or not coupon.get('first_order_only'):
        return None
    phone_digits = _coupon_phone_digits(phone)
    if not phone_digits and not customer_id:
        return 'Enter your phone number to use this first-order offer'
    if len(phone_digits) != 10 and not customer_id:
        return 'Enter a valid 10-digit phone number to use this first-order offer'
    if _customer_has_prior_orders(phone=phone_digits or phone, customer_id=customer_id):
        return 'This welcome offer is only for your first order'
    return None


def _apply_coupon_for_checkout(coupon_code, subtotal, phone=None, customer_id=None):
    """Validate coupon + first-order rules. Returns (discount, error_or_None, coupon_or_None)."""
    code = (coupon_code or '').strip().upper()
    if not code:
        return 0, None, None
    coupon = db_find_one('coupons', {'code': code})
    if not coupon or not coupon.get('active'):
        return 0, 'Invalid coupon code', None
    if coupon.get('expires_at') and coupon['expires_at'] < now_iso():
        return 0, 'Coupon has expired', None
    if subtotal < float(coupon.get('min_subtotal', 0) or 0):
        return 0, f"Minimum order ₹{coupon.get('min_subtotal', 0)} for this coupon", None
    first_err = _coupon_first_order_error(coupon, phone=phone, customer_id=customer_id)
    if first_err:
        return 0, first_err, None
    return _coupon_discount(coupon, subtotal), None, coupon


@app.route('/api/coupons/validate', methods=['POST'])
def api_validate_coupon():
    data = parse_json()
    code = (data.get('code') or '').strip().upper()
    try:
        subtotal = float(data.get('subtotal') or 0)
    except (TypeError, ValueError):
        subtotal = 0
    if not code:
        return jsonify({'valid': False, 'error': 'Enter a coupon code'}), 400
    phone = data.get('phone') or ''
    customer_id = session.get('customer_id') or data.get('customer_id') or ''
    if not phone and customer_id:
        cust = db_find_one('customers', {'id': customer_id})
        if cust:
            phone = cust.get('phone') or ''
    discount, error, coupon = _apply_coupon_for_checkout(
        code, subtotal, phone=phone, customer_id=customer_id
    )
    if error:
        status = 404 if error == 'Invalid coupon code' else 400
        return jsonify({'valid': False, 'error': error}), status
    return jsonify({
        'valid': True,
        'coupon': {
            'code': coupon.get('code'),
            'type': coupon.get('type'),
            'value': coupon.get('value'),
            'first_order_only': bool(coupon.get('first_order_only')),
        },
        'discount': discount,
    })


def log_activity(kind, text, meta=None):
    db_insert('activity', {
        'id': new_id('act_'),
        'kind': kind,  # order | inventory | store | customer | system
        'text': text,
        'meta': meta or {},
        'created_at': now_iso(),
    })


@app.route('/api/orders', methods=['POST'])
def api_place_order():
    data = parse_json()
    settings = get_settings()
    delivery_mode = data.get('delivery_mode', 'delivery')  # delivery | pickup
    required = ['name', 'phone', 'store_id', 'items']
    if delivery_mode == 'delivery':
        required.append('address')
    for r in required:
        if not data.get(r):
            return jsonify({'error': f'Missing {r}'}), 400
    if not data['items']:
        return jsonify({'error': 'Cart is empty'}), 400

    store = db_find_one('stores', {'id': data['store_id']})
    if not store or store.get('status') != 'active':
        return jsonify({'error': 'Invalid store'}), 400

    # Build lines + soft stock check; atomic reserve happens below (no oversell).
    lines = []
    subtotal = 0
    for raw in data['items']:
        pid = raw.get('product_id') or raw.get('id')
        vid = raw.get('variant_id') or (raw.get('variants') or [{}])[0].get('id')
        try:
            qty = int(raw.get('qty', 1))
        except (TypeError, ValueError):
            return jsonify({'error': 'Invalid quantity'}), 400
        if qty < 1 or qty > 500:
            return jsonify({'error': 'Invalid quantity'}), 400
        product = db_find_one('products', {'id': pid})
        if not product or product.get('status') == 'disabled':
            return jsonify({'error': f'Product {pid} not found'}), 400
        inv = db_find_one('inventory', {
            'store_id': data['store_id'], 'product_id': pid, 'variant_id': vid
        })
        if not inv:
            invs = db_find('inventory', {'store_id': data['store_id'], 'product_id': pid})
            inv = invs[0] if invs else None
            if inv:
                vid = inv['variant_id']
        if not inv or inv.get('stock', 0) < qty:
            return jsonify({'error': f'Insufficient stock for {product["name"]}'}), 400

        price = inv['price']
        gst_pct = effective_gst_percent(product)
        lines.append({
            'product_id': pid,
            'variant_id': vid,
            'name': product['name'],
            'qty': qty,
            'price': price,
            'gst_percent': gst_pct,
            'line_total': price * qty,
            'inventory_id': inv['id'],
        })
        subtotal += price * qty

    min_order = float(settings.get('min_order_value') or 0)
    if delivery_mode == 'delivery' and min_order > 0 and subtotal < min_order:
        return jsonify({
            'error': f'Minimum order value is ₹{min_order:,.0f}',
            'min_order_value': min_order,
        }), 400

    # Below the minimum order value the (configurable) delivery charge applies
    if delivery_mode == 'pickup':
        delivery = 0
    else:
        delivery = 0 if subtotal >= float(settings['free_delivery_above']) else float(settings['delivery_fee_below_min'])

    # Coupon (re-validated server-side, including first-order-only rules)
    coupon_code = (data.get('coupon_code') or '').strip().upper()
    discount = 0
    if coupon_code:
        session_customer_id = session.get('customer_id') or ''
        discount, coupon_error, _coupon = _apply_coupon_for_checkout(
            coupon_code,
            subtotal,
            phone=data.get('phone'),
            customer_id=session_customer_id,
        )
        if coupon_error:
            return jsonify({'error': coupon_error}), 400
        if not discount:
            coupon_code = ''

    # GST (informational; prices are treated as inclusive)
    gst_amount = 0
    if settings.get('gst_enabled'):
        for line in lines:
            pct = line.get('gst_percent', 0)
            if pct:
                gst_amount += line['line_total'] * pct / (100 + pct)
    gst_amount = round(gst_amount, 2)

    total = max(0, subtotal - discount) + delivery

    # Atomic stock reserve NOW — concurrent checkouts cannot oversell.
    reserved = []
    for line in lines:
        if not _pos_adjust_stock(line['inventory_id'], -int(line['qty'])):
            for inv_id, qty in reserved:
                _pos_adjust_stock(inv_id, qty)
            return jsonify({
                'error': f'Insufficient stock for {line["name"]} (just sold out)',
            }), 409
        reserved.append((line['inventory_id'], int(line['qty'])))

    # Upsert customer in MongoDB (never wipe password_hash / cart)
    # Same phone merges website + in-store into one customer profile.
    customer = None
    phone_digits = _normalize_customer_phone(data.get('phone'))
    phone_store = phone_digits if len(phone_digits) == 10 else str(data.get('phone') or '').strip()
    if session.get('customer_id'):
        customer = db_find_one('customers', {'id': session.get('customer_id')})
    if not customer and phone_store:
        customer = _find_customer_by_phone(phone_store) if len(phone_digits) == 10 else db_find_one('customers', {'phone': phone_store})
    if not customer:
        customer = {
            'id': new_id('cust_'),
            'name': data['name'],
            'phone': phone_store,
            'email': data.get('email', ''),
            'address': data.get('address', ''),
            'addresses': [],
            'cart': [],
            'preferred_store_id': data.get('store_id', ''),
            'created_at': now_iso(),
        }
        db_insert('customers', customer)
        log_activity('customer', f"New customer {data['name']} registered")
        _remember_order_address(customer, data)
    else:
        # Update contact fields but do not wipe account password / cart mid-edit beyond clearing cart after order
        updates = {
            'phone': phone_store or customer.get('phone', ''),
            'address': data.get('address', customer.get('address', '')),
            'email': data.get('email', customer.get('email', '')),
            'preferred_store_id': data.get('store_id', customer.get('preferred_store_id', '')),
            'cart': [],
            'updated_at': now_iso(),
        }
        # Website checkout can refresh name (shopper typed it); blank names fill from payload only
        if (data.get('name') or '').strip():
            updates['name'] = data['name']
        db_update('customers', {'id': customer['id']}, updates)
        customer = db_find_one('customers', {'id': customer['id']})
        _remember_order_address(customer, data)
    # SECURITY: never auto-login from phone alone (account takeover via guest checkout).
    # Keep session only if the shopper was already authenticated as this customer.
    if session.get('customer_id') and session.get('customer_id') != customer['id']:
        pass  # leave existing login untouched
    # Do not set session['customer_id'] from guest order phone.

    order_id = f"ORD{datetime.utcnow().strftime('%y%m%d')}{uuid.uuid4().hex[:5].upper()}"
    order = {
        'id': new_id('ord_'),
        'order_id': order_id,
        'customer_id': customer['id'],
        'customer_phone': phone_store or str(data['phone']),
        'customer_name': data['name'],
        'store_id': data['store_id'],
        'items': [{k: v for k, v in line.items() if k != 'inventory_id'} for line in lines],
        'subtotal': subtotal,
        'delivery_fee': delivery,
        'coupon_code': coupon_code if discount else '',
        'discount': discount,
        'gst_amount': gst_amount,
        'total': total,
        'status': 'new',
        'inventory_deducted': True,
        'payment_method': data.get('payment_method', 'cod'),
        'delivery_mode': delivery_mode,
        'channel': data.get('channel', 'website'),
        'address': data.get('address', ''),
        'area': data.get('area', ''),
        'pincode': data.get('pincode', ''),
        'notes': data.get('notes', ''),
        'special_instructions': data.get('special_instructions', ''),
        'created_at': now_iso(),
        'updated_at': now_iso(),
    }
    try:
        db_insert('orders', order)
    except Exception:
        for inv_id, qty in reserved:
            _pos_adjust_stock(inv_id, qty)
        raise
    log_activity('order', f"Order {order_id} placed — ₹{total:,.0f} · {store['name']}",
                 {'order_id': order_id, 'store_id': store['id']})

    return jsonify({'ok': True, 'order': order}), 201


# ---------------------------------------------------------------------------
# Admin auth
# ---------------------------------------------------------------------------

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if session.get('admin_ok') and request.method == 'GET':
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        username = (request.form.get('username') or '').strip().lower()
        password = request.form.get('password') or ''
        member, auth_error = _authenticate_staff_credentials(username, password)
        if member:
            set_admin_session(member)
            log_activity('system', f"{member.get('name')} signed in ({member.get('role')})")
            return _apply_admin_no_cache(redirect(url_for('admin_dashboard')))
        error = auth_error or 'Incorrect username or password'

    resp = make_response(render_template(
        'admin/login.html',
        error=error,
        db_mode=db_mode(),
        csrf_token=ensure_csrf_token(),
    ))
    return _apply_admin_no_cache(resp)


@app.route('/admin/logout')
def admin_logout():
    session.clear()
    resp = redirect(url_for('admin_login'))
    resp.headers['Clear-Site-Data'] = '"cache"'
    return _apply_admin_no_cache(resp)


@app.route('/admin/')
@app.route('/admin')
def admin_entry():
    if session.get('admin_ok'):
        return redirect(url_for('admin_dashboard'))
    resp = redirect(url_for('admin_login'))
    return _apply_admin_no_cache(resp)


@app.route('/api/admin/me')
@admin_required
def api_admin_me():
    admin = current_admin()
    return jsonify({
        'ok': True,
        'admin': admin,
        'can': {
            'super': admin_is_super(),
            'manage_staff': admin_is_super(),
            'manage_stores': admin_is_super(),
            'pages': list(ROLE_PAGES.get(admin['role'], [])),
        },
    })


@app.route('/admin/dashboard')
@page_required('dashboard')
def admin_dashboard():
    return render_template('admin/dashboard.html', page='dashboard', db_mode=db_mode())


@app.route('/admin/stores')
@page_required('stores')
def admin_stores():
    return render_template('admin/stores.html', page='stores', db_mode=db_mode())


@app.route('/admin/categories')
@page_required('categories')
def admin_categories():
    return render_template('admin/categories.html', page='categories', db_mode=db_mode())


@app.route('/admin/products')
@page_required('products')
def admin_products():
    return render_template('admin/products.html', page='products', db_mode=db_mode())


@app.route('/admin/inventory')
@page_required('inventory')
def admin_inventory():
    return render_template('admin/inventory.html', page='inventory', db_mode=db_mode())


@app.route('/admin/orders')
@page_required('orders')
def admin_orders():
    return render_template('admin/orders.html', page='orders', db_mode=db_mode())


@app.route('/admin/customers')
@page_required('customers')
def admin_customers():
    return render_template('admin/customers.html', page='customers', db_mode=db_mode())


@app.route('/admin/reports')
@page_required('reports')
def admin_reports():
    return render_template('admin/reports.html', page='reports', db_mode=db_mode())


@app.route('/admin/settings')
@page_required('settings')
def admin_settings():
    return render_template('admin/settings.html', page='settings', db_mode=db_mode())


@app.route('/admin/coupons')
@page_required('coupons')
def admin_coupons():
    return render_template('admin/coupons.html', page='coupons', db_mode=db_mode())


@app.route('/admin/staff')
@page_required('staff')
def admin_staff():
    return render_template('admin/staff.html', page='staff', db_mode=db_mode())


@app.route('/admin/storefront')
@page_required('storefront')
def admin_storefront():
    return render_template('admin/storefront.html', page='storefront', db_mode=db_mode())


@app.route('/admin/in-store')
@page_required('in_store')
def admin_in_store():
    return render_template('admin/in_store.html', page='in_store', db_mode=db_mode())


@app.route('/admin/qr-codes')
@page_required('qr_codes')
def admin_qr_codes():
    return render_template('admin/qr_codes.html', page='qr_codes', db_mode=db_mode())


def _add_months(year, month, delta):
    month += delta
    while month > 12:
        month -= 12
        year += 1
    while month < 1:
        month += 12
        year -= 1
    return year, month


def _created_at_range(period, anchor=None, now=None):
    """Return (start, end_exclusive) ISO date prefixes for created_at range queries."""
    now = now or datetime.utcnow()
    if anchor:
        if period == 'day':
            day = datetime.strptime(anchor, '%Y-%m-%d')
            return anchor, (day + timedelta(days=1)).strftime('%Y-%m-%d')
        if period == 'month':
            year, month = int(anchor[:4]), int(anchor[5:7])
            ny, nm = _add_months(year, month, 1)
            return f'{year:04d}-{month:02d}-01', f'{ny:04d}-{nm:02d}-01'
        if period == 'quarter':
            year = int(anchor[:4])
            quarter = int(anchor[-1])
            start_month = (quarter - 1) * 3 + 1
            ey, em = _add_months(year, start_month, 3)
            return f'{year:04d}-{start_month:02d}-01', f'{ey:04d}-{em:02d}-01'
        if period == 'year':
            year = int(anchor)
            return f'{year:04d}-01-01', f'{year + 1:04d}-01-01'
        return None, None

    if period == 'day':
        start = (now - timedelta(days=13)).strftime('%Y-%m-%d')
        end = (now + timedelta(days=1)).strftime('%Y-%m-%d')
        return start, end
    if period == 'month':
        y, m = _add_months(now.year, now.month, -11)
        return f'{y:04d}-{m:02d}-01', (now + timedelta(days=1)).strftime('%Y-%m-%d')
    if period == 'quarter':
        y = now.year
        q = (now.month - 1) // 3 + 1
        for _ in range(7):
            q -= 1
            if q == 0:
                q = 4
                y -= 1
        start_month = (q - 1) * 3 + 1
        return f'{y:04d}-{start_month:02d}-01', (now + timedelta(days=1)).strftime('%Y-%m-%d')
    if period == 'year':
        return f'{now.year - 4}-01-01', f'{now.year + 1}-01-01'
    return None, None


def _store_date_query(store_ids=None, start=None, end=None):
    """Build a Mongo/local query scoped by store ids and optional created_at range."""
    query = {}
    if store_ids:
        if len(store_ids) == 1:
            query['store_id'] = store_ids[0]
        else:
            query['store_id'] = {'$in': list(store_ids)}
    if start or end:
        created = {}
        if start:
            created['$gte'] = start
        if end:
            created['$lt'] = end
        query['created_at'] = created
    return query


def _cached_collection(name, loader):
    """Short TTL cache for reference collections used by dashboard/reports."""
    cached = _ref_cache.get(name)
    now = datetime.utcnow()
    if cached and (now - cached['at']).total_seconds() < _REF_CACHE_TTL:
        return cached['data']
    data = loader()
    _ref_cache[name] = {'at': now, 'data': data}
    if len(_ref_cache) > 20:
        oldest = sorted(_ref_cache.items(), key=lambda kv: kv[1]['at'])[:10]
        for key, _ in oldest:
            _ref_cache.pop(key, None)
    return data


def _invalidate_ref_cache(*names):
    """Drop cached reference maps so admin mutations stay consistent."""
    if not names:
        _ref_cache.clear()
        _badges_cache.clear()
        _stats_cache.clear()
        return
    for name in names:
        _ref_cache.pop(name, None)
    _badges_cache.clear()
    _stats_cache.clear()


def _status_query_values(status):
    """Include legacy status aliases when filtering orders."""
    status = normalize_status(status)
    if status == 'new':
        return ['new', 'pending', 'placed', 'Placed']
    return [status]


def _ensure_product_inventory_rows(product, default_price=0, default_stock=0):
    """Create missing inventory rows for a product without N+1 existence checks."""
    product_id = product.get('id')
    if not product_id:
        return
    existing = {
        (row.get('store_id'), row.get('variant_id'))
        for row in db_find('inventory', {'product_id': product_id}, projection={
            'store_id': 1, 'variant_id': 1,
        })
    }
    for sid in product.get('store_availability') or []:
        for var in product.get('variants') or []:
            key = (sid, var.get('id'))
            if not key[1] or key in existing:
                continue
            db_insert('inventory', {
                'id': new_id('inv_'),
                'store_id': sid,
                'product_id': product_id,
                'variant_id': var['id'],
                'price': float(default_price or 0),
                'stock': int(default_stock or 0),
                'updated_at': now_iso(),
            })
            existing.add(key)


def _timeline_key(created_at, period):
    """Map an ISO timestamp to a bucket key for day/month/quarter/year."""
    if not created_at:
        return None
    if period == 'day':
        return created_at[:10]
    if period == 'month':
        return created_at[:7]
    if period == 'quarter':
        try:
            year = int(created_at[:4])
            month = int(created_at[5:7])
            quarter = (month - 1) // 3 + 1
            return f'{year}-Q{quarter}'
        except (ValueError, IndexError):
            return None
    if period == 'year':
        return created_at[:4]
    return None


def _default_anchor(period, now=None):
    now = now or datetime.utcnow()
    if period == 'day':
        return now.strftime('%Y-%m-%d')
    if period == 'month':
        return now.strftime('%Y-%m')
    if period == 'quarter':
        return f'{now.year}-Q{(now.month - 1) // 3 + 1}'
    if period == 'year':
        return str(now.year)
    return now.strftime('%Y-%m')


def _normalize_anchor(period, anchor, now=None):
    """Validate/normalize an anchor value, or fall back to the current period."""
    now = now or datetime.utcnow()
    anchor = (anchor or '').strip()
    if not anchor:
        return _default_anchor(period, now)
    if period == 'day' and re.fullmatch(r'\d{4}-\d{2}-\d{2}', anchor):
        try:
            datetime.strptime(anchor, '%Y-%m-%d')
            return anchor
        except ValueError:
            return None
    if period == 'month' and re.fullmatch(r'\d{4}-\d{2}', anchor):
        try:
            datetime.strptime(anchor + '-01', '%Y-%m-%d')
            return anchor
        except ValueError:
            return None
    if period == 'quarter':
        match = re.fullmatch(r'(\d{4})-Q([1-4])', anchor, flags=re.IGNORECASE)
        if match:
            return f'{match.group(1)}-Q{match.group(2)}'
        return None
    if period == 'year' and re.fullmatch(r'\d{4}', anchor):
        return anchor
    return None


def _selection_caption(period, anchor):
    if period == 'day':
        return f'Date {anchor}'
    if period == 'month':
        return f'Month {anchor}'
    if period == 'quarter':
        return f'Quarter {anchor}'
    if period == 'year':
        return f'Year {anchor}'
    return PERIOD_CAPTIONS.get(period, '')


def _selection_label(period, anchor):
    if period == 'day':
        return f'Sales on {anchor}'
    if period == 'month':
        return f'Sales in {anchor}'
    if period == 'quarter':
        return f'Sales in {anchor}'
    if period == 'year':
        return f'Sales in {anchor}'
    return 'Period Sales'


def _matches_selection(created_at, period, anchor):
    return _timeline_key(created_at, period) == anchor


def _selection_bucket_key(created_at, period):
    """Bucket key used inside a selected day/month/quarter/year view."""
    if not created_at:
        return None
    if period == 'day':
        return created_at[:10]
    if period == 'month':
        return created_at[:10]
    if period in ('quarter', 'year'):
        return created_at[:7]
    return None


def _build_selected_timeline_buckets(period, anchor):
    """Build chart buckets for one selected calendar period."""
    buckets = {}
    if period == 'day':
        buckets[anchor] = {'label': anchor, 'sales': 0, 'orders': 0, 'customers': 0}
    elif period == 'month':
        year, month = int(anchor[:4]), int(anchor[5:7])
        days = calendar.monthrange(year, month)[1]
        for day in range(1, days + 1):
            key = f'{year:04d}-{month:02d}-{day:02d}'
            buckets[key] = {'label': key, 'sales': 0, 'orders': 0, 'customers': 0}
    elif period == 'quarter':
        year = int(anchor[:4])
        quarter = int(anchor[-1])
        start_month = (quarter - 1) * 3 + 1
        for month in range(start_month, start_month + 3):
            key = f'{year:04d}-{month:02d}'
            buckets[key] = {'label': key, 'sales': 0, 'orders': 0, 'customers': 0}
    elif period == 'year':
        year = int(anchor)
        for month in range(1, 13):
            key = f'{year:04d}-{month:02d}'
            buckets[key] = {'label': key, 'sales': 0, 'orders': 0, 'customers': 0}
    return buckets


def _build_timeline_buckets(period, now=None):
    """Build empty timeline buckets for the requested rolling period."""
    now = now or datetime.utcnow()
    buckets = {}

    if period == 'day':
        for i in range(13, -1, -1):
            key = (now - timedelta(days=i)).strftime('%Y-%m-%d')
            buckets[key] = {'label': key, 'sales': 0, 'orders': 0, 'customers': 0}
    elif period == 'month':
        y, m = now.year, now.month
        keys = []
        for _ in range(12):
            keys.append(f'{y:04d}-{m:02d}')
            m -= 1
            if m == 0:
                m = 12
                y -= 1
        keys.reverse()
        for key in keys:
            buckets[key] = {'label': key, 'sales': 0, 'orders': 0, 'customers': 0}
    elif period == 'quarter':
        y = now.year
        q = (now.month - 1) // 3 + 1
        keys = []
        for _ in range(8):
            keys.append(f'{y}-Q{q}')
            q -= 1
            if q == 0:
                q = 4
                y -= 1
        keys.reverse()
        for key in keys:
            buckets[key] = {'label': key, 'sales': 0, 'orders': 0, 'customers': 0}
    elif period == 'year':
        for i in range(4, -1, -1):
            key = str(now.year - i)
            buckets[key] = {'label': key, 'sales': 0, 'orders': 0, 'customers': 0}
    else:
        # default to month
        return _build_timeline_buckets('month', now)

    return buckets


PERIOD_CAPTIONS = {
    'day': 'Last 14 days',
    'month': 'Last 12 months',
    'quarter': 'Last 8 quarters',
    'year': 'Last 5 years',
}

@app.route('/api/admin/stats')
@admin_required
def api_admin_stats():
    period = request.args.get('period', 'month')
    if period not in ('day', 'month', 'quarter', 'year'):
        period = 'month'
    store_ids = resolve_store_ids()
    anchor_raw = request.args.get('anchor')
    selected_mode = anchor_raw is not None
    now = datetime.utcnow()
    anchor = None
    if selected_mode:
        anchor = _normalize_anchor(period, anchor_raw, now)
        if not anchor:
            return jsonify({'error': 'Invalid period selection'}), 400

    cache_key = f'{session.get("admin_user_id")}|{",".join(store_ids)}|{period}|{anchor or ""}'
    cached = _stats_cache.get(cache_key)
    if cached and (datetime.utcnow() - cached['at']).total_seconds() < _STATS_CACHE_TTL:
        return jsonify(cached['payload'])

    start, end = _created_at_range(period, anchor if selected_mode else None, now)
    order_query = _store_date_query(store_ids, start, end)
    customer_query = {}
    if start or end:
        created = {}
        if start:
            created['$gte'] = start
        if end:
            created['$lt'] = end
        customer_query['created_at'] = created

    orders = db_find('orders', order_query, projection=_ORDER_STATS_PROJECTION)
    customers = db_find('customers', customer_query, projection=_CUSTOMER_STATS_PROJECTION)

    all_stores = _cached_collection('stores', lambda: db_find('stores'))
    stores_by_id = {s['id']: s for s in all_stores}
    if store_ids:
        allowed = set(store_ids)
        stores = [s for s in all_stores if s.get('id') in allowed] or all_stores
    else:
        stores = all_stores

    products_by_id = _cached_collection(
        'products_by_id',
        lambda: {p['id']: p for p in db_find('products')}
    )
    categories_by_id = _cached_collection(
        'categories_by_id',
        lambda: {c['id']: c for c in db_find('categories')}
    )

    if selected_mode:
        buckets = _build_selected_timeline_buckets(period, anchor)
        scoped_orders = orders
        scoped_customers = customers
        period_caption = _selection_caption(period, anchor)
        selection_label = _selection_label(period, anchor)
    else:
        buckets = _build_timeline_buckets(period, now)
        scoped_orders = orders
        scoped_customers = customers
        period_caption = PERIOD_CAPTIONS.get(period, '')
        selection_label = None

    store_sales = {
        s['id']: {'store_id': s['id'], 'name': s['name'], 'sales': 0, 'orders': 0}
        for s in stores
    }
    status_counts = {}
    product_agg = {}
    pending = 0
    sales_today = 0
    sales_month = 0
    sales_quarter = 0
    sales_year = 0
    sales_selected = 0
    today = now.strftime('%Y-%m-%d')
    month = now.strftime('%Y-%m')
    quarter = f'{now.year}-Q{(now.month - 1) // 3 + 1}'
    year = str(now.year)
    keys_ordered = list(buckets.keys())
    prev_key = keys_ordered[-2] if len(keys_ordered) > 1 else None
    curr_key = keys_ordered[-1] if keys_ordered else None
    store_prev = {s['id']: 0 for s in stores}
    store_curr = {s['id']: 0 for s in stores}
    open_statuses = ('new', 'confirmed', 'ready', 'out_for_delivery')

    for o in scoped_orders:
        total = o.get('total', 0) or 0
        sid = o.get('store_id')
        created = o.get('created_at') or ''
        st = normalize_status(o.get('status'))
        status_counts[st] = status_counts.get(st, 0) + 1
        if st in open_statuses:
            pending += 1
        if sid in store_sales:
            store_sales[sid]['sales'] += total
            store_sales[sid]['orders'] += 1

        bucket_key = (
            _selection_bucket_key(created, period)
            if selected_mode else _timeline_key(created, period)
        )
        if bucket_key in buckets:
            buckets[bucket_key]['sales'] += total
            buckets[bucket_key]['orders'] += 1
            for it in o.get('items') or []:
                pid = it.get('product_id')
                if not pid:
                    continue
                row = product_agg.get(pid)
                if row is None:
                    row = {
                        'product_id': pid,
                        'name': it.get('name', pid),
                        'qty': 0,
                        'revenue': 0,
                    }
                    product_agg[pid] = row
                row['qty'] += it.get('qty', 0) or 0
                row['revenue'] += (it.get('price', 0) or 0) * (it.get('qty', 0) or 0)

        if prev_key and bucket_key == prev_key and sid in store_prev:
            store_prev[sid] += total
        if curr_key and bucket_key == curr_key and sid in store_curr:
            store_curr[sid] += total

        if created[:10] == today:
            sales_today += total
        if created[:7] == month:
            sales_month += total
        if _timeline_key(created, 'quarter') == quarter:
            sales_quarter += total
        if created[:4] == year:
            sales_year += total
        sales_selected += total

    for c in scoped_customers:
        bucket_key = (
            _selection_bucket_key(c.get('created_at'), period)
            if selected_mode else _timeline_key(c.get('created_at'), period)
        )
        if bucket_key in buckets:
            buckets[bucket_key]['customers'] += 1

    if selected_mode:
        if period == 'day':
            sales_today = sales_selected
        elif period == 'month':
            sales_month = sales_selected
        elif period == 'quarter':
            sales_quarter = sales_selected
        elif period == 'year':
            sales_year = sales_selected
        orders_selected = len(scoped_orders)
        customers_selected = len(scoped_customers)
    else:
        sales_selected = None
        orders_selected = None
        customers_selected = None

    def _delta(curr, prev):
        if prev <= 0:
            return None
        return round((curr - prev) / prev * 100, 1)

    bucket_list = list(buckets.values())
    curr_bucket = bucket_list[-1] if bucket_list else {'sales': 0, 'orders': 0}
    prev_bucket = bucket_list[-2] if len(bucket_list) > 1 else {'sales': 0, 'orders': 0}
    deltas = {
        'sales_pct': _delta(curr_bucket.get('sales', 0), prev_bucket.get('sales', 0)),
        'orders_pct': _delta(curr_bucket.get('orders', 0), prev_bucket.get('orders', 0)),
    }

    for sid, row in store_sales.items():
        row['change_pct'] = _delta(store_curr.get(sid, 0), store_prev.get(sid, 0))

    top_products = sorted(product_agg.values(), key=lambda x: x['revenue'], reverse=True)[:5]
    for tp in top_products:
        p = products_by_id.get(tp['product_id'])
        if p:
            cat = categories_by_id.get(p.get('category_id'))
            tp['category'] = cat['name'] if cat else ''

    settings = get_settings()
    threshold = int(settings['low_stock_threshold'])
    inv_query = _store_date_query(store_ids) if store_ids else {}
    inventory_rows = db_find('inventory', inv_query)
    max_stock = max((i.get('stock', 0) for i in inventory_rows), default=1) or 1
    low_stock = []
    prod_totals = {}
    for inv in inventory_rows:
        stock = inv.get('stock', 0) or 0
        pid = inv.get('product_id')
        prod_totals[pid] = prod_totals.get(pid, 0) + stock
        if stock <= threshold:
            p = products_by_id.get(pid)
            s = stores_by_id.get(inv.get('store_id'))
            low_stock.append({
                **inv,
                'product_name': p['name'] if p else pid,
                'store_name': s['name'] if s else inv.get('store_id'),
                'pct': min(100, round(stock / max_stock * 100)),
            })

    inv_health = []
    prod_max = max(prod_totals.values(), default=1) or 1
    for pid, total_stock in sorted(prod_totals.items(), key=lambda x: x[1])[:8]:
        p = products_by_id.get(pid)
        inv_health.append({
            'product_id': pid,
            'name': p['name'] if p else pid,
            'stock': total_stock,
            'pct': min(100, round(total_stock / prod_max * 100)),
            'low': total_stock <= threshold,
        })

    staff = _cached_collection('staff', lambda: db_find('staff'))
    if store_ids:
        allowed_staff = set(store_ids)
        staff_scoped = [
            m for m in staff
            if not m.get('store_id') or m.get('store_id') in allowed_staff
        ]
    else:
        staff_scoped = staff

    recent = sorted(scoped_orders, key=lambda x: x.get('created_at', ''), reverse=True)[:8]
    for o in recent:
        o['status'] = normalize_status(o.get('status'))
        s = stores_by_id.get(o.get('store_id'))
        o['store_name'] = s['name'] if s else ''

    sales_period = sum(b['sales'] for b in buckets.values())
    orders_in_buckets = sum(b['orders'] for b in buckets.values())
    kpis = {
        'sales_today': sales_today,
        'sales_month': sales_month,
        'sales_quarter': sales_quarter,
        'sales_year': sales_year,
        'sales_period': sales_period,
        'orders_total': len(scoped_orders),
        'customers_total': (
            len(scoped_customers) if selected_mode
            else len(_cached_collection(
                'customer_ids',
                lambda: db_find('customers', projection={'id': 1})
            ))
        ),
        'pending_orders': pending,
        'cancelled_orders': status_counts.get('cancelled', 0),
        'products_total': len(products_by_id),
        'stores_active': sum(1 for s in stores if s.get('status') == 'active'),
        'stores_total': len(stores),
        'staff_on_duty': sum(1 for m in staff_scoped if m.get('on_duty')),
        'staff_total': len(staff_scoped),
        'low_stock_count': len(low_stock),
        'avg_order_value': round(sales_period / max(1, orders_in_buckets)),
    }
    if selected_mode:
        kpis['sales_selected'] = sales_selected
        kpis['orders_selected'] = orders_selected
        kpis['customers_selected'] = customers_selected

    on_duty_staff = []
    for m in staff_scoped[:12]:
        on_duty_staff.append({
            'id': m.get('id'),
            'name': m.get('name'),
            'role': m.get('role'),
            'store_id': m.get('store_id') or '',
            'store_name': (stores_by_id.get(m.get('store_id')) or {}).get('name', 'All Stores') if m.get('store_id') else 'All Stores',
            'on_duty': bool(m.get('on_duty')),
        })

    activity = db_find('activity', sort=[('created_at', -1)], limit=12)

    payload = {
        'period': period,
        'anchor': anchor,
        'period_caption': period_caption,
        'selection_label': selection_label,
        'generated_at': now_iso(),
        'kpis': kpis,
        'deltas': deltas,
        'timeline': list(buckets.values()),
        'store_sales': list(store_sales.values()),
        'status_counts': status_counts,
        'top_products': top_products,
        'inventory_health': inv_health,
        'low_stock': low_stock[:20],
        'recent_orders': recent,
        'low_stock_threshold': threshold,
        'activity': activity,
        'on_duty_staff': on_duty_staff,
    }
    _stats_cache[cache_key] = {'at': datetime.utcnow(), 'payload': payload}
    # Keep cache map small
    if len(_stats_cache) > 40:
        oldest = sorted(_stats_cache.items(), key=lambda kv: kv[1]['at'])[:20]
        for key, _ in oldest:
            _stats_cache.pop(key, None)
    return jsonify(payload)


# --- Stores CRUD ---

@app.route('/api/admin/stores', methods=['GET', 'POST'])
@admin_required
def api_admin_stores():
    if request.method == 'GET':
        stores = db_find('stores', sort=[('name', 1)])
        scoped = resolve_store_scope(request.args.get('store_id'))
        if scoped:
            stores = [s for s in stores if s.get('id') == scoped]
        return jsonify(stores)
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can create stores'}), 403
    data = parse_json()
    settings = get_settings()
    store = {
        'id': new_id('store_'),
        'name': data.get('name', '').strip(),
        'tag': data.get('tag', ''),
        'address': data.get('address', ''),
        'contact': data.get('contact', ''),
        'hours': data.get('hours', '7 AM – 10 PM'),
        'status': data.get('status', 'active'),
        'delivery_radius_km': float(data.get('delivery_radius_km', settings['default_delivery_radius_km']) or 0),
        'manager': data.get('manager', ''),
        'created_at': now_iso(),
    }
    if not store['name']:
        return jsonify({'error': 'Store name required'}), 400
    db_insert('stores', store)
    _invalidate_ref_cache('stores')
    log_activity('store', f"Store {store['name']} added")
    return jsonify(store), 201


@app.route('/api/admin/stores/<store_id>', methods=['PUT', 'DELETE'])
@admin_required
def api_admin_store_detail(store_id):
    store = db_find_one('stores', {'id': store_id})
    if not store:
        return jsonify({'error': 'Store not found'}), 404

    if request.method == 'DELETE':
        if not admin_is_super():
            return jsonify({'error': 'Only Super Admin can delete stores'}), 403
        remaining = db_count('stores', {'id': {'$ne': store_id}})
        if remaining < 1:
            return jsonify({'error': 'Keep at least one store'}), 400
        # Keep orders, products, customers, QR units and other stores.
        # Drop only this store's inventory and unlink it from product availability.
        if _use_mongo and _mongo_db is not None:
            _mongo_db.products.update_many(
                {'store_availability': store_id},
                {'$pull': {'store_availability': store_id}},
            )
        db_delete('inventory', {'store_id': store_id})
        db_delete('stores', {'id': store_id})
        _invalidate_ref_cache()
        log_activity('system', f"Store {store.get('name', store_id)} deleted")
        return jsonify({'ok': True})
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can edit stores'}), 403
    data = parse_json()
    allowed = ['name', 'tag', 'address', 'contact', 'hours', 'status', 'delivery_radius_km', 'manager']
    updates = {k: data[k] for k in allowed if k in data}
    if 'delivery_radius_km' in updates:
        updates['delivery_radius_km'] = float(updates['delivery_radius_km'] or 0)
    updates['updated_at'] = now_iso()
    db_update('stores', {'id': store_id}, updates)
    _invalidate_ref_cache('stores')
    return jsonify(db_find_one('stores', {'id': store_id}))


# --- Categories CRUD ---

@app.route('/api/admin/categories', methods=['GET', 'POST'])
@admin_required
def api_admin_categories():
    if request.method == 'GET':
        return jsonify(db_find('categories', sort=[('sort_order', 1)]))
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can manage categories'}), 403
    data = parse_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Name required'}), 400
    slug = data.get('slug') or name.lower().replace(' ', '-')
    cat = {
        'id': new_id('cat_'),
        'name': name,
        'slug': slug,
        'code': '',
        'enabled': data.get('enabled', True),
        'seo_title': data.get('seo_title', name),
        'seo_description': data.get('seo_description', ''),
        'banner': data.get('banner', ''),
        'parameters': normalize_parameters(data.get('parameters')),
        'sort_order': int(data.get('sort_order', 99)),
        'created_at': now_iso(),
    }
    db_insert('categories', cat)
    ensure_category_code(cat)
    _invalidate_ref_cache('categories')
    _storefront_cache['data'] = None
    _storefront_cache['at'] = 0
    return jsonify(cat), 201


@app.route('/api/admin/categories/<cat_id>', methods=['PUT', 'DELETE'])
@admin_required
def api_admin_category_detail(cat_id):
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can manage categories'}), 403
    category = db_find_one('categories', {'id': cat_id})
    if not category:
        return jsonify({'error': 'Category not found'}), 404

    if request.method == 'DELETE':
        # Keep products; clear their category link so nothing is orphaned wrongly.
        blanked = db_update_many(
            'products',
            {'category_id': cat_id},
            {'category_id': '', 'updated_at': now_iso()},
        )
        # Drop deleted category from CMS Product Range picks.
        content = get_storefront_content()
        pr = dict(content.get('product_range') or {})
        ids = list(pr.get('category_ids') or [])
        if cat_id in ids:
            pr['category_ids'] = [x for x in ids if x != cat_id]
            save_storefront_content({'product_range': pr})
        else:
            _storefront_cache['data'] = None
            _storefront_cache['at'] = 0
        delete_upload_file(category.get('banner'))
        db_delete('categories', {'id': cat_id})
        _invalidate_ref_cache('categories', 'products_by_id')
        log_activity(
            'system',
            f"Category {category.get('name', cat_id)} deleted"
            + (f' ({blanked} products uncategorized)' if blanked else '')
        )
        return jsonify({'ok': True, 'products_uncategorized': blanked})

    data = parse_json()
    allowed = [
        'name', 'slug', 'enabled', 'seo_title', 'seo_description', 'banner',
        'parameters', 'sort_order'
    ]
    updates = {k: data[k] for k in allowed if k in data}
    if 'parameters' in updates:
        updates['parameters'] = normalize_parameters(updates['parameters'])
    if 'sort_order' in updates:
        updates['sort_order'] = int(updates['sort_order'])
    updates['updated_at'] = now_iso()
    if 'banner' in updates and updates.get('banner') != category.get('banner'):
        delete_upload_file(category.get('banner'))
    db_update('categories', {'id': cat_id}, updates)
    _invalidate_ref_cache('categories')
    _storefront_cache['data'] = None
    _storefront_cache['at'] = 0
    return jsonify(db_find_one('categories', {'id': cat_id}))


# --- Products CRUD ---

@app.route('/api/admin/products', methods=['GET', 'POST'])
@admin_required
def api_admin_products():
    if request.method == 'GET':
        products = db_find('products')
        categories = _cached_collection('categories', lambda: db_find('categories'))
        cat_names = {c['id']: c.get('name', '') for c in categories}
        lite = (request.args.get('lite') or '').strip().lower() in ('1', 'true', 'yes')
        out = []
        for p in products:
            row = dict(p)
            row['category_name'] = cat_names.get(p.get('category_id'), '')
            if lite:
                # Fast dropdown / QR filters — skip heavy fields
                images = row.get('images') or []
                out.append({
                    'id': row.get('id'),
                    'name': row.get('name'),
                    'sku': row.get('sku', ''),
                    'category_id': row.get('category_id', ''),
                    'category_name': row.get('category_name', ''),
                    'status': row.get('status', 'available'),
                    'variants': row.get('variants') or [],
                    'store_availability': row.get('store_availability') or [],
                    'images': images[:1],
                    'featured': bool(row.get('featured')),
                    'bestseller': bool(row.get('bestseller')),
                })
            else:
                out.append(row)
        return jsonify(out)

    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can manage products'}), 403
    data = parse_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Product name required'}), 400
    variants = data.get('variants') or [{'id': 'v1', 'label': 'Default', 'sku_suffix': 'DEF', 'unit': 'unit'}]
    for v in variants:
        if not v.get('id'):
            v['id'] = new_id('v')
    stores = _cached_collection('stores', lambda: db_find('stores'))
    category = db_find_one('categories', {'id': data.get('category_id', '')})
    parameters = (
        normalize_parameters(data.get('parameters'))
        if 'parameters' in data
        else normalize_parameters((category or {}).get('parameters'))
    )
    category_id = data.get('category_id', '')
    product = {
        'id': new_id('p'),
        'name': name,
        'description': data.get('description', ''),
        'sku': data.get('sku') or f'FAM-{uuid.uuid4().hex[:6].upper()}',
        'category_id': category_id,
        'images': data.get('images', []),
        'status': data.get('status', 'available'),
        'gst_percent': (
            FROZEN_GST_PERCENT if category_id == FROZEN_FOOD_CATEGORY_ID
            else float(data.get('gst_percent', 0) or 0)
        ),
        'expiry_info': data.get('expiry_info', ''),
        'nutritional_info': data.get('nutritional_info', ''),
        'parameters': parameters,
        'seo_title': data.get('seo_title', name),
        'seo_description': data.get('seo_description', ''),
        'featured': bool(data.get('featured', False)),
        'bestseller': bool(data.get('bestseller', False)),
        'inventory_model': data.get('inventory_model', 'variant'),
        'variants': variants,
        'store_availability': data.get('store_availability', [s['id'] for s in stores]),
        'created_at': now_iso(),
        'updated_at': now_iso(),
    }
    db_insert('products', product)
    _ensure_product_inventory_rows(
        product,
        default_price=data.get('default_price', 0),
        default_stock=data.get('default_stock', 0),
    )
    _invalidate_ref_cache('products_by_id')
    return jsonify(product), 201


@app.route('/api/admin/products/<product_id>', methods=['PUT', 'DELETE'])
@admin_required
def api_admin_product_detail(product_id):
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can manage products'}), 403
    product = db_find_one('products', {'id': product_id})
    if not product:
        return jsonify({'error': 'Product not found'}), 404

    if request.method == 'DELETE':
        for url in product.get('images') or []:
            delete_upload_file(url)
        db_delete('products', {'id': product_id})
        db_delete('inventory', {'product_id': product_id})
        _invalidate_ref_cache('products_by_id')
        log_activity('system', f"Product {product.get('name', product_id)} deleted")
        return jsonify({'ok': True})

    data = parse_json()
    allowed = [
        'name', 'description', 'sku', 'category_id', 'images', 'status',
        'gst_percent', 'expiry_info', 'nutritional_info', 'seo_title', 'seo_description',
        'featured', 'bestseller', 'inventory_model', 'variants', 'store_availability',
        'parameters'
    ]
    updates = {k: data[k] for k in allowed if k in data}
    if 'parameters' in updates:
        updates['parameters'] = normalize_parameters(updates['parameters'])
    if 'gst_percent' in updates:
        updates['gst_percent'] = float(updates['gst_percent'] or 0)
    next_category = updates.get('category_id', product.get('category_id'))
    if next_category == FROZEN_FOOD_CATEGORY_ID:
        updates['gst_percent'] = FROZEN_GST_PERCENT
    updates['updated_at'] = now_iso()
    if 'images' in updates:
        purge_removed_uploads(product.get('images') or [], updates.get('images') or [])
    db_update('products', {'id': product_id}, updates)

    # Ensure inventory rows exist for new store/variant combos
    product = db_find_one('products', {'id': product_id})
    if product:
        _ensure_product_inventory_rows(product, default_price=0, default_stock=0)
    _invalidate_ref_cache('products_by_id')
    return jsonify(product)


@app.route('/api/admin/products/<product_id>/image', methods=['POST', 'DELETE'])
@admin_required
def api_admin_product_image(product_id):
    product = db_find_one('products', {'id': product_id})
    if not product:
        return jsonify({'error': 'Product not found'}), 404

    if request.method == 'DELETE':
        url = (request.args.get('url') or parse_json().get('url') or '').strip()
        images = list(product.get('images') or [])
        if url not in images:
            return jsonify({'error': 'Image not found on product'}), 404
        images.remove(url)
        # Persist the updated image list in MongoDB, then remove the file from disk.
        db_update('products', {'id': product_id}, {'images': images, 'updated_at': now_iso()})
        delete_upload_file(url)
        log_activity('system', f"Product image removed from {product.get('name', product_id)}")
        return jsonify({'ok': True, 'images': images})

    if 'image' not in request.files:
        return jsonify({'error': 'No image file'}), 400
    f = request.files['image']
    if not f.filename or not allowed_file(f.filename):
        return jsonify({'error': 'Invalid image type'}), 400
    safe_name = secure_filename(f.filename) or 'upload.bin'
    ext = safe_name.rsplit('.', 1)[-1].lower() if '.' in safe_name else ''
    if ext not in ALLOWED_IMAGE_EXT:
        return jsonify({'error': 'Invalid image type'}), 400
    filename = f'{product_id}_{uuid.uuid4().hex[:8]}.{ext}'
    data = f.read()
    if not data:
        return jsonify({'error': 'Empty image file'}), 400
    ok_img, img_err = validate_image_bytes(data, ext)
    if not ok_img:
        return jsonify({'error': img_err or 'Invalid image content'}), 400
    url = save_upload_bytes(
        'products',
        filename,
        data,
        content_type=f.mimetype or _content_type_for_ext(ext),
    )
    images = list(product.get('images') or [])
    images.append(url)
    db_update('products', {'id': product_id}, {'images': images, 'updated_at': now_iso()})
    return jsonify({'ok': True, 'url': url, 'images': images})


# --- Inventory & pricing ---

@app.route('/api/admin/inventory', methods=['GET', 'POST'])
@admin_required
def api_admin_inventory():
    if request.method == 'POST':
        data = parse_json()
        row = db_find_one('inventory', {'id': data.get('inventory_id')})
        if not row:
            return jsonify({'error': 'Inventory row not found'}), 404
        denied = assert_store_access(row.get('store_id'))
        if denied:
            return denied
        try:
            quantity = int(data.get('quantity', 0))
        except (TypeError, ValueError):
            return jsonify({'error': 'Quantity must be a whole number'}), 400
        if quantity < 1 or quantity > 1000000:
            return jsonify({'error': 'Quantity must be between 1 and 1,000,000'}), 400
        updated = db_increment('inventory', {'id': row['id']}, 'stock', quantity)
        # Do not sync/create QR units here — that made Add Stock extremely slow.
        # Stock numbers update instantly; QR units come from Generate QR + punch.
        _badges_cache.clear()
        log_activity(
            'inventory',
            f"Added {quantity} units of stock",
            {'inventory_id': row['id'], 'store_id': row.get('store_id'),
             'product_id': row.get('product_id'), 'quantity': quantity},
        )
        return jsonify(updated)

    store_id = resolve_store_scope(request.args.get('store_id'))
    if store_id:
        pass
    elif not admin_is_super():
        return jsonify({'error': 'Your account is not assigned to a store'}), 403
    return jsonify(_enrich_inventory_rows(store_id))


@app.route('/api/admin/inventory/<inv_id>', methods=['PUT'])
@admin_required
def api_admin_inventory_update(inv_id):
    row = db_find_one('inventory', {'id': inv_id})
    if not row:
        return jsonify({'error': 'Inventory row not found'}), 404
    denied = assert_store_access(row.get('store_id'))
    if denied:
        return denied
    data = parse_json()
    updates = {}
    try:
        if 'price' in data:
            updates['price'] = float(data['price'])
        if 'stock' in data:
            updates['stock'] = int(data['stock'])
    except (TypeError, ValueError):
        return jsonify({'error': 'Price and stock must be valid numbers'}), 400
    if updates.get('price', 0) < 0 or updates.get('stock', 0) < 0:
        return jsonify({'error': 'Price and stock cannot be negative'}), 400
    updates['updated_at'] = now_iso()
    db_update('inventory', {'id': inv_id}, updates)
    updated = db_find_one('inventory', {'id': inv_id})
    # Fast inventory path: price/stock only. No QR mint/void here (POS billing untouched).
    _badges_cache.clear()
    return jsonify(updated)


# --- Orders ---

@app.route('/api/admin/orders')
@admin_required
def api_admin_orders():
    status = request.args.get('status')
    store_id = resolve_store_scope(request.args.get('store_id'))
    q = (request.args.get('q') or '').strip().lower()
    focus = (request.args.get('focus') or '').strip()
    page = max(1, int(request.args.get('page', 1)))
    per_page = min(100, int(request.args.get('per_page', 20)))
    if store_id:
        pass
    elif not admin_is_super():
        return jsonify({'error': 'Your account is not assigned to a store'}), 403

    query = {}
    if store_id:
        query['store_id'] = store_id
    if status:
        query['status'] = {'$in': _status_query_values(status)}
    if focus:
        query['$or'] = [{'order_id': focus}, {'id': focus}]
    elif q:
        rx = {'$regex': re.escape(q), '$options': 'i'}
        query['$or'] = [
            {'order_id': rx},
            {'customer_name': rx},
            {'customer_phone': {'$regex': re.escape(q)}},
        ]

    total = db_count('orders', query)
    start = (page - 1) * per_page
    chunk = db_find('orders', query, sort=[('created_at', -1)], skip=start, limit=per_page)
    stores_by_id = {
        s['id']: s for s in _cached_collection('stores', lambda: db_find('stores'))
    }
    for o in chunk:
        o['status'] = normalize_status(o.get('status'))
        s = stores_by_id.get(o.get('store_id'))
        o['store_name'] = s['name'] if s else ''
    settings = get_settings()
    return jsonify({'items': chunk, 'total': total, 'page': page, 'per_page': per_page,
                    'statuses': settings['order_statuses']})


def _normalize_admin_order_items(order, raw_items):
    """Validate edited order lines against the order store and live catalog."""
    normalized = []
    required = {}
    for raw in raw_items or []:
        product_id = raw.get('product_id')
        variant_id = raw.get('variant_id')
        product = db_find_one('products', {'id': product_id})
        inventory = db_find_one('inventory', {
            'store_id': order.get('store_id'),
            'product_id': product_id,
            'variant_id': variant_id,
        })
        if not product or not inventory:
            raise ValueError('An item or variant is unavailable at this store')
        try:
            qty = int(raw.get('qty', 0))
            price = round(float(raw.get('price', inventory.get('price', 0))), 2)
        except (TypeError, ValueError):
            raise ValueError('Item quantity and price must be valid numbers') from None
        if qty < 1 or qty > 500:
            raise ValueError('Item quantity must be between 1 and 500')
        if price < 0:
            raise ValueError('Item price cannot be negative')
        variant = next(
            (v for v in product.get('variants') or [] if v.get('id') == variant_id),
            {},
        )
        key = (product_id, variant_id)
        required[key] = required.get(key, 0) + qty
        normalized.append({
            'product_id': product_id,
            'variant_id': variant_id,
            'name': product.get('name', ''),
            'variant_label': variant.get('label', ''),
            'qty': qty,
            'price': price,
            'gst_percent': effective_gst_percent(product),
            'line_total': round(price * qty, 2),
        })
    if not normalized:
        raise ValueError('An order must contain at least one item')
    return normalized, required


def _has_stock_for_order(store_id, required):
    for (product_id, variant_id), qty in required.items():
        inventory = db_find_one('inventory', {
            'store_id': store_id,
            'product_id': product_id,
            'variant_id': variant_id,
        })
        if not inventory or int(inventory.get('stock', 0) or 0) < qty:
            return False
    return True


@app.route('/api/admin/orders/<order_id>', methods=['PUT', 'DELETE'])
@admin_required
def api_admin_order_update(order_id):
    order = db_find_one('orders', {'order_id': order_id}) or db_find_one('orders', {'id': order_id})
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    denied = assert_store_access(order.get('store_id'))
    if denied:
        return denied

    if request.method == 'DELETE':
        if order.get('inventory_deducted'):
            _apply_inventory_delta(order, +1, sync_qr_on_restore=False)
            log_activity('inventory', f"Stock restored for deleted order {order.get('order_id')}")
        db_delete('orders', {'id': order['id']})
        log_activity('order', f"Order {order.get('order_id')} deleted")
        return jsonify({'ok': True})

    data = parse_json()
    updates = {}
    old_status = normalize_status(order.get('status'))
    new_status = normalize_status(data.get('status', old_status))
    valid_statuses = set(get_settings().get('order_statuses') or [])
    if new_status not in valid_statuses:
        return jsonify({'error': 'Invalid order status'}), 400

    candidate = dict(order)
    items_changed = 'items' in data
    try:
        normalized_items, required = _normalize_admin_order_items(
            order, data.get('items', order.get('items') or [])
        )
    except ValueError as exc:
        return jsonify({'error': public_error(exc, 'Invalid order items')}), 400
    candidate['items'] = normalized_items

    active_statuses = {'confirmed', 'ready', 'out_for_delivery', 'delivered'}
    was_deducted = bool(order.get('inventory_deducted'))
    should_deduct = new_status in active_statuses
    inventory_rebalanced = items_changed or (was_deducted != should_deduct)

    if inventory_rebalanced and was_deducted:
        if not _apply_inventory_delta(order, +1):
            return jsonify({'error': 'Could not restore previous stock. Try again.'}), 409
    if inventory_rebalanced and should_deduct:
        if not _apply_inventory_delta(candidate, -1):
            if was_deducted:
                _apply_inventory_delta(order, -1)
            return jsonify({'error': 'Insufficient stock for the edited order'}), 409

    updates['status'] = new_status
    updates['inventory_deducted'] = should_deduct
    updates['items'] = normalized_items
    subtotal = round(sum(i['line_total'] for i in normalized_items), 2)
    try:
        discount = round(max(0, min(float(data.get('discount', order.get('discount', 0)) or 0), subtotal)), 2)
        delivery = round(max(0, float(data.get('delivery_fee', order.get('delivery_fee', 0)) or 0)), 2)
    except (TypeError, ValueError):
        if inventory_rebalanced:
            _apply_inventory_delta(candidate, +1) if should_deduct else None
            _apply_inventory_delta(order, -1) if was_deducted else None
        return jsonify({'error': 'Discount and delivery fee must be valid numbers'}), 400
    updates.update({
        'subtotal': subtotal,
        'discount': discount,
        'delivery_fee': delivery,
        'total': round(subtotal - discount + delivery, 2),
        'gst_amount': round(sum(
            item['line_total'] * item['gst_percent'] / (100 + item['gst_percent'])
            for item in normalized_items if item.get('gst_percent')
        ), 2) if get_settings().get('gst_enabled') else 0,
    })
    for field in (
        'customer_name', 'customer_phone', 'address', 'notes',
        'special_instructions', 'payment_method', 'delivery_mode',
    ):
        if field in data:
            updates[field] = str(data[field]).strip()

    if new_status != old_status:
        log_activity('order', f"Order {order.get('order_id')} → {new_status.replace('_', ' ')}")
    if inventory_rebalanced:
        action = 'deducted' if should_deduct else 'restored'
        log_activity('inventory', f"Stock {action} for edited order {order.get('order_id')}")

    updates['updated_at'] = now_iso()
    db_update('orders', {'id': order['id']}, updates)
    updated_order = db_find_one('orders', {'id': order['id']})

    phone = updates.get('customer_phone', order.get('customer_phone', '')).strip()
    name = updates.get('customer_name', order.get('customer_name', '')).strip()
    if phone:
        customer = (
            db_find_one('customers', {'id': order.get('customer_id')})
            or db_find_one('customers', {'phone': phone})
        )
        if customer:
            db_update('customers', {'id': customer['id']}, {
                'name': name or customer.get('name', ''),
                'phone': phone,
                'updated_at': now_iso(),
            })
            if order.get('customer_id') != customer['id']:
                db_update('orders', {'id': order['id']}, {'customer_id': customer['id']})
        else:
            customer_id = new_id('cust_')
            db_insert('customers', {
                'id': customer_id, 'name': name or 'Customer', 'phone': phone,
                'email': '', 'address': updates.get('address', order.get('address', '')),
                'created_at': now_iso(),
            })
            db_update('orders', {'id': order['id']}, {'customer_id': customer_id})
    return jsonify(db_find_one('orders', {'id': order['id']}))


# --- Customers (paginated) ---

@app.route('/api/admin/customers')
@admin_required
def api_admin_customers():
    q = request.args.get('q', '').strip().lower()
    focus = (request.args.get('focus') or '').strip()
    page = max(1, int(request.args.get('page', 1)))
    per_page = min(100, int(request.args.get('per_page', 20)))
    store_id = resolve_store_scope(request.args.get('store_id'))

    if focus:
        customers = db_find('customers', {'id': focus})
    elif q:
        rx = {'$regex': re.escape(q), '$options': 'i'}
        customers = db_find('customers', {'$or': [
            {'name': rx},
            {'phone': {'$regex': re.escape(q)}},
            {'email': rx},
        ]})
    elif store_id:
        slim_orders = db_find(
            'orders',
            {'store_id': store_id},
            projection={'customer_id': 1, 'customer_phone': 1},
        )
        store_customer_ids = {o['customer_id'] for o in slim_orders if o.get('customer_id')}
        store_phones = {str(o['customer_phone']) for o in slim_orders if o.get('customer_phone')}
        merged = {}
        for c in db_find('customers', {'preferred_store_id': store_id}):
            merged[c['id']] = c
        if store_customer_ids:
            for c in db_find('customers', {'id': {'$in': list(store_customer_ids)}}):
                merged[c['id']] = c
        if store_phones:
            for c in db_find('customers', {'phone': {'$in': list(store_phones)}}):
                merged[c['id']] = c
        customers = list(merged.values())
    else:
        customers = db_find('customers')

    if store_id and (focus or q):
        slim_orders = db_find(
            'orders',
            {'store_id': store_id},
            projection={'customer_id': 1, 'customer_phone': 1},
        )
        store_customer_ids = {o['customer_id'] for o in slim_orders if o.get('customer_id')}
        store_phones = {str(o['customer_phone']) for o in slim_orders if o.get('customer_phone')}
        customers = [
            c for c in customers
            if c.get('id') in store_customer_ids
            or (c.get('phone') and str(c.get('phone')) in store_phones)
            or c.get('preferred_store_id') == store_id
        ]

    customers = sorted(customers, key=lambda x: x.get('created_at', ''), reverse=True)
    total = len(customers)
    start = (page - 1) * per_page
    chunk = customers[start:start + per_page]

    # Load only orders needed for this page of customers
    chunk_ids = [c.get('id') for c in chunk if c.get('id')]
    chunk_phones = [str(c.get('phone')) for c in chunk if c.get('phone')]
    order_clauses = []
    if chunk_ids:
        order_clauses.append({'customer_id': {'$in': chunk_ids}})
    if chunk_phones:
        order_clauses.append({'customer_phone': {'$in': chunk_phones}})
    related_orders = []
    if order_clauses:
        order_query = {'$or': order_clauses} if len(order_clauses) > 1 else order_clauses[0]
        if store_id:
            order_query = {'$and': [order_query, {'store_id': store_id}]}
        related_orders = db_find('orders', order_query, projection=_CUSTOMER_ORDER_PROJECTION)

    orders_by_customer = {}
    for order in related_orders:
        keys = set()
        if order.get('customer_id'):
            keys.add(('id', order['customer_id']))
        if order.get('customer_phone'):
            keys.add(('phone', str(order['customer_phone'])))
        for key in keys:
            orders_by_customer.setdefault(key, []).append(order)

    for c in chunk:
        matched = []
        seen = set()
        for key in (('id', c.get('id')), ('phone', str(c.get('phone') or ''))):
            if not key[1]:
                continue
            for order in orders_by_customer.get(key, []):
                oid = order.get('id') or order.get('order_id')
                if oid in seen:
                    continue
                seen.add(oid)
                matched.append(order)
        matched = sorted(matched, key=lambda x: x.get('created_at', ''), reverse=True)
        c['order_count'] = len(matched)
        c['lifetime_value'] = sum(o.get('total', 0) for o in matched)
        c['orders'] = matched[:10]
        c['has_account'] = bool(c.get('password_hash'))
        c.pop('password_hash', None)
    return jsonify({'items': chunk, 'total': total, 'page': page, 'per_page': per_page})


@app.route('/api/admin/customers/<customer_id>', methods=['PUT', 'DELETE'])
@admin_required
def api_admin_customer_detail(customer_id):
    customer = db_find_one('customers', {'id': customer_id})
    if not customer:
        return jsonify({'error': 'Customer not found'}), 404

    if request.method == 'DELETE':
        if not admin_is_super():
            return jsonify({'error': 'Only Super Admin (abhi) can remove customer details'}), 403
        db_delete('customers', {'id': customer_id})
        log_activity('customer', f"Customer {customer.get('name') or customer_id} removed")
        return jsonify({'ok': True})

    # PUT — Super Admin only (login: abhi)
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin (abhi) can edit customer details'}), 403

    data = parse_json()
    name = (data.get('name') or '').strip()
    phone = re.sub(r'\D', '', str(data.get('phone') or ''))
    email = (data.get('email') or '').strip()
    address = (data.get('address') or '').strip()
    preferred_store_id = (data.get('preferred_store_id') or '').strip()

    if not name:
        return jsonify({'error': 'Customer name is required'}), 400
    if len(phone) < 10:
        return jsonify({'error': 'Enter a valid 10-digit phone number'}), 400

    conflict = db_find_one('customers', {'phone': phone})
    if conflict and conflict.get('id') != customer_id:
        return jsonify({'error': 'Another customer already uses this phone number'}), 400

    if preferred_store_id and not db_find_one('stores', {'id': preferred_store_id}):
        return jsonify({'error': 'Preferred store not found'}), 400

    updates = {
        'name': name[:120],
        'phone': phone[-10:] if len(phone) > 10 else phone,
        'email': email[:160],
        'address': address[:240],
        'preferred_store_id': preferred_store_id,
        'updated_at': now_iso(),
        'updated_by': session.get('admin_username') or session.get('admin_name') or 'abhi',
    }

    # Keep default address line in sync when legacy address field is edited
    addresses = _normalize_addresses(customer)
    if address:
        if addresses:
            for a in addresses:
                if a.get('is_default'):
                    a['line1'] = address[:240]
                    break
            else:
                addresses[0]['line1'] = address[:240]
            updates['addresses'] = addresses
        else:
            updates['addresses'] = [{
                'id': new_id('addr_'),
                'label': 'Home',
                'line1': address[:240],
                'area': '',
                'pincode': '',
                'is_default': True,
            }]

    db_update('customers', {'id': customer_id}, updates)
    # Keep order history labels in sync for this customer
    for order in db_find('orders', {'customer_id': customer_id}):
        q = {'id': order['id']} if order.get('id') else {'order_id': order.get('order_id')}
        db_update('orders', q, {
            'customer_name': updates['name'],
            'customer_phone': updates['phone'],
        })
    # Also match by previous phone if customer_id was missing on older orders
    old_phone = customer.get('phone')
    if old_phone and old_phone != updates['phone']:
        for order in db_find('orders', {'customer_phone': old_phone}):
            if order.get('customer_id') and order.get('customer_id') != customer_id:
                continue
            q = {'id': order['id']} if order.get('id') else {'order_id': order.get('order_id')}
            db_update('orders', q, {
                'customer_id': customer_id,
                'customer_name': updates['name'],
                'customer_phone': updates['phone'],
            })

    log_activity('customer', f"Customer {updates['name']} updated by {updates['updated_by']}")
    refreshed = db_find_one('customers', {'id': customer_id}) or {}
    refreshed.pop('password_hash', None)
    return jsonify(refreshed)

# --- Storefront CMS ---

@app.route('/api/admin/storefront-content', methods=['GET', 'PUT'])
@admin_required
def api_admin_storefront_content():
    if request.method == 'GET':
        return jsonify(get_storefront_content())
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can edit storefront content'}), 403
    content = save_storefront_content(parse_json())
    log_activity('system', 'Storefront content updated')
    return jsonify(content)


@app.route('/api/admin/content-image', methods=['POST'])
@admin_required
def api_admin_content_image():
    """Generic image upload used by the Storefront Content CMS (hero, why-us,
    category banners, and custom section photos). Bytes are stored in Mongo
    (media collection) so Vercel can serve them without local disk."""
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can upload content images'}), 403
    if 'image' not in request.files:
        return jsonify({'error': 'No image file'}), 400
    f = request.files['image']
    if not f.filename or not allowed_file(f.filename):
        return jsonify({'error': 'Invalid image type'}), 400
    data = f.read()
    ok_img, img_err = validate_image_bytes(data)
    if not ok_img:
        return jsonify({'error': img_err or 'Invalid image content'}), 400
    try:
        from io import BytesIO
        f.stream = BytesIO(data)
        f.seek = f.stream.seek
        f.tell = f.stream.tell
        f.read = f.stream.read
        url = save_upload_file('content', f)
    except ValueError as e:
        return jsonify({'error': 'Upload failed'}), 400
    return jsonify({'ok': True, 'url': url})


@app.route('/api/admin/media/sync', methods=['POST'])
@admin_required
def api_admin_media_sync():
    """Push any local upload files into the media store (useful after offline edits)."""
    count = sync_local_uploads_to_media()
    return jsonify({'ok': True, 'synced': count, 'db': db_mode()})


# --- In-store POS ---

def _pos_adjust_stock(inventory_id, quantity_delta, sync_qr=True):
    """Atomically adjust a POS inventory row; optionally keep qr_units in sync."""
    _require_mongo()
    query = {'id': inventory_id}
    if quantity_delta < 0:
        query['stock'] = {'$gte': abs(quantity_delta)}
    result = _mongo_db.inventory.update_one(
        query,
        {'$inc': {'stock': quantity_delta}, '$set': {'updated_at': now_iso()}},
    )
    ok = result.modified_count == 1
    # POS restore path only (positive delta) — unchanged from prior billing strategy.
    if ok and quantity_delta > 0 and sync_qr:
        inv = db_find_one('inventory', {'id': inventory_id})
        if inv:
            sync_qr_units_for_inventory_row(inv, create_missing=True)
    return ok


def _pos_catalog(store_id):
    """Lightweight in-store catalog — only in-stock variants for one store."""
    if not store_id:
        return {'store_id': '', 'categories': [], 'products': []}
    categories = db_find('categories', {'enabled': True}, projection={'id': 1, 'name': 1, 'code': 1})
    products = db_find('products', projection={
        'id': 1, 'name': 1, 'sku': 1, 'category_id': 1, 'status': 1, 'variants': 1,
    })
    inventory_rows = db_find('inventory', {'store_id': store_id})
    inv_by_product = {}
    for row in inventory_rows:
        inv_by_product.setdefault(row.get('product_id'), []).append(row)
    catalog_products = []
    for p in products:
        if (p.get('status') or 'available') == 'disabled':
            continue
        related = inv_by_product.get(p.get('id')) or []
        store_inventory = []
        for r in related:
            stock = int(r.get('stock') or 0)
            if stock < 1:
                continue
            store_inventory.append({
                'id': r.get('id'),
                'variant_id': r.get('variant_id'),
                'price': float(r.get('price') or 0),
                'stock': stock,
            })
        if not store_inventory:
            continue
        catalog_products.append({
            'id': p.get('id'),
            'name': p.get('name'),
            'sku': p.get('sku') or '',
            'category_id': p.get('category_id') or '',
            'variants': p.get('variants') or [],
            'store_inventory': store_inventory,
        })
    catalog_products.sort(key=lambda r: (r.get('name') or '').lower())
    categories.sort(key=lambda c: (c.get('name') or '').lower())
    return {
        'store_id': store_id,
        'categories': categories,
        'products': catalog_products,
    }


def _create_pos_order(data, staff):
    """Shared in-store billing for admin session + mobile Bearer auth.

    Returns (payload_dict, http_status). Success payload: {'ok': True, 'order': ...}.
    """
    staff = staff or {}
    is_super = staff.get('role') == ROLE_SUPER
    store_id = (data.get('store_id') or '').strip()
    if not is_super:
        locked = (staff.get('store_id') or '').strip()
        if not locked:
            return {'error': 'Your account is not assigned to a store'}, 403
        store_id = locked
    if not store_id:
        return {'error': 'Select a store'}, 400

    raw_items = data.get('items') or []
    payment_method = data.get('payment_method', 'cash')
    if payment_method not in ('cash', 'card', 'upi'):
        return {'error': 'Invalid payment method'}, 400
    store = db_find_one('stores', {'id': store_id})
    if not store or store.get('status') != 'active':
        return {'error': 'Select an active store'}, 400
    if not raw_items:
        return {'error': 'Add at least one item'}, 400

    if not raw_items:
        return {'error': 'Add at least one item'}, 400

    product_ids = list({raw.get('product_id') for raw in raw_items if raw.get('product_id')})
    products_by_id = {
        p['id']: p for p in db_find('products', {'id': {'$in': product_ids}})
    } if product_ids else {}
    inventory_by_key = {}
    for inv in db_find('inventory', {'store_id': store_id, 'product_id': {'$in': product_ids}}):
        inventory_by_key[(inv.get('product_id'), inv.get('variant_id'))] = inv

    lines = []
    subtotal = 0.0
    deductions = []
    claimed_by_line = []
    for raw in raw_items:
        try:
            qty = int(raw.get('qty', 0))
        except (TypeError, ValueError):
            return {'error': 'Invalid quantity'}, 400
        if qty < 1 or qty > 500:
            return {'error': 'Invalid quantity'}, 400
        product_id = raw.get('product_id')
        variant_id = raw.get('variant_id')
        product = products_by_id.get(product_id)
        inv = inventory_by_key.get((product_id, variant_id))
        if not product or not inv:
            return {'error': 'Product or variant is unavailable at this store'}, 400
        if inv.get('stock', 0) < qty:
            return {
                'error': f'Only {inv.get("stock", 0)} units available for {product["name"]}'
            }, 409
        preferred_unit_ids = []
        if raw.get('unit_id'):
            preferred_unit_ids.append(raw.get('unit_id'))
        for uid in (raw.get('unit_ids') or []):
            preferred_unit_ids.append(uid)
        claimed = []
        # Manual POS adds (no scanned QR) — stock-only, no QR sync during checkout.
        if preferred_unit_ids:
            claimed = claim_qr_units_for_sale(
                store_id, product_id, variant_id, qty,
                preferred_unit_ids=preferred_unit_ids,
                sync_missing=False,
            )
            if len(claimed) < qty:
                return {
                    'error': f'QR unit unavailable for {product["name"]}. Rescan or refresh.'
                }, 409
        variant = next(
            (v for v in product.get('variants') or [] if v.get('id') == variant_id),
            {},
        )
        price = float(inv.get('price', 0))
        line_total = round(price * qty, 2)
        unit_payload = [{
            'unit_id': u.get('id'),
            'qr_code': u.get('code'),
            'unit_serial': u.get('unit_serial'),
        } for u in claimed]
        lines.append({
            'product_id': product_id,
            'variant_id': variant_id,
            'name': product['name'],
            'variant_label': variant.get('label', ''),
            'qty': qty,
            'price': price,
            'gst_percent': effective_gst_percent(product),
            'line_total': line_total,
            'qr_units': unit_payload,
            'qr_codes': [u['qr_code'] for u in unit_payload if u.get('qr_code')],
            'unit_serials': [u['unit_serial'] for u in unit_payload if u.get('unit_serial')],
        })
        deductions.append((inv['id'], qty))
        claimed_by_line.append([u.get('id') for u in claimed])
        subtotal += line_total

    try:
        discount = float(data.get('discount', 0) or 0)
    except (TypeError, ValueError):
        return {'error': 'Invalid discount'}, 400
    discount = round(max(0, min(discount, subtotal)), 2)

    deducted = []
    for inventory_id, qty in deductions:
        if not _pos_adjust_stock(inventory_id, -qty):
            for completed_id, completed_qty in deducted:
                _pos_adjust_stock(completed_id, completed_qty)
            return {'error': 'Stock changed while billing. Please refresh and try again.'}, 409
        deducted.append((inventory_id, qty))

    phone = _normalize_customer_phone(data.get('customer_phone') or '')
    name = (data.get('customer_name') or '').strip() or 'Walk-in Customer'
    customer_id = ''
    if phone and len(phone) == 10:
        customer = _find_customer_by_phone(phone)
        if customer:
            customer_id = customer['id']
            # Keep existing profile name — never override on repeat POS visits.
            updates = {'updated_at': now_iso()}
            if _normalize_customer_phone(customer.get('phone')) != phone:
                updates['phone'] = phone
            db_update('customers', {'id': customer_id}, updates)
            # Prefer stored name on the bill when staff left name blank / walk-in default
            stored_name = (customer.get('name') or '').strip()
            if stored_name and name in ('', 'Walk-in Customer'):
                name = stored_name
        else:
            customer_id = new_id('cust_')
            db_insert('customers', {
                'id': customer_id,
                'name': name,
                'phone': phone,
                'email': '',
                'address': '',
                'created_at': now_iso(),
            })
    elif str(data.get('customer_phone') or '').strip():
        # Invalid / incomplete phone — keep on the bill only, skip customer merge
        phone = str(data.get('customer_phone') or '').strip()

    gst_amount = 0.0
    if get_settings().get('gst_enabled'):
        for line in lines:
            pct = line['gst_percent']
            if pct:
                gst_amount += line['line_total'] * pct / (100 + pct)
    gst_amount = round(gst_amount, 2)
    total = round(subtotal - discount, 2)
    timestamp = datetime.utcnow()
    order_id = f"POS{timestamp.strftime('%y%m%d')}{uuid.uuid4().hex[:5].upper()}"
    order = {
        'id': new_id('ord_'),
        'order_id': order_id,
        'customer_id': customer_id,
        'customer_phone': phone,
        'customer_name': name,
        'store_id': store_id,
        'items': lines,
        'subtotal': round(subtotal, 2),
        'delivery_fee': 0,
        'discount': discount,
        'gst_amount': gst_amount,
        'total': total,
        'status': 'delivered',
        'inventory_deducted': True,
        'payment_method': payment_method,
        'delivery_mode': 'in_store',
        'channel': 'in_store',
        'address': store.get('address', ''),
        'notes': data.get('notes', ''),
        'staff_id': staff.get('id') or '',
        'staff_name': (
            staff.get('name')
            or (data.get('staff_name') or '').strip()
            or 'Staff'
        ),
        'created_at': now_iso(),
        'updated_at': now_iso(),
    }
    try:
        db_insert('orders', order)
    except Exception:
        for inventory_id, qty in deducted:
            _pos_adjust_stock(inventory_id, qty)
        raise
    sold_ids = [uid for group in claimed_by_line for uid in group]
    mark_qr_units_sold(sold_ids, order_id=order_id)
    log_activity(
        'order',
        f'In-store bill {order_id} created — ₹{total:,.0f} · {store["name"]}',
        {
            'order_id': order_id,
            'store_id': store_id,
            'channel': 'in_store',
            'qr_units': len(sold_ids),
            'staff_id': staff.get('id'),
        },
    )
    order['receipt'] = _build_receipt_data(order)
    return {'ok': True, 'order': order}, 201


def _list_pos_orders(store_id, limit=20):
    limit = min(100, max(1, int(limit or 20)))
    query = {'channel': 'in_store'}
    if store_id:
        query['store_id'] = store_id
    orders = db_find(
        'orders',
        query,
        sort=[('created_at', -1)],
        limit=limit,
        projection={
            'id': 1,
            'order_id': 1,
            'store_id': 1,
            'customer_name': 1,
            'customer_phone': 1,
            'total': 1,
            'payment_method': 1,
            'created_at': 1,
            'status': 1,
            'channel': 1,
        },
    )
    stores = {s['id']: s['name'] for s in _cached_collection('stores', lambda: db_find('stores', projection={'id': 1, 'name': 1}))}
    for order in orders:
        order['store_name'] = stores.get(order.get('store_id'), '')
    return orders


def _ensure_store_inventory_coverage(store_id):
    """Create any missing inventory rows for products assigned to this store (one pass)."""
    if not store_id:
        return 0
    products = db_find(
        'products',
        {'store_availability': store_id, 'status': {'$ne': 'disabled'}},
        projection={'id': 1, 'variants': 1, 'store_availability': 1},
    )
    existing = {
        (row.get('product_id'), row.get('variant_id'))
        for row in db_find(
            'inventory',
            {'store_id': store_id},
            projection={'product_id': 1, 'variant_id': 1},
        )
    }
    created = 0
    now = now_iso()
    for product in products:
        product_id = product.get('id')
        if not product_id:
            continue
        for var in product.get('variants') or []:
            vid = var.get('id')
            if not vid:
                continue
            key = (product_id, vid)
            if key in existing:
                continue
            db_insert('inventory', {
                'id': new_id('inv_'),
                'store_id': store_id,
                'product_id': product_id,
                'variant_id': vid,
                'price': 0,
                'stock': 0,
                'updated_at': now,
            })
            existing.add(key)
            created += 1
    return created


def _enrich_inventory_rows(store_id):
    low_stock_threshold = int(get_settings().get('low_stock_threshold', 10))
    # Keep product↔store inventory rows complete so Add Stock lists every assigned product.
    if store_id:
        _ensure_store_inventory_coverage(store_id)
    query = {'store_id': store_id} if store_id else {}
    rows = db_find('inventory', query)
    products_by_id = _cached_collection(
        'products_by_id',
        lambda: {p['id']: p for p in db_find('products')}
    )
    stores_by_id = {
        s['id']: s for s in _cached_collection('stores', lambda: db_find('stores'))
    }
    categories_by_id = {c['id']: c for c in db_find('categories')}
    enriched = []
    for r in rows:
        p = products_by_id.get(r.get('product_id'))
        if not p:
            continue
        if (p.get('status') or 'available') == 'disabled':
            continue
        s = stores_by_id.get(r.get('store_id'))
        variant_label = ''
        for v in p.get('variants') or []:
            if v.get('id') == r.get('variant_id'):
                variant_label = v.get('label', '')
                break
        cat = categories_by_id.get(p.get('category_id')) or {}
        enriched.append({
            **r,
            'product_name': p.get('name') or '',
            'sku': p.get('sku', '') or '',
            'category_id': p.get('category_id') or '',
            'category_name': cat.get('name') or '',
            'variant_label': variant_label,
            'store_name': s['name'] if s else '',
            'inventory_model': p.get('inventory_model') if p else 'variant',
            'low_stock': int(r.get('stock', 0) or 0) <= low_stock_threshold,
            'low_stock_threshold': low_stock_threshold,
        })
    enriched.sort(key=lambda row: ((row.get('product_name') or '').lower(), row.get('variant_label') or ''))
    return enriched


def _format_receipt_phone(phone):
    digits = re.sub(r'\D', '', str(phone or ''))
    if len(digits) == 10:
        return f'{digits[:5]} {digits[5:]}'
    return phone or '—'


def _format_receipt_datetime(iso_str):
    if not iso_str:
        return ''
    try:
        raw = str(iso_str).replace('Z', '+00:00')
        dt = datetime.fromisoformat(raw)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        ist = timezone(timedelta(hours=5, minutes=30))
        return dt.astimezone(ist).strftime('%d/%m/%Y %H:%M')
    except (TypeError, ValueError):
        return str(iso_str)[:16].replace('T', ' ')


def _receipt_item_label(item):
    name = item.get('name') or ''
    variant = (item.get('variant_label') or '').strip()
    if variant and variant not in name:
        return f'{name} ({variant})'
    return name


def _build_receipt_data(order):
    """Structured payload for 80mm thermal receipts (Essae PR-55 and similar)."""
    settings = get_settings()
    store = db_find_one('stores', {'id': order.get('store_id')}) or {}
    gst_amount = float(order.get('gst_amount') or 0)
    half_gst = round(gst_amount / 2, 2) if gst_amount else 0.0
    discount = float(order.get('discount') or 0)
    payment = (order.get('payment_method') or 'cash').lower()
    payment_labels = {
        'cash': 'Cash',
        'card': 'Card',
        'upi': 'UPI',
        'cod': 'Cash on Delivery',
    }
    channel = order.get('channel') or order.get('delivery_mode') or ''
    return {
        'business_name': 'FISH AND MEAT',
        'title': 'TAX INVOICE',
        'address': store.get('address') or order.get('address') or '',
        'gstin': settings.get('gst_number') or '',
        'fssai': settings.get('fssai_number') or '',
        'invoice': order.get('order_id') or order.get('id') or '',
        'date': _format_receipt_datetime(order.get('created_at')),
        'customer': order.get('customer_name') or 'Walk-in Customer',
        'mobile': _format_receipt_phone(order.get('customer_phone')),
        'order_type': 'In-Store' if channel == 'in_store' else 'Delivery',
        'items': [{
            'name': _receipt_item_label(it),
            'qty': it.get('qty', 0),
            'rate': f"{float(it.get('price', 0) or 0):,.2f}",
            'amount': f"{float(it.get('line_total', 0) or (it.get('price', 0) * it.get('qty', 0))):,.2f}",
        } for it in (order.get('items') or [])],
        'subtotal': f"{float(order.get('subtotal', 0) or 0):,.2f}",
        'discount': f"{discount:,.2f}" if discount else '',
        'cgst': f"{half_gst:,.2f}" if half_gst else '',
        'sgst': f"{half_gst:,.2f}" if half_gst else '',
        'delivery': f"{float(order.get('delivery_fee', 0) or 0):,.2f}",
        'total': f"{float(order.get('total', 0) or 0):,.2f}",
        'payment_line': f"Paid by {payment_labels.get(payment, payment.upper())}",
        'footer_lines': ['Thank you. Please visit again.'],
        'fine_print': [
            'Frozen items: keep at -18°C.',
            'No return on perishable goods.',
        ],
    }


def _render_thermal_receipt_html(order, auto_print=False):
    return render_template(
        'receipt/thermal.html',
        receipt=_build_receipt_data(order),
        auto_print=auto_print,
    )


def _find_order_for_receipt(order_id):
    return db_find_one('orders', {'order_id': order_id}) or db_find_one('orders', {'id': order_id})


def _build_invoice_pdf(order):
    """Return BytesIO PDF for an order invoice."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    settings = get_settings()
    store = db_find_one('stores', {'id': order.get('store_id')}) or {}

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Heading1'], textColor=colors.HexColor('#1E3A22'), fontSize=17)
    sub_style = ParagraphStyle('S', parent=styles['Normal'], textColor=colors.HexColor('#55594F'), fontSize=9)
    head_style = ParagraphStyle('H', parent=styles['Heading2'], textColor=colors.HexColor('#A5342A'), fontSize=12)

    story = [Paragraph('FISH AND MEAT — Tax Invoice', title_style)]
    meta_lines = [f"Invoice for Order {order['order_id']} · {(order.get('created_at') or '')[:10]}"]
    if settings.get('gst_number'):
        meta_lines.append(f"GSTIN: {settings['gst_number']}")
    if settings.get('fssai_number'):
        meta_lines.append(f"FSSAI: {settings['fssai_number']}")
    if settings.get('halal_certified'):
        meta_lines.append('Halal Certified')
    story.append(Paragraph(' · '.join(meta_lines), sub_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"Billed to: {order.get('customer_name')} · {order.get('customer_phone')}<br/>"
        f"{order.get('address') or 'Store pickup'}<br/>"
        f"Store: {store.get('name', '')} · {store.get('address', '')}", sub_style))
    story.append(Spacer(1, 14))

    rows = [['Item', 'Qty', 'Rate (₹)', 'GST %', 'Amount (₹)']]
    for it in order.get('items') or []:
        rows.append([
            it.get('name', ''), str(it.get('qty', 0)),
            f"{it.get('price', 0):,.2f}",
            f"{it.get('gst_percent', 0):g}",
            f"{it.get('price', 0) * it.get('qty', 0):,.2f}",
        ])
    rows.append(['', '', '', 'Subtotal', f"{order.get('subtotal', 0):,.2f}"])
    if order.get('discount'):
        rows.append(['', '', '', f"Discount ({order.get('coupon_code', '')})", f"-{order['discount']:,.2f}"])
    rows.append(['', '', '', 'Delivery', f"{order.get('delivery_fee', 0):,.2f}"])
    if order.get('gst_amount'):
        rows.append(['', '', '', 'GST included', f"{order['gst_amount']:,.2f}"])
    rows.append(['', '', '', 'Total', f"{order.get('total', 0):,.2f}"])

    t = Table(rows, colWidths=[2.6 * inch, 0.7 * inch, 1.1 * inch, 1.3 * inch, 1.3 * inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A22')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E7E2D6')),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('FONTNAME', (3, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (3, -1), (-1, -1), colors.HexColor('#A5342A')),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))
    story.append(Paragraph(
        f"Payment: {(order.get('payment_method') or 'cod').upper()} · Mode: {(order.get('delivery_mode') or 'delivery').title()}",
        head_style))

    doc.build(story)
    buf.seek(0)
    return buf


@app.route('/api/admin/pos/customer-lookup', methods=['GET'])
@admin_required
def api_admin_pos_customer_lookup():
    """Lookup returning customer by 10-digit phone for in-store billing autofill."""
    phone = request.args.get('phone') or ''
    digits = _normalize_customer_phone(phone)
    if len(digits) != 10:
        return jsonify({'found': False, 'error': 'Enter a valid 10-digit mobile number'}), 400
    profile = _lookup_customer_profile_by_phone(digits)
    if not profile or not profile.get('name'):
        return jsonify({'found': False, 'phone': digits})
    return jsonify({'found': True, 'customer': profile})


@app.route('/api/admin/pos/catalog', methods=['GET'])
@admin_required
def api_admin_pos_catalog():
    store_id = resolve_store_scope(request.args.get('store_id'))
    if not store_id:
        return jsonify({'error': 'Select a store'}), 400
    denied = assert_store_access(store_id)
    if denied:
        return denied
    return jsonify(_pos_catalog(store_id))


@app.route('/api/admin/pos/orders', methods=['GET', 'POST'])
@admin_required
def api_admin_pos_orders():
    if request.method == 'GET':
        store_id = resolve_store_scope(request.args.get('store_id'))
        return jsonify(_list_pos_orders(store_id, request.args.get('limit', 20)))

    data = parse_json()
    store_id = resolve_store_scope(data.get('store_id'))
    denied = assert_store_access(store_id)
    if denied:
        return denied
    data = dict(data or {})
    data['store_id'] = store_id
    payload, status = _create_pos_order(data, current_admin())
    return jsonify(payload), status


# --- QR Codes (Super Admin) ---

def _enrich_qr_product_row(product, categories_by_id, stores_by_id, inventory_rows):
    cat = categories_by_id.get(product.get('category_id')) or {}
    store_ids = product.get('store_availability') or []
    related = [r for r in inventory_rows if r.get('product_id') == product.get('id')]
    total_stock = sum(int(r.get('stock', 0) or 0) for r in related)
    prices = [float(r.get('price', 0) or 0) for r in related if float(r.get('price', 0) or 0) > 0]
    store_names = []
    store_details = []
    for sid in store_ids:
        store = stores_by_id.get(sid)
        if store:
            store_names.append(store.get('name', sid))
    for row in related:
        store = stores_by_id.get(row.get('store_id'))
        variant_label = ''
        for v in product.get('variants') or []:
            if v.get('id') == row.get('variant_id'):
                variant_label = v.get('label', '')
                break
        store_details.append({
            'store_id': row.get('store_id'),
            'store_name': store.get('name') if store else row.get('store_id'),
            'variant_id': row.get('variant_id'),
            'variant_label': variant_label,
            'price': float(row.get('price', 0) or 0),
            'stock': int(row.get('stock', 0) or 0),
            'inventory_id': row.get('id'),
        })
    return {
        'id': product.get('id'),
        'name': product.get('name'),
        'sku': product.get('sku', ''),
        'category_id': product.get('category_id', ''),
        'category_name': cat.get('name', ''),
        'category_code': (product.get('qr_category_code') or cat.get('code') or ''),
        'product_code': product.get('qr_product_code') or '',
        'qr_serial': product.get('qr_serial') or '',
        'qr_uid': product_qr_uid(product),
        'qr_code': product.get('qr_code') or '',
        'qr_generated': bool((product.get('qr_code') or '').strip()),
        'qr_generated_at': product.get('qr_generated_at') or product.get('updated_at') or product.get('created_at') or '',
        'status': product.get('status', 'available'),
        'total_stock': total_stock,
        'price_min': min(prices) if prices else 0,
        'price_max': max(prices) if prices else 0,
        'stores': store_names,
        'store_details': store_details,
        'variants': product.get('variants') or [],
        'updated_at': product.get('updated_at') or product.get('created_at') or '',
    }


@app.route('/api/admin/qr-codes', methods=['GET'])
@admin_required
def api_admin_qr_codes():
    """Return unique QR units (slim payload). Optional pagination + filters for speed."""
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can manage QR codes'}), 403
    sync_flag = (request.args.get('sync') or '').strip().lower() in ('1', 'true', 'yes')
    backfilled = backfill_all_product_qrs() if sync_flag else 0
    store_id = (request.args.get('store_id') or '').strip()
    category_id = (request.args.get('category_id') or '').strip()
    product_id = (request.args.get('product_id') or '').strip()
    status = (request.args.get('status') or 'active').strip() or 'active'
    include_catalog = (request.args.get('catalog') or '').strip().lower() in ('1', 'true', 'yes')
    try:
        limit = min(500, max(1, int(request.args.get('limit') or 250)))
        offset = max(0, int(request.args.get('offset') or 0))
    except (TypeError, ValueError):
        limit, offset = 250, 0

    categories_by_id = {c['id']: c for c in db_find('categories')}
    stores_by_id = {s['id']: s for s in db_find('stores')}
    # Only load products we need for unit enrichment (not full catalog by default)
    products_by_id = {p['id']: p for p in db_find('products')}

    query = {}
    if status == 'active':
        query['status'] = {'$in': ['pending', 'in_stock']}
    elif status != 'all':
        query['status'] = status
    if store_id:
        query['store_id'] = store_id
    if product_id:
        query['product_id'] = product_id

    units = db_find('qr_units', query, sort=[('created_at', -1)])
    items = []
    for unit in units:
        product = products_by_id.get(unit.get('product_id'))
        if not product:
            continue
        if category_id and product.get('category_id') != category_id:
            continue
        cat = categories_by_id.get(product.get('category_id')) or {}
        store = stores_by_id.get(unit.get('store_id')) or {}
        variant_label = ''
        for v in product.get('variants') or []:
            if v.get('id') == unit.get('variant_id'):
                variant_label = v.get('label') or ''
                break
        serial = (unit.get('unit_serial') or (unit.get('code') or '')[-3:]).upper()
        items.append({
            'id': unit.get('id'),
            'unit_id': unit.get('id'),
            'product_id': product.get('id'),
            'name': product.get('name'),
            'sku': product.get('sku', ''),
            'category_id': product.get('category_id', ''),
            'category_name': cat.get('name', ''),
            'variant_id': unit.get('variant_id'),
            'variant_label': variant_label or unit.get('variant_id') or '—',
            'store_id': unit.get('store_id'),
            'store_name': store.get('name', ''),
            'stock': 1,
            'price': float(unit.get('price') or 0),
            'qr_code': unit.get('code') or '',
            'qr_serial': serial,
            'unit_serial': serial,
            'qr_uid': serial,
            'qr_generated': True,
            'qr_generated_at': unit.get('created_at') or '',
            'status': unit.get('status') or 'pending',
            'unit_status': unit.get('status') or 'pending',
        })

    total = len(items)
    page = items[offset:offset + limit]
    payload = {
        'items': page,
        'units': page,
        'backfilled': backfilled,
        'unit_count': total,
        'limit': limit,
        'offset': offset,
        'has_more': (offset + limit) < total,
    }
    if include_catalog:
        inventory_rows = db_find('inventory')
        catalog = [
            _enrich_qr_product_row(p, categories_by_id, stores_by_id, inventory_rows)
            for p in products_by_id.values()
        ]
        catalog.sort(key=lambda r: (r.get('name') or '').lower())
        payload['products'] = catalog
    return jsonify(payload)


@app.route('/api/admin/qr-codes/unit/<unit_id>/image')
@admin_required
def api_admin_qr_unit_image(unit_id):
    unit = db_find_one('qr_units', {'id': unit_id})
    if not unit:
        return jsonify({'error': 'QR unit not found'}), 404
    code = (unit.get('code') or '').strip()
    if not code:
        return jsonify({'error': 'QR code missing'}), 404
    png = _qr_png_bytes(code, box_size=10, border=2)
    resp = send_file(png, mimetype='image/png')
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/api/admin/products/<product_id>/qr-units', methods=['GET'])
@admin_required
@page_required('products')
def api_admin_product_qr_units(product_id):
    """List unique QR units for one product (all stores for Super; own store for Store Admin)."""
    product = db_find_one('products', {'id': product_id})
    if not product:
        return jsonify({'error': 'Product not found'}), 404
    store_id = (request.args.get('store_id') or '').strip()
    if not admin_is_super():
        locked = (session.get('admin_store_id') or '').strip()
        if not locked:
            return jsonify({'error': 'Your account is not assigned to a store'}), 403
        store_id = locked
    query = {'product_id': product_id}
    if store_id:
        query['store_id'] = store_id
    status = (request.args.get('status') or 'active').strip() or 'active'
    if status == 'active':
        query['status'] = {'$in': ['pending', 'in_stock']}
    elif status != 'all':
        query['status'] = status
    categories_by_id = {c['id']: c for c in db_find('categories')}
    stores_by_id = {s['id']: s for s in db_find('stores')}
    cat = categories_by_id.get(product.get('category_id')) or {}
    units = db_find('qr_units', query, sort=[('created_at', -1)])
    items = []
    for unit in units:
        store = stores_by_id.get(unit.get('store_id')) or {}
        variant_label = ''
        for v in product.get('variants') or []:
            if v.get('id') == unit.get('variant_id'):
                variant_label = v.get('label') or ''
                break
        serial = (unit.get('unit_serial') or (unit.get('code') or '')[-3:]).upper()
        items.append({
            'id': unit.get('id'),
            'unit_id': unit.get('id'),
            'product_id': product.get('id'),
            'name': product.get('name'),
            'sku': product.get('sku', ''),
            'category_id': product.get('category_id', ''),
            'category_name': cat.get('name', ''),
            'variant_id': unit.get('variant_id'),
            'variant_label': variant_label or unit.get('variant_id') or '—',
            'store_id': unit.get('store_id'),
            'store_name': store.get('name', ''),
            'stock': 1,
            'price': float(unit.get('price') or 0),
            'qr_code': unit.get('code') or '',
            'qr_serial': serial,
            'unit_serial': serial,
            'qr_uid': serial,
            'qr_generated': True,
            'qr_generated_at': unit.get('created_at') or '',
            'status': unit.get('status') or 'pending',
            'unit_status': unit.get('status') or 'pending',
        })
    return jsonify({
        'product': {
            'id': product.get('id'),
            'name': product.get('name'),
            'sku': product.get('sku', ''),
        },
        'items': items,
        'units': items,
        'can_delete': admin_can_manage_qr_units(),
    })


@app.route('/api/admin/qr-codes/unit/<unit_id>', methods=['DELETE'])
@admin_required
def api_admin_qr_unit_delete(unit_id):
    """Delete one unique QR unit. Store Admin + Super Admin only. Syncs inventory if in_stock."""
    if not admin_can_manage_qr_units():
        return jsonify({'error': 'Only Super Admin or Store Admin can delete QR units'}), 403
    unit = db_find_one('qr_units', {'id': unit_id})
    if not unit:
        return jsonify({'error': 'QR unit not found'}), 404
    denied = assert_store_access(unit.get('store_id'))
    if denied:
        return denied
    status = (unit.get('status') or '').strip()
    if status == 'sold':
        return jsonify({'error': 'Sold QR units cannot be deleted'}), 409

    stock_before = None
    stock_after = None
    if status == 'in_stock':
        inv = db_find_one('inventory', {
            'store_id': unit.get('store_id'),
            'product_id': unit.get('product_id'),
            'variant_id': unit.get('variant_id') or 'v1',
        })
        if inv:
            stock_before = int(inv.get('stock') or 0)
            if stock_before > 0:
                updated = db_increment('inventory', {'id': inv['id']}, 'stock', -1)
                stock_after = int((updated or {}).get('stock') or 0)
            else:
                stock_after = stock_before

    deleted = db_delete('qr_units', {'id': unit_id})
    if not deleted:
        return jsonify({'error': 'Could not delete QR unit'}), 500

    _invalidate_ref_cache('products_by_id')
    _badges_cache.clear()
    product = db_find_one('products', {'id': unit.get('product_id')}) or {}
    log_activity(
        'inventory',
        f"QR unit deleted · {product.get('name') or unit.get('product_id')} · "
        f"{unit.get('unit_serial') or unit.get('code')} · was {status or 'pending'}",
        {
            'unit_id': unit_id,
            'product_id': unit.get('product_id'),
            'store_id': unit.get('store_id'),
            'status_was': status,
            'stock_before': stock_before,
            'stock_after': stock_after,
        },
    )
    return jsonify({
        'ok': True,
        'deleted_id': unit_id,
        'status_was': status,
        'stock_before': stock_before,
        'stock_after': stock_after,
    })


@app.route('/api/admin/qr-codes/<product_id>/image')
@admin_required
def api_admin_qr_image(product_id):
    """Serve QR PNG — prefers a live unit code, falls back to product template."""
    unit = db_find_one('qr_units', {'id': product_id})
    if unit and unit.get('code'):
        png = _qr_png_bytes(unit['code'], box_size=10, border=2)
        resp = send_file(png, mimetype='image/png')
        resp.headers['Cache-Control'] = 'no-store'
        return resp
    product = db_find_one('products', {'id': product_id})
    if not product:
        return jsonify({'error': 'Product not found'}), 404
    units = db_find('qr_units', {'product_id': product_id, 'status': 'in_stock'}, limit=1)
    if units and units[0].get('code'):
        code = units[0]['code']
    else:
        code = (product.get('qr_code') or '').strip()
        if not code:
            category = db_find_one('categories', {'id': product.get('category_id')})
            if category:
                apply_qr_to_product(product, category=category)
                product = db_find_one('products', {'id': product_id})
                code = (product.get('qr_code') or '').strip()
    if not code:
        return jsonify({'error': 'QR not generated for this product'}), 404
    png = _qr_png_bytes(code, box_size=10, border=2)
    resp = send_file(png, mimetype='image/png')
    resp.headers['Cache-Control'] = 'no-store'
    return resp


@app.route('/api/admin/qr-codes/lookup', methods=['GET'])
@admin_required
def api_admin_qr_lookup():
    code = request.args.get('code') or request.args.get('q') or ''
    unit = find_qr_unit_by_code(code)
    product = find_product_by_qr(code)
    if not product:
        return jsonify({'error': 'QR code not found'}), 404
    store_id = resolve_store_scope(request.args.get('store_id'))
    if unit and store_id and unit.get('store_id') and unit.get('store_id') != store_id:
        return jsonify({'error': 'This QR belongs to a different store'}), 409
    if unit and unit.get('status') == 'sold':
        return jsonify({'error': 'This unit QR was already sold'}), 409
    categories_by_id = {c['id']: c for c in db_find('categories')}
    stores_by_id = {s['id']: s for s in db_find('stores')}
    inv_query = {'product_id': product['id']}
    if store_id:
        inv_query['store_id'] = store_id
    elif unit:
        inv_query['store_id'] = unit.get('store_id')
    inventory_rows = db_find('inventory', inv_query)
    enriched = _enrich_qr_product_row(product, categories_by_id, stores_by_id, inventory_rows)
    preferred_variant = (unit or {}).get('variant_id')
    preferred = None
    for detail in enriched.get('store_details') or []:
        if preferred_variant and detail.get('variant_id') != preferred_variant:
            continue
        if store_id and detail.get('store_id') != store_id:
            continue
        preferred = detail
        if detail.get('stock', 0) > 0:
            break
    if not preferred and enriched.get('store_details'):
        preferred = enriched['store_details'][0]
    enriched['preferred_variant_id'] = preferred_variant or (preferred or {}).get('variant_id') or (
        (product.get('variants') or [{}])[0].get('id')
    )
    enriched['preferred_price'] = float((unit or {}).get('price') or (preferred or {}).get('price') or 0)
    enriched['preferred_stock'] = int((preferred or {}).get('stock') or 0)
    if unit:
        enriched['unit_id'] = unit.get('id')
        enriched['qr_code'] = unit.get('code')
        enriched['qr_serial'] = unit.get('unit_serial')
        enriched['qr_uid'] = unit.get('unit_serial')
        enriched['unit_status'] = unit.get('status')
    return jsonify(enriched)


@app.route('/api/admin/qr-codes/generate', methods=['POST'])
@admin_required
def api_admin_qr_generate():
    """Create or update a product QR and sync inventory for the chosen store."""
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can generate QR codes'}), 403
    data = parse_json()
    category_id = (data.get('category_id') or '').strip()
    category = db_find_one('categories', {'id': category_id})
    if not category:
        return jsonify({'error': 'Select a valid category'}), 400
    store_id = (data.get('store_id') or '').strip()
    store = db_find_one('stores', {'id': store_id})
    if not store:
        return jsonify({'error': 'Select a valid store'}), 400
    try:
        price = float(data.get('price', 0) or 0)
        stock = int(data.get('stock', 0) or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'Price and stock must be valid numbers'}), 400
    if price < 0 or stock < 0:
        return jsonify({'error': 'Price and stock cannot be negative'}), 400
    if stock > 1000000:
        return jsonify({'error': 'Stock is too large'}), 400

    product_id = (data.get('product_id') or '').strip()
    product_name = (data.get('product_name') or '').strip()
    variant_label = (data.get('variant_label') or '1 kg').strip() or '1 kg'
    variant_unit = (data.get('variant_unit') or variant_label).strip()
    regenerate = bool(data.get('regenerate', False))
    created = False

    if product_id:
        product = db_find_one('products', {'id': product_id})
        if not product:
            return jsonify({'error': 'Product not found'}), 404
        if product_name:
            product['name'] = product_name
    else:
        if not product_name:
            return jsonify({'error': 'Product name is required'}), 400
        stores = db_find('stores')
        product = {
            'id': new_id('p'),
            'name': product_name,
            'description': data.get('description', ''),
            'sku': data.get('sku') or f'FAM-{uuid.uuid4().hex[:6].upper()}',
            'category_id': category_id,
            'images': [],
            'status': 'available',
            'gst_percent': float(data.get('gst_percent', 0) or 0),
            'expiry_info': '',
            'nutritional_info': '',
            'parameters': normalize_parameters(category.get('parameters')),
            'seo_title': product_name,
            'seo_description': '',
            'featured': False,
            'bestseller': False,
            'inventory_model': 'variant',
            'variants': [{
                'id': 'v1',
                'label': variant_label,
                'sku_suffix': _sanitize_code_chars(variant_unit, 4) or 'UNIT',
                'unit': variant_unit,
            }],
            'store_availability': [s['id'] for s in stores] if data.get('all_stores') else [store_id],
            'created_at': now_iso(),
            'updated_at': now_iso(),
        }
        db_insert('products', product)
        created = True

    # Keep category in sync
    if product_id:
        product['category_id'] = category_id

    try:
        apply_qr_to_product(product, category=category, regenerate=regenerate)
    except ValueError as exc:
        return jsonify({'error': public_error(exc, 'Could not generate QR')}), 400

    # Ensure selected store is available
    availability = list(product.get('store_availability') or [])
    if store_id not in availability:
        availability.append(store_id)
    product['store_availability'] = availability
    product['updated_at'] = now_iso()

    variants = product.get('variants') or []
    if not variants:
        variants = [{'id': 'v1', 'label': variant_label, 'sku_suffix': 'UNIT', 'unit': variant_unit}]
        product['variants'] = variants
    variant_id = (data.get('variant_id') or '').strip() or variants[0].get('id')
    if not any(v.get('id') == variant_id for v in variants):
        variant_id = variants[0].get('id')

    db_update('products', {'id': product['id']}, {
        'name': product.get('name'),
        'category_id': category_id,
        'qr_code': product.get('qr_code'),
        'qr_category_code': product.get('qr_category_code'),
        'qr_product_code': product.get('qr_product_code'),
        'qr_serial': product.get('qr_serial'),
        'qr_generated_at': product.get('qr_generated_at') or now_iso(),
        'store_availability': availability,
        'variants': variants,
        'status': product.get('status', 'available'),
        'updated_at': product['updated_at'],
    })
    product = db_find_one('products', {'id': product['id']})
    # Ensure inventory row exists for price metadata — do NOT increase stock on generate
    _ensure_product_inventory_rows(product, default_price=price, default_stock=0)
    inv = db_find_one('inventory', {
        'store_id': store_id,
        'product_id': product['id'],
        'variant_id': variant_id,
    })
    if not inv:
        inv = {
            'id': new_id('inv_'),
            'store_id': store_id,
            'product_id': product['id'],
            'variant_id': variant_id,
            'price': price if price > 0 else 0,
            'stock': 0,
            'updated_at': now_iso(),
        }
        db_insert('inventory', inv)
    elif price > 0:
        db_update('inventory', {'id': inv['id']}, {
            'price': price,
            'updated_at': now_iso(),
        })
        inv = db_find_one('inventory', {'id': inv['id']})

    # Generate unique QRs as pending — inventory stock only increases after punch
    qty = stock if stock > 0 else int(data.get('qty') or 0)
    if qty < 1:
        return jsonify({'error': 'Enter how many unique QR codes to generate (at least 1)'}), 400
    new_units = create_qr_units(
        product, store_id, variant_id, qty,
        price=float((inv or {}).get('price') or price or 0),
        status='pending',
    )
    created_unit_ids = [u.get('id') for u in new_units if u.get('id')]
    created_unit_codes = [u.get('code') for u in new_units if u.get('code')]
    created_unit_serials = [u.get('unit_serial') for u in new_units if u.get('unit_serial')]

    _invalidate_ref_cache('products_by_id')
    _badges_cache.clear()
    log_activity(
        'inventory',
        f"QR generated (pending punch) · {product.get('name')} · +{len(created_unit_ids)} · {store.get('name')}",
        {
            'product_id': product['id'],
            'store_id': store_id,
            'units_created': len(created_unit_ids),
            'status': 'pending',
        },
    )
    categories_by_id = {c['id']: c for c in db_find('categories')}
    stores_by_id = {s['id']: s for s in db_find('stores')}
    inventory_rows = db_find('inventory', {'product_id': product['id']})
    enriched = _enrich_qr_product_row(product, categories_by_id, stores_by_id, inventory_rows)
    enriched['units_created'] = len(created_unit_ids)
    return jsonify({
        'ok': True,
        'created': created,
        'product': enriched,
        'units_created': len(created_unit_ids),
        'created_unit_ids': created_unit_ids,
        'created_unit_codes': created_unit_codes,
        'created_unit_serials': created_unit_serials,
        'inventory_unchanged': True,
    }), (201 if created else 200)


def _qr_png_bytes(code, box_size=8, border=2):
    """Render a QR payload to PNG bytes for PDF embedding."""
    import qrcode
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M,
                       box_size=box_size, border=border)
    qr.add_data(code)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#1E3A22', back_color='white')
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


def _build_qr_pdf(products):
    """One A4 page per QR unit (or legacy product row): QR image + code + name."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak, KeepTogether
    from reportlab.lib.enums import TA_CENTER

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=48, leftMargin=48, topMargin=56, bottomMargin=48,
    )
    styles = getSampleStyleSheet()
    brand = ParagraphStyle(
        'FamBrand', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=18, textColor='#1E3A22',
        alignment=TA_CENTER, spaceAfter=8,
    )
    code_style = ParagraphStyle(
        'FamCode', parent=styles['Normal'],
        fontName='Courier-Bold', fontSize=14, textColor='#1E3A22',
        alignment=TA_CENTER, spaceBefore=14, spaceAfter=8,
    )
    meta = ParagraphStyle(
        'FamMeta', parent=styles['Normal'],
        fontName='Helvetica', fontSize=12, textColor='#55594F',
        alignment=TA_CENTER, leading=16,
    )
    story = []
    for idx, product in enumerate(products):
        code = (product.get('qr_code') or '').strip().upper()
        if not code:
            continue
        png = _qr_png_bytes(code, box_size=10, border=2)
        img = Image(png, width=3.2 * inch, height=3.2 * inch)
        img.hAlign = 'CENTER'
        unique = (
            product.get('unit_serial')
            or product.get('qr_serial')
            or product.get('qr_uid')
            or (code[-3:] if len(code) >= 3 else '')
        )
        block = [
            Paragraph('FISH AND MEAT', brand),
            Spacer(1, 18),
            img,
            Paragraph(code, code_style),
            Paragraph(
                f"<b>{product.get('name') or 'Product'}</b>",
                meta,
            ),
            Spacer(1, 6),
            Paragraph(
                f"{product.get('category_name') or ''}"
                f"{(' · ' + product.get('sku')) if product.get('sku') else ''}"
                f"{(' · Unique ' + str(unique)) if unique else ''}"
                f"{(' · ' + product.get('store_name')) if product.get('store_name') else ''}",
                meta,
            ),
        ]
        story.append(KeepTogether(block))
        if idx < len(products) - 1:
            story.append(PageBreak())
    if not story:
        raise ValueError('No QR codes to print')
    doc.build(story)
    buf.seek(0)
    return buf


def _units_for_qr_print(unit_ids=None, product_ids=None, codes=None):
    """Resolve print rows from unit ids, product ids, or raw QR codes."""
    categories_by_id = {c['id']: c for c in db_find('categories')}
    stores_by_id = {s['id']: s for s in db_find('stores')}
    products_by_id = {p['id']: p for p in db_find('products')}
    rows = []
    seen = set()

    def append_unit(unit):
        if not unit or not unit.get('code'):
            return
        uid = unit.get('id') or unit.get('code')
        if uid in seen:
            return
        seen.add(uid)
        product = products_by_id.get(unit.get('product_id')) or {}
        cat = categories_by_id.get(product.get('category_id')) or {}
        store = stores_by_id.get(unit.get('store_id')) or {}
        rows.append({
            'id': unit.get('id'),
            'name': product.get('name') or 'Product',
            'sku': product.get('sku') or '',
            'category_name': cat.get('name') or '',
            'store_name': store.get('name') or '',
            'qr_code': unit.get('code'),
            'unit_serial': unit.get('unit_serial') or '',
            'qr_uid': unit.get('unit_serial') or '',
            'qr_serial': unit.get('unit_serial') or '',
        })

    for uid in unit_ids or []:
        unit = db_find_one('qr_units', {'id': str(uid).strip()})
        if unit:
            append_unit(unit)

    for code in codes or []:
        unit = find_qr_unit_by_code(code)
        if unit:
            append_unit(unit)

    for pid in product_ids or []:
        # Prefer in-stock units for this product; fall back to template once
        units = db_find('qr_units', {
            'product_id': str(pid).strip(),
            'status': 'in_stock',
        }, sort=[('created_at', 1)])
        if units:
            for unit in units:
                append_unit(unit)
            continue
        product = products_by_id.get(str(pid).strip())
        if product and product.get('qr_code'):
            key = 'p:' + product['id']
            if key in seen:
                continue
            seen.add(key)
            cat = categories_by_id.get(product.get('category_id')) or {}
            rows.append({
                'id': product['id'],
                'name': product.get('name'),
                'sku': product.get('sku') or '',
                'category_name': cat.get('name') or '',
                'store_name': '',
                'qr_code': product.get('qr_code'),
                'unit_serial': '',
                'qr_uid': product_qr_uid(product),
                'qr_serial': product.get('qr_serial') or '',
            })
    return rows


def _products_for_qr_print(product_ids):
    """Back-compat wrapper — expands products into their in-stock unit QRs."""
    return _units_for_qr_print(product_ids=product_ids)


@app.route('/api/admin/qr-codes/print', methods=['POST'])
@admin_required
def api_admin_qr_print():
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can print QR codes'}), 403
    data = parse_json()
    unit_ids = data.get('unit_ids') or []
    product_ids = data.get('product_ids') or []
    if isinstance(unit_ids, str):
        unit_ids = [unit_ids]
    if isinstance(product_ids, str):
        product_ids = [product_ids]
    unit_ids = [str(x).strip() for x in unit_ids if str(x).strip()]
    product_ids = [str(x).strip() for x in product_ids if str(x).strip()]
    if not unit_ids and not product_ids:
        return jsonify({'error': 'Select at least one QR unit'}), 400
    if len(unit_ids) + len(product_ids) > 500:
        return jsonify({'error': 'Select up to 500 QR units at a time'}), 400
    rows = _units_for_qr_print(unit_ids=unit_ids, product_ids=product_ids)
    if not rows:
        return jsonify({'error': 'None of the selected items have a QR code yet'}), 400
    try:
        pdf = _build_qr_pdf(rows)
    except Exception as exc:  # noqa: BLE001
        return jsonify({'error': 'Could not build PDF'}), 500
    filename = f'fam_qr_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(pdf, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/api/mobile/qr-print', methods=['POST'])
@mobile_auth_required
def api_mobile_qr_print():
    """Download a multi-page QR PDF for punched / selected units."""
    data = parse_json()
    unit_ids = data.get('unit_ids') or []
    product_ids = data.get('product_ids') or []
    codes = data.get('qr_codes') or []
    if isinstance(unit_ids, str):
        unit_ids = [unit_ids]
    if isinstance(product_ids, str):
        product_ids = [product_ids]
    unit_ids = [str(x).strip() for x in unit_ids if str(x).strip()]
    product_ids = [str(x).strip() for x in product_ids if str(x).strip()]
    if not unit_ids and not product_ids and codes:
        for code in codes:
            unit = find_qr_unit_by_code(code)
            if unit:
                unit_ids.append(unit['id'])
            else:
                product = find_product_by_qr(code)
                if product and product.get('id'):
                    product_ids.append(product['id'])
    seen = set()
    ordered_units = []
    for uid in unit_ids:
        if uid not in seen:
            seen.add(uid)
            ordered_units.append(uid)
    seen_p = set()
    ordered_products = []
    for pid in product_ids:
        if pid not in seen_p:
            seen_p.add(pid)
            ordered_products.append(pid)
    if not ordered_units and not ordered_products:
        return jsonify({'error': 'No QR units to print'}), 400
    rows = _units_for_qr_print(unit_ids=ordered_units, product_ids=ordered_products, codes=codes)
    if not rows:
        return jsonify({'error': 'No QR codes found'}), 400
    try:
        pdf = _build_qr_pdf(rows)
    except Exception as exc:  # noqa: BLE001
        return jsonify({'error': 'Could not build PDF'}), 500
    filename = f'fam_punch_qr_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(pdf, as_attachment=True, download_name=filename, mimetype='application/pdf')


# --- Mobile APK APIs (Wi‑Fi + Bearer token) ---

@app.route('/api/mobile', methods=['OPTIONS'])
@app.route('/api/mobile/<path:_rest>', methods=['OPTIONS'])
def api_mobile_options(_rest=None):
    return ('', 204)


@app.route('/api/mobile/login', methods=['POST'])
def api_mobile_login():
    data = parse_json()
    member, error = _authenticate_staff_credentials(data.get('username'), data.get('password'))
    if error:
        return jsonify({'error': error}), 401
    token = issue_mobile_token(member)
    log_activity('system', f"{member.get('name')} signed in via mobile APK ({member.get('role')})")
    return jsonify({
        'ok': True,
        'token': token,
        'admin': {
            'id': member.get('id'),
            'name': member.get('name') or member.get('username') or 'Staff',
            'username': member.get('username') or '',
            'role': member.get('role') or ROLE_STORE,
            'store_id': member.get('store_id') or '',
        },
        'api_base': '/api/mobile',
    })


@app.route('/api/mobile/me', methods=['GET'])
@mobile_auth_required
def api_mobile_me():
    staff = request.mobile_staff
    return jsonify({'ok': True, 'admin': staff})


@app.route('/api/mobile/stores', methods=['GET'])
@mobile_auth_required
def api_mobile_stores():
    staff = request.mobile_staff
    stores = [s for s in db_find('stores', sort=[('name', 1)]) if s.get('status') == 'active']
    if staff.get('role') != ROLE_SUPER:
        locked = (staff.get('store_id') or '').strip()
        stores = [s for s in stores if s.get('id') == locked] if locked else []
    return jsonify(stores)


@app.route('/api/mobile/dashboard', methods=['GET'])
@mobile_auth_required
def api_mobile_dashboard():
    """Sales cards for the signed-in staff (same metrics flavor as admin dashboard)."""
    staff = request.mobile_staff
    store_id = ''
    if staff.get('role') != ROLE_SUPER:
        store_id = (staff.get('store_id') or '').strip()
    else:
        store_id = (request.args.get('store_id') or '').strip()

    today = datetime.utcnow().strftime('%Y-%m-%d')
    tomorrow = (datetime.utcnow() + timedelta(days=1)).strftime('%Y-%m-%d')
    query = {}
    if store_id:
        query['store_id'] = store_id
    # Only today's window + open statuses — avoid loading 500 full order docs every home open
    today_query = dict(query)
    today_query['created_at'] = {'$gte': today, '$lt': tomorrow}
    today_orders = db_find(
        'orders',
        today_query,
        projection={'total': 1, 'status': 1, 'created_at': 1},
        limit=500,
    )
    today_sales = sum(float(o.get('total', 0) or 0) for o in today_orders)
    delivered_today = sum(
        1 for o in today_orders if normalize_status(o.get('status')) == 'delivered'
    )
    open_query = dict(query)
    open_query['status'] = {'$in': list(_OPEN_ORDER_STATUSES)}
    open_count = db_count('orders', open_query)
    inv_query = {'store_id': store_id} if store_id else {}
    inventory = db_find('inventory', inv_query, projection={'stock': 1, 'id': 1})
    threshold = int(get_settings().get('low_stock_threshold', 10))
    low_stock = 0
    total_stock = 0
    for r in inventory:
        stock = int(r.get('stock', 0) or 0)
        total_stock += stock
        if stock <= threshold:
            low_stock += 1
    return jsonify({
        'ok': True,
        'admin': staff,
        'cards': {
            'today_sales': round(today_sales, 2),
            'today_orders': len(today_orders),
            'open_orders': open_count,
            'delivered_today': delivered_today,
            'low_stock': low_stock,
            'total_stock_units': total_stock,
        },
    })


@app.route('/api/mobile/qr-lookup', methods=['GET'])
@mobile_auth_required
def api_mobile_qr_lookup():
    """Fast QR resolve for punch — only pending unit QRs are punchable."""
    code = request.args.get('code') or ''
    purpose = (request.args.get('purpose') or 'punch').strip().lower()
    unit = find_qr_unit_by_code(code)
    if not unit:
        return jsonify({'error': 'Scan a unique unit QR (not a catalog/product template)'}), 404
    product = db_find_one('products', {'id': unit.get('product_id')})
    if not product:
        return jsonify({'error': 'QR code not found'}), 404
    store_id = (request.args.get('store_id') or '').strip()
    staff = request.mobile_staff
    if staff.get('role') != ROLE_SUPER:
        store_id = (staff.get('store_id') or store_id or '').strip()
    if store_id and unit.get('store_id') and unit.get('store_id') != store_id:
        return jsonify({'error': 'This QR belongs to a different store'}), 409

    unit_status = (unit.get('status') or '').strip()
    if purpose == 'punch':
        if unit_status == 'in_stock':
            return jsonify({'error': 'Already in inventory — cannot punch again'}), 409
        if unit_status == 'sold':
            return jsonify({'error': 'This QR was already sold'}), 409
        if unit_status == 'void':
            return jsonify({'error': 'This QR is void'}), 409
        if unit_status not in ('pending', ''):
            return jsonify({'error': f'QR status is {unit_status} — not punchable'}), 409
    elif purpose == 'sale':
        if unit_status == 'pending' or unit_status == '':
            return jsonify({'error': 'Punch this QR into stock before billing'}), 409
        if unit_status == 'sold':
            return jsonify({'error': 'This QR was already sold'}), 409
        if unit_status == 'void':
            return jsonify({'error': 'This QR is void'}), 409
        if unit_status != 'in_stock':
            return jsonify({'error': f'QR status is {unit_status} — not sellable'}), 409

    categories_by_id = {c['id']: c for c in db_find('categories')}
    stores_by_id = {s['id']: s for s in db_find('stores')}
    inv_query = {'product_id': product['id'], 'store_id': unit.get('store_id') or store_id}
    inventory_rows = db_find('inventory', inv_query) if inv_query.get('store_id') else db_find('inventory', {'product_id': product['id']})
    enriched = _enrich_qr_product_row(product, categories_by_id, stores_by_id, inventory_rows)
    preferred = None
    for detail in enriched.get('store_details') or []:
        if detail.get('variant_id') == unit.get('variant_id'):
            preferred = detail
            break
    if not preferred and enriched.get('store_details'):
        preferred = enriched['store_details'][0]
    enriched['id'] = product['id']
    enriched['preferred_variant_id'] = unit.get('variant_id') or (preferred or {}).get('variant_id')
    enriched['inventory_id'] = (preferred or {}).get('inventory_id')
    enriched['preferred_price'] = float(unit.get('price') or (preferred or {}).get('price') or 0)
    enriched['preferred_stock'] = int((preferred or {}).get('stock') or 0)
    enriched['unit_id'] = unit.get('id')
    enriched['qr_code'] = unit.get('code')
    enriched['qr_serial'] = unit.get('unit_serial')
    enriched['qr_uid'] = unit.get('unit_serial')
    enriched['unit_status'] = unit_status or 'pending'
    enriched['store_id'] = unit.get('store_id')
    enriched['punchable'] = unit_status in ('pending', '')
    enriched['sellable'] = unit_status == 'in_stock'
    enriched['product_id'] = product.get('id')
    enriched['variant_id'] = unit.get('variant_id') or enriched.get('preferred_variant_id')
    return jsonify(enriched)


@app.route('/api/mobile/punch', methods=['POST'])
@mobile_auth_required
def api_mobile_punch():
    """Punch ONE pending unit QR into inventory (qty 1). Rejects already in-stock units."""
    staff = request.mobile_staff
    data = parse_json()
    store_id = (data.get('store_id') or '').strip()
    if staff.get('role') != ROLE_SUPER:
        locked = (staff.get('store_id') or '').strip()
        if not locked:
            return jsonify({'error': 'Your account is not assigned to a store'}), 403
        store_id = locked
    store = db_find_one('stores', {'id': store_id})
    if not store:
        return jsonify({'error': 'Select a valid store'}), 400
    items = data.get('items') or []
    if not items:
        return jsonify({'error': 'Scan one QR to punch'}), 400
    if len(items) > 1:
        return jsonify({'error': 'Only one product can be punched at a time'}), 400

    raw = items[0]
    code = raw.get('qr_code') or raw.get('code') or ''
    unit_id = (raw.get('unit_id') or '').strip()
    unit = db_find_one('qr_units', {'id': unit_id}) if unit_id else find_qr_unit_by_code(code)
    if not unit:
        return jsonify({'error': 'Scan a unique pending unit QR'}), 404
    if unit.get('store_id') and unit.get('store_id') != store_id:
        return jsonify({'error': 'This QR belongs to a different store'}), 409
    status = (unit.get('status') or '').strip()
    if status == 'in_stock':
        return jsonify({'error': 'Already in inventory — cannot punch again'}), 409
    if status == 'sold':
        return jsonify({'error': 'This QR was already sold'}), 409
    if status not in ('pending', ''):
        return jsonify({'error': 'This QR is not awaiting punch'}), 409

    product = db_find_one('products', {'id': unit.get('product_id')})
    if not product:
        return jsonify({'error': 'Product not found for this QR'}), 404
    variant_id = unit.get('variant_id') or 'v1'
    availability = list(product.get('store_availability') or [])
    if store_id not in availability:
        availability.append(store_id)
        db_update('products', {'id': product['id']}, {
            'store_availability': availability,
            'updated_at': now_iso(),
        })
    _ensure_product_inventory_rows(product, default_price=float(unit.get('price') or 0), default_stock=0)
    inv = db_find_one('inventory', {
        'store_id': store_id,
        'product_id': product['id'],
        'variant_id': variant_id,
    })
    if not inv:
        inv = {
            'id': new_id('inv_'),
            'store_id': store_id,
            'product_id': product['id'],
            'variant_id': variant_id,
            'price': float(unit.get('price') or 0),
            'stock': 0,
            'updated_at': now_iso(),
        }
        db_insert('inventory', inv)

    before_stock = int(inv.get('stock') or 0)
    # Claim pending unit (prevents double-punch). Empty status treated as pending.
    claim_status = status if status else 'pending'
    claim_updates = {
        'status': 'in_stock',
        'punched_at': now_iso(),
        'punched_by': staff.get('id') or '',
        'updated_at': now_iso(),
    }
    claimed = db_update('qr_units', {'id': unit['id'], 'status': claim_status}, claim_updates)
    if not claimed and not status:
        # Legacy rows with missing status field
        current = db_find_one('qr_units', {'id': unit['id']})
        if current and (current.get('status') or '') in ('', 'pending'):
            claimed = db_update('qr_units', {'id': unit['id']}, claim_updates)
    if not claimed:
        return jsonify({'error': 'Already in inventory — cannot punch again'}), 409

    updated_row = db_increment('inventory', {'id': inv['id']}, 'stock', 1)
    if unit.get('price') is not None and float(unit.get('price') or 0) > 0:
        db_update('inventory', {'id': inv['id']}, {
            'price': float(unit.get('price') or 0),
            'updated_at': now_iso(),
        })
        updated_row = db_find_one('inventory', {'id': inv['id']})

    _invalidate_ref_cache('products_by_id')
    _badges_cache.clear()
    updated = [{
        'qr_code': unit.get('code'),
        'unit_id': unit.get('id'),
        'unit_serial': unit.get('unit_serial'),
        'product_id': product['id'],
        'product_name': product.get('name'),
        'variant_id': variant_id,
        'qty_added': 1,
        'stock_before': before_stock,
        'stock': int((updated_row or {}).get('stock', 0) or 0),
        'price': float((updated_row or {}).get('price', 0) or 0),
        'unit_serials': [unit.get('unit_serial')],
        'unit_ids': [unit.get('id')],
    }]
    log_activity(
        'inventory',
        f"Mobile punch · {product.get('name')} · UID {unit.get('unit_serial')} · "
        f"{store.get('name')} · {staff.get('name')}",
        {'store_id': store_id, 'unit_id': unit.get('id'), 'staff_id': staff.get('id')},
    )
    return jsonify({
        'ok': True,
        'updated': updated,
        'created_unit_ids': [unit.get('id')],
    })


@app.route('/api/mobile/catalog', methods=['GET'])
@mobile_auth_required
def api_mobile_catalog():
    """Categories + products for Generate & Print on mobile."""
    categories = db_find('categories')
    products = db_find('products')
    inventory_rows = db_find('inventory')
    inv_by_product = {}
    for row in inventory_rows:
        inv_by_product.setdefault(row.get('product_id'), []).append(row)
    catalog_products = []
    for p in products:
        if (p.get('status') or 'available') == 'disabled':
            continue
        related = inv_by_product.get(p.get('id')) or []
        prices = [float(r.get('price') or 0) for r in related if float(r.get('price') or 0) > 0]
        catalog_products.append({
            'id': p.get('id'),
            'name': p.get('name'),
            'sku': p.get('sku') or '',
            'category_id': p.get('category_id') or '',
            'variants': p.get('variants') or [],
            'status': p.get('status') or 'available',
            'price_min': min(prices) if prices else 0,
            'store_inventory': [{
                'store_id': r.get('store_id'),
                'variant_id': r.get('variant_id'),
                'price': float(r.get('price') or 0),
                'stock': int(r.get('stock') or 0),
            } for r in related],
        })
    catalog_products.sort(key=lambda r: (r.get('name') or '').lower())
    categories.sort(key=lambda c: (c.get('name') or '').lower())
    return jsonify({
        'categories': [{
            'id': c.get('id'),
            'name': c.get('name'),
            'code': c.get('code') or '',
        } for c in categories],
        'products': catalog_products,
    })


@app.route('/api/mobile/qr-units', methods=['GET'])
@mobile_auth_required
def api_mobile_qr_units():
    """Pending + in-stock unique QR units for Print QR (newest first)."""
    staff = request.mobile_staff
    store_id = (request.args.get('store_id') or '').strip()
    if staff.get('role') != ROLE_SUPER:
        store_id = (staff.get('store_id') or store_id or '').strip()
    if not store_id:
        return jsonify({'error': 'Select a store'}), 400
    try:
        limit = min(300, max(1, int(request.args.get('limit', 150))))
    except (TypeError, ValueError):
        limit = 150
    units = db_find('qr_units', {
        'store_id': store_id,
        'status': {'$in': ['pending', 'in_stock']},
    }, sort=[('created_at', -1)], limit=limit)
    product_ids = list({u.get('product_id') for u in units if u.get('product_id')})
    products_by_id = {}
    if product_ids:
        for p in db_find('products', {'id': {'$in': product_ids}}, projection={
            'id': 1, 'name': 1, 'sku': 1, 'category_id': 1, 'variants': 1
        }):
            products_by_id[p['id']] = p
    category_ids = list({(p.get('category_id') or '') for p in products_by_id.values() if p.get('category_id')})
    categories_by_id = {}
    if category_ids:
        for c in db_find('categories', {'id': {'$in': category_ids}}, projection={'id': 1, 'name': 1}):
            categories_by_id[c['id']] = c
    items = []
    for unit in units:
        product = products_by_id.get(unit.get('product_id')) or {}
        cat = categories_by_id.get(product.get('category_id')) or {}
        serial = (unit.get('unit_serial') or (unit.get('code') or '')[-3:]).upper()
        variant_label = ''
        for v in product.get('variants') or []:
            if v.get('id') == unit.get('variant_id'):
                variant_label = v.get('label') or ''
                break
        items.append({
            'id': unit.get('id'),
            'unit_id': unit.get('id'),
            'product_id': product.get('id'),
            'name': product.get('name') or 'Product',
            'sku': product.get('sku') or '',
            'category_name': cat.get('name') or '',
            'variant_label': variant_label or '—',
            'qr_code': unit.get('code') or '',
            'unit_serial': serial,
            'qr_uid': serial,
            'status': unit.get('status') or 'pending',
            'created_at': unit.get('created_at') or '',
            'price': float(unit.get('price') or 0),
        })
    return jsonify({'items': items, 'store_id': store_id, 'count': len(items)})


@app.route('/api/mobile/qr-generate', methods=['POST'])
@mobile_auth_required
def api_mobile_qr_generate():
    """Generate N unique pending QR units — inventory stock only increases after punch."""
    staff = request.mobile_staff
    data = parse_json()
    store_id = (data.get('store_id') or '').strip()
    if staff.get('role') != ROLE_SUPER:
        locked = (staff.get('store_id') or '').strip()
        if not locked:
            return jsonify({'error': 'Your account is not assigned to a store'}), 403
        store_id = locked
    store = db_find_one('stores', {'id': store_id})
    if not store:
        return jsonify({'error': 'Select a valid store'}), 400
    category_id = (data.get('category_id') or '').strip()
    product_id = (data.get('product_id') or '').strip()
    category = db_find_one('categories', {'id': category_id})
    product = db_find_one('products', {'id': product_id})
    if not category or not product:
        return jsonify({'error': 'Select a valid category and product'}), 400
    if product.get('category_id') != category_id:
        return jsonify({'error': 'Product does not belong to that category'}), 400
    try:
        qty = int(data.get('qty') or data.get('stock') or 0)
        price = float(data.get('price') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'Quantity and price must be valid numbers'}), 400
    if qty < 1 or qty > 500:
        return jsonify({'error': 'Quantity must be between 1 and 500'}), 400

    ensure_product_template_codes(product, category=category)
    variants = product.get('variants') or []
    variant_id = (data.get('variant_id') or '').strip() or (variants[0].get('id') if variants else 'v1')
    availability = list(product.get('store_availability') or [])
    if store_id not in availability:
        availability.append(store_id)
        db_update('products', {'id': product['id']}, {
            'store_availability': availability,
            'updated_at': now_iso(),
        })
        product['store_availability'] = availability
    # Ensure price row exists — do NOT add stock on generate
    _ensure_product_inventory_rows(product, default_price=price, default_stock=0)
    inv = db_find_one('inventory', {
        'store_id': store_id,
        'product_id': product['id'],
        'variant_id': variant_id,
    })
    if not inv:
        inv = {
            'id': new_id('inv_'),
            'store_id': store_id,
            'product_id': product['id'],
            'variant_id': variant_id,
            'price': price if price > 0 else 0,
            'stock': 0,
            'updated_at': now_iso(),
        }
        db_insert('inventory', inv)
    elif price > 0:
        db_update('inventory', {'id': inv['id']}, {
            'price': price,
            'updated_at': now_iso(),
        })
        inv = db_find_one('inventory', {'id': inv['id']})

    product = db_find_one('products', {'id': product['id']})
    new_units = create_qr_units(
        product, store_id, variant_id, qty,
        price=float((inv or {}).get('price') or price or 0),
        status='pending',
    )
    created_unit_ids = [u.get('id') for u in new_units if u.get('id')]
    _invalidate_ref_cache('products_by_id')
    _badges_cache.clear()
    log_activity(
        'inventory',
        f"Mobile Generate & Print (pending) · +{len(created_unit_ids)} · "
        f"{product.get('name')} · {store.get('name')} · {staff.get('name')}",
        {
            'store_id': store_id,
            'product_id': product['id'],
            'units_created': len(created_unit_ids),
            'status': 'pending',
            'staff_id': staff.get('id'),
        },
    )
    return jsonify({
        'ok': True,
        'units_created': len(created_unit_ids),
        'created_unit_ids': created_unit_ids,
        'created_unit_serials': [u.get('unit_serial') for u in new_units],
        'stock': int((inv or {}).get('stock') or 0),
        'inventory_unchanged': True,
        'status': 'pending',
    })


def _mobile_resolve_store(staff, requested=None):
    requested = (requested or '').strip()
    if staff.get('role') == ROLE_SUPER:
        return requested
    return (staff.get('store_id') or '').strip()


def _mobile_assert_store(staff, store_id):
    if staff.get('role') == ROLE_SUPER:
        return None
    locked = (staff.get('store_id') or '').strip()
    if not locked:
        return jsonify({'error': 'Your account is not assigned to a store'}), 403
    if store_id and store_id != locked:
        return jsonify({'error': 'You can only access your assigned store'}), 403
    return None


def _mobile_can_manage_inventory(staff):
    return staff.get('role') in (ROLE_SUPER, ROLE_STORE)


@app.route('/api/mobile/pos/catalog', methods=['GET'])
@mobile_auth_required
def api_mobile_pos_catalog():
    """Products with in-stock variants for the selected store (billing)."""
    staff = request.mobile_staff
    store_id = _mobile_resolve_store(staff, request.args.get('store_id'))
    if not store_id:
        return jsonify({'error': 'Select a store'}), 400
    denied = _mobile_assert_store(staff, store_id)
    if denied:
        return denied
    categories = db_find('categories')
    products = db_find('products')
    inventory_rows = db_find('inventory', {'store_id': store_id})
    inv_by_product = {}
    for row in inventory_rows:
        inv_by_product.setdefault(row.get('product_id'), []).append(row)
    catalog_products = []
    for p in products:
        if (p.get('status') or 'available') == 'disabled':
            continue
        related = inv_by_product.get(p.get('id')) or []
        variants = []
        for r in related:
            stock = int(r.get('stock') or 0)
            if stock < 1:
                continue
            variant_label = ''
            for v in p.get('variants') or []:
                if v.get('id') == r.get('variant_id'):
                    variant_label = v.get('label') or ''
                    break
            variants.append({
                'inventory_id': r.get('id'),
                'variant_id': r.get('variant_id'),
                'variant_label': variant_label or '—',
                'price': float(r.get('price') or 0),
                'stock': stock,
            })
        if not variants:
            continue
        catalog_products.append({
            'id': p.get('id'),
            'name': p.get('name'),
            'sku': p.get('sku') or '',
            'category_id': p.get('category_id') or '',
            'variants': variants,
        })
    catalog_products.sort(key=lambda r: (r.get('name') or '').lower())
    categories.sort(key=lambda c: (c.get('name') or '').lower())
    return jsonify({
        'store_id': store_id,
        'categories': [{
            'id': c.get('id'),
            'name': c.get('name'),
            'code': c.get('code') or '',
        } for c in categories],
        'products': catalog_products,
    })


@app.route('/api/mobile/pos/orders', methods=['GET', 'POST'])
@mobile_auth_required
def api_mobile_pos_orders():
    staff = request.mobile_staff
    if request.method == 'GET':
        store_id = _mobile_resolve_store(staff, request.args.get('store_id'))
        denied = _mobile_assert_store(staff, store_id) if store_id else None
        if denied:
            return denied
        if staff.get('role') != ROLE_SUPER and not store_id:
            return jsonify({'error': 'Your account is not assigned to a store'}), 403
        orders = _list_pos_orders(store_id, request.args.get('limit', 40))
        # Object wrapper keeps clients resilient; `items` is the canonical list.
        return jsonify({'ok': True, 'items': orders, 'orders': orders, 'count': len(orders)})

    data = parse_json()
    payload, status = _create_pos_order(data, staff)
    return jsonify(payload), status


@app.route('/api/mobile/pos/invoice/<order_id>', methods=['GET'])
@mobile_auth_required
def api_mobile_pos_invoice(order_id):
    staff = request.mobile_staff
    order = _find_order_for_receipt(order_id)
    if not order or order.get('channel') != 'in_store':
        return jsonify({'error': 'Bill not found'}), 404
    denied = _mobile_assert_store(staff, order.get('store_id'))
    if denied:
        return denied
    buf = _build_invoice_pdf(order)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'invoice_{order["order_id"]}.pdf',
        mimetype='application/pdf',
    )


@app.route('/api/mobile/pos/receipt/<order_id>', methods=['GET'])
@mobile_auth_required
def api_mobile_pos_receipt(order_id):
    staff = request.mobile_staff
    order = _find_order_for_receipt(order_id)
    if not order or order.get('channel') != 'in_store':
        return jsonify({'error': 'Bill not found'}), 404
    denied = _mobile_assert_store(staff, order.get('store_id'))
    if denied:
        return denied
    if request.args.get('format') == 'json':
        return jsonify({'ok': True, 'receipt': _build_receipt_data(order)})
    auto_print = request.args.get('auto') == '1' or request.args.get('print') == '1'
    return make_response(_render_thermal_receipt_html(order, auto_print=auto_print))


@app.route('/api/mobile/inventory', methods=['GET', 'POST'])
@mobile_auth_required
def api_mobile_inventory():
    staff = request.mobile_staff
    if request.method == 'POST':
        if not _mobile_can_manage_inventory(staff):
            return jsonify({'error': 'Only Super Admin / Store Admin can edit inventory'}), 403
        data = parse_json()
        row = db_find_one('inventory', {'id': data.get('inventory_id')})
        if not row:
            return jsonify({'error': 'Inventory row not found'}), 404
        denied = _mobile_assert_store(staff, row.get('store_id'))
        if denied:
            return denied
        try:
            quantity = int(data.get('quantity', 0))
        except (TypeError, ValueError):
            return jsonify({'error': 'Quantity must be a whole number'}), 400
        if quantity < 1 or quantity > 1000000:
            return jsonify({'error': 'Quantity must be between 1 and 1,000,000'}), 400
        updated = db_increment('inventory', {'id': row['id']}, 'stock', quantity)
        # Keep mobile Add Stock fast — no QR minting on quantity bumps.
        _badges_cache.clear()
        log_activity(
            'inventory',
            f"Mobile added {quantity} units of stock · {staff.get('name')}",
            {
                'inventory_id': row['id'],
                'store_id': row.get('store_id'),
                'product_id': row.get('product_id'),
                'quantity': quantity,
                'staff_id': staff.get('id'),
            },
        )
        return jsonify(updated)

    store_id = _mobile_resolve_store(staff, request.args.get('store_id'))
    if not store_id and staff.get('role') != ROLE_SUPER:
        return jsonify({'error': 'Your account is not assigned to a store'}), 403
    if store_id:
        denied = _mobile_assert_store(staff, store_id)
        if denied:
            return denied
    return jsonify(_enrich_inventory_rows(store_id))


@app.route('/api/mobile/inventory/<inv_id>', methods=['PUT'])
@mobile_auth_required
def api_mobile_inventory_update(inv_id):
    staff = request.mobile_staff
    if not _mobile_can_manage_inventory(staff):
        return jsonify({'error': 'Only Super Admin / Store Admin can edit inventory'}), 403
    row = db_find_one('inventory', {'id': inv_id})
    if not row:
        return jsonify({'error': 'Inventory row not found'}), 404
    denied = _mobile_assert_store(staff, row.get('store_id'))
    if denied:
        return denied
    data = parse_json()
    updates = {}
    try:
        if 'price' in data:
            updates['price'] = float(data['price'])
        if 'stock' in data:
            updates['stock'] = int(data['stock'])
    except (TypeError, ValueError):
        return jsonify({'error': 'Price and stock must be valid numbers'}), 400
    if updates.get('price', 0) < 0 or updates.get('stock', 0) < 0:
        return jsonify({'error': 'Price and stock cannot be negative'}), 400
    updates['updated_at'] = now_iso()
    db_update('inventory', {'id': inv_id}, updates)
    updated = db_find_one('inventory', {'id': inv_id})
    # Fast mobile inventory path — no QR mint/void (POS billing untouched).
    _badges_cache.clear()
    log_activity(
        'inventory',
        f"Mobile inventory update · {staff.get('name')}",
        {'inventory_id': inv_id, 'updates': updates, 'staff_id': staff.get('id')},
    )
    return jsonify(updated)


# --- Settings ---

@app.route('/api/admin/settings', methods=['GET', 'PUT'])
@admin_required
def api_admin_settings():
    if request.method == 'GET':
        return jsonify(get_settings())
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can change settings'}), 403
    data = parse_json()
    numeric = ['low_stock_threshold', 'min_order_value', 'delivery_fee_below_min',
               'free_delivery_above', 'default_delivery_radius_km']
    for k in numeric:
        if k in data:
            try:
                data[k] = float(data[k])
            except (TypeError, ValueError):
                data.pop(k)
    if 'order_statuses' in data and isinstance(data['order_statuses'], str):
        data['order_statuses'] = [s.strip() for s in data['order_statuses'].split(',') if s.strip()]
    result = save_settings(data)
    log_activity('system', 'Settings updated')
    return jsonify(result)


# --- Coupons CRUD ---

@app.route('/api/admin/coupons', methods=['GET', 'POST'])
@admin_required
def api_admin_coupons():
    if request.method == 'GET':
        return jsonify(db_find('coupons', sort=[('created_at', -1)]))
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can manage coupons'}), 403
    data = parse_json()
    code = (data.get('code') or '').strip().upper()
    if not code:
        return jsonify({'error': 'Coupon code required'}), 400
    if db_find_one('coupons', {'code': code}):
        return jsonify({'error': 'Coupon code already exists'}), 400
    coupon = {
        'id': new_id('cpn_'),
        'code': code,
        'type': data.get('type', 'percent'),  # percent | flat
        'value': float(data.get('value', 0) or 0),
        'max_discount': float(data.get('max_discount') or 0) or None,
        'min_subtotal': float(data.get('min_subtotal', 0) or 0),
        'expires_at': data.get('expires_at', ''),
        'first_order_only': bool(data.get('first_order_only', False)),
        'active': bool(data.get('active', True)),
        'created_at': now_iso(),
    }
    db_insert('coupons', coupon)
    return jsonify(coupon), 201


@app.route('/api/admin/coupons/<coupon_id>', methods=['PUT', 'DELETE'])
@admin_required
def api_admin_coupon_detail(coupon_id):
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can manage coupons'}), 403
    if request.method == 'DELETE':
        db_delete('coupons', {'id': coupon_id})
        return jsonify({'ok': True})
    data = parse_json()
    allowed = ['type', 'value', 'max_discount', 'min_subtotal', 'expires_at', 'active', 'first_order_only']
    updates = {k: data[k] for k in allowed if k in data}
    if 'value' in updates:
        updates['value'] = float(updates['value'] or 0)
    if 'min_subtotal' in updates:
        updates['min_subtotal'] = float(updates['min_subtotal'] or 0)
    if 'first_order_only' in updates:
        updates['first_order_only'] = bool(updates['first_order_only'])
    updates['updated_at'] = now_iso()
    db_update('coupons', {'id': coupon_id}, updates)
    return jsonify(db_find_one('coupons', {'id': coupon_id}))


# --- Staff (login accounts + duty status) ---

def reveal_staff_password(member):
    if not member:
        return ''
    if _is_locked_recovery_staff(member):
        return RECOVERY_PASSWORD
    stored = unseal_staff_password(member.get('password_vault') or '')
    if stored:
        return stored
    hashed = member.get('password_hash') or ''
    if not hashed:
        return ''
    for candidate in (ADMIN_PASSWORD, 'abhi123', RECOVERY_PASSWORD):
        if candidate and verify_password(hashed, candidate):
            return candidate
    return ''


def hydrate_staff_password_vault(member):
    """Save a viewable copy when we can recover it, without changing the login hash."""
    if not member or not member.get('id'):
        return member
    if member.get('password_vault') and unseal_staff_password(member.get('password_vault')):
        return member
    plain = reveal_staff_password(member)
    if not plain:
        return member
    sealed = seal_staff_password(plain)
    db_update('staff', {'id': member['id']}, {
        'password_vault': sealed,
        'updated_at': now_iso(),
    })
    member['password_vault'] = sealed
    return member


def _staff_public(member, stores=None):
    stores = stores or {s['id']: s['name'] for s in db_find('stores')}
    row = {
        'id': member.get('id'),
        'name': member.get('name', ''),
        'username': member.get('username', ''),
        'role': member.get('role', ROLE_STORE),
        'store_id': member.get('store_id', ''),
        'store_name': stores.get(member.get('store_id'), 'All Stores'),
        'phone': member.get('phone', ''),
        'on_duty': bool(member.get('on_duty')),
        'active': member.get('active') is not False,
        'has_login': bool(member.get('password_hash') and member.get('username')),
        'locked_recovery': _is_locked_recovery_staff(member),
        'created_at': member.get('created_at'),
        'updated_at': member.get('updated_at'),
    }
    if admin_is_super():
        row['password'] = reveal_staff_password(member)
        row['password_known'] = bool(row['password'])
    return row


@app.route('/api/admin/staff', methods=['GET', 'POST'])
@admin_required
def api_admin_staff():
    stores = {s['id']: s['name'] for s in db_find('stores')}
    if request.method == 'GET':
        _dedupe_recovery_staff()
        staff = db_find('staff', sort=[('name', 1)])
        if admin_is_super():
            staff = [hydrate_staff_password_vault(m) for m in staff]
        scoped = resolve_store_scope(request.args.get('store_id'))
        role = session.get('admin_role')
        if role == ROLE_BILLING:
            staff = [m for m in staff if m.get('id') == session.get('admin_user_id')]
        elif scoped and not admin_is_super():
            staff = [
                m for m in staff
                if m.get('store_id') == scoped or m.get('id') == session.get('admin_user_id')
            ]
        return jsonify({
            'items': [_staff_public(m, stores) for m in staff],
            'roles': STAFF_ROLES if admin_is_super() else [ROLE_STORE, ROLE_BILLING],
        })

    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can create staff logins'}), 403
    data = parse_json()
    name = (data.get('name') or '').strip()
    username = (data.get('username') or '').strip().lower()
    password = data.get('password') or ''
    role = data.get('role') or ROLE_STORE
    store_id = data.get('store_id') or ''
    if not name:
        return jsonify({'error': 'Name required'}), 400
    if role not in STAFF_ROLES:
        return jsonify({'error': 'Invalid role'}), 400
    if role in (ROLE_STORE, ROLE_BILLING) and not store_id:
        return jsonify({'error': 'Store Admin and Billing Staff must be assigned to a store'}), 400
    if role == ROLE_SUPER:
        store_id = ''
    if not username:
        return jsonify({'error': 'Username required for admin login'}), 400
    if len(password) < 4:
        return jsonify({'error': 'Password must be at least 4 characters'}), 400
    if username == RECOVERY_USERNAME:
        return jsonify({'error': 'That username is reserved for the recovery Super Admin'}), 409
    if any((m.get('username') or '').lower() == username for m in db_find('staff')):
        return jsonify({'error': 'Username already exists'}), 409
    member = {
        'id': new_id('stf_'),
        'name': name,
        'username': username,
        'password_hash': hash_password(password),
        'password_vault': seal_staff_password(password),
        'role': role,
        'store_id': store_id,
        'phone': data.get('phone', ''),
        'on_duty': bool(data.get('on_duty', True)),
        'active': True,
        'created_at': now_iso(),
    }
    db_insert('staff', member)
    log_activity('system', f"Staff account {username} created ({role})")
    return jsonify(_staff_public(member, stores)), 201


@app.route('/api/admin/staff/<staff_id>', methods=['PUT', 'DELETE'])
@admin_required
def api_admin_staff_detail(staff_id):
    member = db_find_one('staff', {'id': staff_id})
    if not member:
        return jsonify({'error': 'Staff not found'}), 404
    stores = {s['id']: s['name'] for s in db_find('stores')}

    if request.method == 'DELETE':
        if not admin_is_super():
            return jsonify({'error': 'Only Super Admin can remove staff'}), 403
        if member.get('username') == 'abhi':
            return jsonify({'error': 'Cannot delete the primary Super Admin account'}), 400
        if _is_locked_recovery_staff(member):
            removed = _dedupe_recovery_staff()
            if not removed and _is_canonical_recovery_staff(member):
                return jsonify({'error': 'Cannot delete the recovery Super Admin account'}), 400
            return jsonify({'ok': True, 'removed_duplicates': removed})
        db_delete('staff', {'id': staff_id})
        return jsonify({'ok': True})

    data = parse_json()
    # Staff may toggle their own duty; Super Admin can edit everything.
    if not admin_is_super():
        if staff_id != session.get('admin_user_id'):
            # Store Admin can toggle duty for their store team
            if session.get('admin_role') != ROLE_STORE or member.get('store_id') != session.get('admin_store_id'):
                return jsonify({'error': 'Forbidden'}), 403
            if set(data.keys()) - {'on_duty'}:
                return jsonify({'error': 'You can only update duty status'}), 403
        updates = {'on_duty': bool(data.get('on_duty', member.get('on_duty'))), 'updated_at': now_iso()}
        db_update('staff', {'id': staff_id}, updates)
        return jsonify(_staff_public(db_find_one('staff', {'id': staff_id}), stores))

    updates = {}
    for key in ('name', 'phone', 'on_duty', 'active'):
        if key in data:
            updates[key] = data[key]
    if _is_locked_recovery_staff(member):
        if data.get('active') is False:
            return jsonify({'error': 'Cannot disable the recovery Super Admin account'}), 400
        if 'username' in data and (data.get('username') or '').strip().lower() != RECOVERY_USERNAME:
            return jsonify({'error': 'Cannot rename the recovery Super Admin account'}), 400
        if 'role' in data and data.get('role') != ROLE_SUPER:
            return jsonify({'error': 'Cannot change the recovery Super Admin role'}), 400
        data.pop('password', None)
    if 'role' in data:
        role = data['role']
        if role not in STAFF_ROLES:
            return jsonify({'error': 'Invalid role'}), 400
        updates['role'] = role
    if 'store_id' in data:
        updates['store_id'] = data.get('store_id') or ''
    role = updates.get('role', member.get('role'))
    store_id = updates.get('store_id', member.get('store_id') or '')
    if role == ROLE_SUPER:
        updates['store_id'] = ''
    elif role in (ROLE_STORE, ROLE_BILLING) and not store_id:
        return jsonify({'error': 'Store Admin and Billing Staff must be assigned to a store'}), 400
    if 'username' in data:
        username = (data.get('username') or '').strip().lower()
        if not username:
            return jsonify({'error': 'Username required'}), 400
        if username == RECOVERY_USERNAME and not _is_locked_recovery_staff(member):
            return jsonify({'error': 'That username is reserved for the recovery Super Admin'}), 409
        clash = next(
            (m for m in db_find('staff')
             if (m.get('username') or '').lower() == username and m.get('id') != staff_id),
            None,
        )
        if clash:
            return jsonify({'error': 'Username already exists'}), 409
        updates['username'] = username
    if data.get('password'):
        if len(data['password']) < 4:
            return jsonify({'error': 'Password must be at least 4 characters'}), 400
        updates['password_hash'] = hash_password(data['password'])
        updates['password_vault'] = seal_staff_password(data['password'])
    if 'on_duty' in updates:
        updates['on_duty'] = bool(updates['on_duty'])
    if 'active' in updates:
        updates['active'] = bool(updates['active'])
    updates['updated_at'] = now_iso()
    db_update('staff', {'id': staff_id}, updates)
    return jsonify(_staff_public(db_find_one('staff', {'id': staff_id}), stores))


# --- Activity feed & nav badges ---

@app.route('/api/admin/activity')
@admin_required
def api_admin_activity():
    limit = min(50, int(request.args.get('limit', 20)))
    rows = db_find('activity', sort=[('created_at', -1)], limit=limit)
    return jsonify(rows)


@app.route('/api/admin/security-events')
@admin_required
def api_admin_security_events():
    if not admin_is_super():
        return jsonify({'error': 'Only Super Admin can view security events'}), 403
    return jsonify({'items': recent_security_events(min(200, int(request.args.get('limit', 100))))})


@app.route('/api/admin/badges')
@admin_required
def api_admin_badges():
    settings = get_settings()
    threshold = int(settings['low_stock_threshold'])
    store_id = resolve_store_scope(request.args.get('store_id'))
    cache_key = store_id or '__all__'
    cached = _badges_cache.get(cache_key)
    if cached and (datetime.utcnow() - cached['at']).total_seconds() < _BADGES_CACHE_TTL:
        return jsonify(cached['payload'])

    order_query = {'status': {'$in': list(_OPEN_ORDER_STATUSES)}}
    inv_query = {'stock': {'$lte': threshold}}
    if store_id:
        order_query['store_id'] = store_id
        inv_query['store_id'] = store_id
    payload = {
        'orders': db_count('orders', order_query),
        'inventory': db_count('inventory', inv_query),
    }
    _badges_cache[cache_key] = {'at': datetime.utcnow(), 'payload': payload}
    if len(_badges_cache) > 20:
        oldest = sorted(_badges_cache.items(), key=lambda kv: kv[1]['at'])[:10]
        for key, _ in oldest:
            _badges_cache.pop(key, None)
    return jsonify(payload)


# --- Global search ---

@app.route('/api/admin/search')
@admin_required
def api_admin_search():
    q = (request.args.get('q') or '').strip().lower()
    if len(q) < 2:
        return jsonify({'orders': [], 'products': [], 'customers': [], 'categories': [], 'staff': []})
    rx = {'$regex': re.escape(q), '$options': 'i'}
    phone_rx = {'$regex': re.escape(q)}
    orders = db_find(
        'orders',
        {'$or': [{'order_id': rx}, {'customer_name': rx}, {'customer_phone': phone_rx}]},
        sort=[('created_at', -1)],
        limit=6,
        projection={
            'order_id': 1, 'customer_name': 1, 'total': 1, 'status': 1,
        },
    )
    products = db_find(
        'products',
        {'$or': [{'name': rx}, {'sku': rx}]},
        limit=6,
        projection={'id': 1, 'name': 1, 'sku': 1},
    )
    customers = db_find(
        'customers',
        {'$or': [{'name': rx}, {'phone': phone_rx}, {'email': rx}]},
        limit=6,
        projection={'id': 1, 'name': 1, 'phone': 1},
    )
    categories = db_find(
        'categories',
        {'$or': [{'name': rx}, {'slug': rx}]},
        limit=6,
        projection={'id': 1, 'name': 1, 'slug': 1},
    )
    staff = db_find(
        'staff',
        {'$or': [{'name': rx}, {'phone': phone_rx}, {'role': rx}]},
        limit=6,
        projection={'id': 1, 'name': 1, 'role': 1, 'phone': 1},
    )
    return jsonify({
        'orders': [{'order_id': o['order_id'], 'customer_name': o.get('customer_name'),
                    'total': o.get('total'), 'status': normalize_status(o.get('status'))} for o in orders],
        'products': [{'id': p['id'], 'name': p['name'], 'sku': p.get('sku')} for p in products],
        'customers': [{'id': c['id'], 'name': c['name'], 'phone': c.get('phone')} for c in customers],
        'categories': [{'id': c['id'], 'name': c['name'], 'slug': c.get('slug')} for c in categories],
        'staff': [{'id': m['id'], 'name': m['name'], 'role': m.get('role'), 'phone': m.get('phone')} for m in staff],
    })


# --- Tax invoice PDF ---

@app.route('/api/admin/orders/<order_id>/invoice')
@admin_required
def api_admin_order_invoice(order_id):
    order = _find_order_for_receipt(order_id)
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    denied = assert_store_access(order.get('store_id'))
    if denied:
        return denied
    buf = _build_invoice_pdf(order)
    return send_file(buf, as_attachment=True, download_name=f'invoice_{order["order_id"]}.pdf',
                     mimetype='application/pdf')


@app.route('/api/admin/orders/<order_id>/receipt')
@admin_required
def api_admin_order_receipt(order_id):
    order = _find_order_for_receipt(order_id)
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    denied = assert_store_access(order.get('store_id'))
    if denied:
        return denied
    if request.args.get('format') == 'json':
        return jsonify({'ok': True, 'receipt': _build_receipt_data(order)})
    auto_print = request.args.get('auto') == '1' or request.args.get('print') == '1'
    return make_response(_render_thermal_receipt_html(order, auto_print=auto_print))


# --- Reports PDF / XLSX ---

def _report_dataset(store_id=None, period=None, anchor=None, store_ids=None):
    if store_ids is None:
        if store_id:
            store_ids = [store_id]
        else:
            store_ids = []

    period_caption = 'All time'
    start = end = None
    if period in ('day', 'month', 'quarter', 'year'):
        if anchor is not None:
            normalized = _normalize_anchor(period, anchor)
            if not normalized:
                raise ValueError('Invalid period selection')
            period_caption = _selection_caption(period, normalized)
            anchor = normalized
            start, end = _created_at_range(period, anchor)
        else:
            period_caption = PERIOD_CAPTIONS.get(period, 'All time')
            start, end = _created_at_range(period, None)

    orders = db_find('orders', _store_date_query(store_ids, start, end))
    if period in ('day', 'month', 'quarter', 'year') and anchor is None:
        buckets = _build_timeline_buckets(period)
        valid_keys = set(buckets.keys())
        orders = [o for o in orders if _timeline_key(o.get('created_at'), period) in valid_keys]

    stores = {s['id']: s for s in _cached_collection('stores', lambda: db_find('stores'))}
    customer_query = {}
    if start or end:
        created = {}
        if start:
            created['$gte'] = start
        if end:
            created['$lt'] = end
        customer_query['created_at'] = created
    customers = db_find('customers', customer_query)

    store_sales = {}
    product_perf = {}
    status_counts = {}
    sales_total = 0
    for o in orders:
        total = o.get('total', 0) or 0
        sales_total += total
        sid = o.get('store_id')
        row = store_sales.get(sid)
        if row is None:
            row = {'name': stores.get(sid, {}).get('name', sid), 'sales': 0, 'orders': 0}
            store_sales[sid] = row
        row['sales'] += total
        row['orders'] += 1
        st = normalize_status(o.get('status'))
        status_counts[st] = status_counts.get(st, 0) + 1
        for it in o.get('items') or []:
            pid = it.get('product_id')
            if not pid:
                continue
            pref = product_perf.get(pid)
            if pref is None:
                pref = {'name': it.get('name', pid), 'qty': 0, 'revenue': 0}
                product_perf[pid] = pref
            pref['qty'] += it.get('qty', 0) or 0
            pref['revenue'] += (it.get('price', 0) or 0) * (it.get('qty', 0) or 0)

    # Inventory snapshot (batch product lookup — no N+1)
    settings = get_settings()
    threshold = int(settings['low_stock_threshold'])
    products_by_id = _cached_collection(
        'products_by_id',
        lambda: {p['id']: p for p in db_find('products')}
    )
    inv_query = _store_date_query(store_ids) if store_ids else {}
    inventory = []
    for inv in db_find('inventory', inv_query):
        p = products_by_id.get(inv.get('product_id'))
        variant_label = ''
        if p:
            for v in p.get('variants') or []:
                if v.get('id') == inv.get('variant_id'):
                    variant_label = v.get('label', '')
                    break
        stock = inv.get('stock', 0) or 0
        inventory.append({
            'product': p['name'] if p else inv.get('product_id'),
            'variant': variant_label,
            'store': stores.get(inv.get('store_id'), {}).get('name', inv.get('store_id')),
            'price': inv.get('price', 0),
            'stock': stock,
            'low': stock <= threshold,
        })

    return {
        'orders': sorted(orders, key=lambda x: x.get('created_at', ''), reverse=True),
        'customers': sorted(customers, key=lambda x: x.get('created_at', ''), reverse=True),
        'customers_total': len(customers),
        'sales_total': sales_total,
        'orders_total': len(orders),
        'store_sales': list(store_sales.values()),
        'avg_order': (sales_total / len(orders)) if orders else 0,
        'product_perf': sorted(product_perf.values(), key=lambda x: x['revenue'], reverse=True),
        'inventory': sorted(inventory, key=lambda x: x['stock']),
        'status_counts': status_counts,
        'period': period,
        'anchor': anchor,
        'period_caption': period_caption,
    }


@app.route('/api/admin/reports/xlsx')
@admin_required
def api_report_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    store_ids = resolve_store_ids()
    period = request.args.get('period', 'month')
    if period not in ('day', 'month', 'quarter', 'year'):
        period = 'month'
    anchor = request.args.get('anchor')
    try:
        data = _report_dataset(store_ids=store_ids, period=period, anchor=anchor)
    except ValueError as exc:
        return jsonify({'error': public_error(exc, 'Invalid report request')}), 400
    wb = Workbook()

    ws = wb.active
    ws.title = 'Summary'
    header_fill = PatternFill('solid', fgColor='1E3A22')
    header_font = Font(color='FFFFFF', bold=True)
    gold_fill = PatternFill('solid', fgColor='E7B430')

    ws['A1'] = 'Fish and Meat — Sales Report'
    ws['A1'].font = Font(bold=True, size=14, color='1E3A22')
    ws['A2'] = f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")} · {data["period_caption"]}'
    ws['A3'] = f'Storage: {db_mode()}'

    ws['A5'] = 'KPI'
    ws['B5'] = 'Value'
    for cell in (ws['A5'], ws['B5']):
        cell.fill = header_fill
        cell.font = header_font

    kpis = [
        ('Total Sales (₹)', data['sales_total']),
        ('Total Orders', data['orders_total']),
        ('Customers', data['customers_total']),
        ('Avg Order Value (₹)', round(data['avg_order'], 2)),
    ]
    for i, (k, v) in enumerate(kpis, start=6):
        ws[f'A{i}'] = k
        ws[f'B{i}'] = v

    ws['A11'] = 'Store-wise Contribution'
    ws['A11'].font = Font(bold=True, color='A5342A')
    ws['A12'] = 'Store'
    ws['B12'] = 'Orders'
    ws['C12'] = 'Sales (₹)'
    for col in ('A', 'B', 'C'):
        ws[f'{col}12'].fill = gold_fill
        ws[f'{col}12'].font = Font(bold=True, color='1E3A22')
    for i, s in enumerate(data['store_sales'], start=13):
        ws[f'A{i}'] = s['name']
        ws[f'B{i}'] = s['orders']
        ws[f'C{i}'] = s['sales']

    ws2 = wb.create_sheet('Order History')
    headers = ['Order ID', 'Date', 'Customer', 'Phone', 'Store', 'Status', 'Items', 'Total']
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    store_lookup = {s['id']: s['name'] for s in db_find('stores')}
    for row, o in enumerate(data['orders'], 2):
        ws2.cell(row, 1, o.get('order_id'))
        ws2.cell(row, 2, (o.get('created_at') or '')[:16])
        ws2.cell(row, 3, o.get('customer_name'))
        ws2.cell(row, 4, o.get('customer_phone'))
        ws2.cell(row, 5, store_lookup.get(o.get('store_id'), ''))
        ws2.cell(row, 6, o.get('status'))
        ws2.cell(row, 7, ', '.join(f"{i.get('name')} x{i.get('qty')}" for i in o.get('items') or []))
        ws2.cell(row, 8, o.get('total'))

    ws3 = wb.create_sheet('Product Performance')
    for col, h in enumerate(['Product', 'Units Sold', 'Revenue (₹)'], 1):
        cell = ws3.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for row, p in enumerate(data['product_perf'], 2):
        ws3.cell(row, 1, p['name'])
        ws3.cell(row, 2, p['qty'])
        ws3.cell(row, 3, p['revenue'])

    ws4 = wb.create_sheet('Inventory')
    for col, h in enumerate(['Product', 'Variant', 'Store', 'Price (₹)', 'Stock', 'Low Stock?'], 1):
        cell = ws4.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for row, r in enumerate(data['inventory'], 2):
        ws4.cell(row, 1, r['product'])
        ws4.cell(row, 2, r['variant'])
        ws4.cell(row, 3, r['store'])
        ws4.cell(row, 4, r['price'])
        ws4.cell(row, 5, r['stock'])
        ws4.cell(row, 6, 'YES' if r['low'] else '')

    ws5 = wb.create_sheet('Customers')
    for col, h in enumerate(['Name', 'Phone', 'Email', 'Address', 'Joined'], 1):
        cell = ws5.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for row, c in enumerate(data['customers'], 2):
        ws5.cell(row, 1, c.get('name'))
        ws5.cell(row, 2, c.get('phone'))
        ws5.cell(row, 3, c.get('email'))
        ws5.cell(row, 4, c.get('address'))
        ws5.cell(row, 5, (c.get('created_at') or '')[:10])

    ws6 = wb.create_sheet('Order Status')
    for col, h in enumerate(['Status', 'Count'], 1):
        cell = ws6.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for row, (st, count) in enumerate(sorted(data['status_counts'].items()), 2):
        ws6.cell(row, 1, st.replace('_', ' ').title())
        ws6.cell(row, 2, count)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'fam_report_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/admin/reports/pdf')
@admin_required
def api_report_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    store_ids = resolve_store_ids()
    period = request.args.get('period', 'month')
    if period not in ('day', 'month', 'quarter', 'year'):
        period = 'month'
    anchor = request.args.get('anchor')
    try:
        data = _report_dataset(store_ids=store_ids, period=period, anchor=anchor)
    except ValueError as exc:
        return jsonify({'error': public_error(exc, 'Invalid report request')}), 400

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleFAM', parent=styles['Heading1'], textColor=colors.HexColor('#1E3A22'), fontSize=18)
    sub_style = ParagraphStyle('SubFAM', parent=styles['Normal'], textColor=colors.HexColor('#55594F'), fontSize=10)
    brand_style = ParagraphStyle('Brand', parent=styles['Heading2'], textColor=colors.HexColor('#A5342A'), fontSize=12)

    story = []
    story.append(Paragraph('FISH AND MEAT — Admin Sales Report', title_style))
    story.append(Paragraph(f'Generated {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")} · {data["period_caption"]} · DB: {db_mode()}', sub_style))
    story.append(Spacer(1, 16))

    kpi_data = [
        ['Total Sales', f"₹{data['sales_total']:,.0f}"],
        ['Total Orders', str(data['orders_total'])],
        ['Customers', str(data['customers_total'])],
        ['Avg Order Value', f"₹{data['avg_order']:,.0f}"],
    ]
    kpi_table = Table(kpi_data, colWidths=[3 * inch, 2.5 * inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FBF6EC')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1E3A22')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#A5342A')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('PADDING', (0, 0), (-1, -1), 10),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E7E2D6')),
        ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor('#E7E2D6')),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph('Store-wise Contribution', brand_style))
    store_rows = [['Store', 'Orders', 'Sales (₹)']]
    for s in data['store_sales']:
        store_rows.append([s['name'], str(s['orders']), f"{s['sales']:,.0f}"])
    st = Table(store_rows, colWidths=[2.5 * inch, 1.5 * inch, 2 * inch])
    st.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A22')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFFDF8')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E7E2D6')),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(st)
    story.append(Spacer(1, 20))

    story.append(Paragraph('Order History (latest 40)', brand_style))
    order_rows = [['Order', 'Date', 'Customer', 'Status', 'Total']]
    for o in data['orders'][:40]:
        order_rows.append([
            o.get('order_id', ''),
            (o.get('created_at') or '')[:10],
            (o.get('customer_name') or '')[:18],
            o.get('status', ''),
            f"₹{o.get('total', 0):,.0f}",
        ])
    ot = Table(order_rows, colWidths=[1.2 * inch, 1 * inch, 1.6 * inch, 1.2 * inch, 1 * inch])
    ot.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E7B430')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1E3A22')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E7E2D6')),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBF6EC')]),
    ]))
    story.append(ot)
    story.append(Spacer(1, 20))

    story.append(Paragraph('Product Performance (top 15)', brand_style))
    pp_rows = [['Product', 'Units Sold', 'Revenue (₹)']]
    for p in data['product_perf'][:15]:
        pp_rows.append([p['name'][:32], str(p['qty']), f"{p['revenue']:,.0f}"])
    pt = Table(pp_rows, colWidths=[3 * inch, 1.4 * inch, 1.6 * inch])
    pt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A22')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E7E2D6')),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBF6EC')]),
    ]))
    story.append(pt)
    story.append(Spacer(1, 20))

    story.append(Paragraph('Inventory — Lowest Stock First (top 20)', brand_style))
    inv_rows = [['Product', 'Variant', 'Store', 'Price (₹)', 'Stock']]
    for r in data['inventory'][:20]:
        inv_rows.append([
            r['product'][:26], r['variant'][:14], r['store'][:14],
            f"{r['price']:,.0f}", ('LOW · ' if r['low'] else '') + str(r['stock']),
        ])
    it = Table(inv_rows, colWidths=[1.9 * inch, 1.1 * inch, 1.1 * inch, 0.9 * inch, 1 * inch])
    it.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E7B430')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1E3A22')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E7E2D6')),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FBF6EC')]),
    ]))
    story.append(it)
    story.append(Spacer(1, 20))

    story.append(Paragraph('Order Status Summary', brand_style))
    st_rows = [['Status', 'Count']]
    for st, count in sorted(data['status_counts'].items()):
        st_rows.append([st.replace('_', ' ').title(), str(count)])
    stt = Table(st_rows, colWidths=[3 * inch, 1.5 * inch])
    stt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A22')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E7E2D6')),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(stt)

    doc.build(story)
    buf.seek(0)
    filename = f'fam_report_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


# ---------------------------------------------------------------------------
# Main — both entry styles are supported:
#   1) python app.py
#   2) gunicorn -c gunicorn.conf.py app:app   (nginx → gunicorn multi-worker)
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    print('Fish and Meat server')
    print(f'  Storefront: http://127.0.0.1:5000/')
    print(f'  Admin:      http://127.0.0.1:5000/admin/login')
    print(f'  Database:   {db_mode()}')
    print('  Dev run:    python app.py')
    print('  Prod run:   gunicorn -c gunicorn.conf.py "app:app"')
    debug = FLASK_DEBUG and not IS_PRODUCTION
    if IS_PRODUCTION and FLASK_DEBUG:
        print('  WARNING: FLASK_DEBUG ignored in production')
    # Prefer localhost bind in production; 0.0.0.0 only when explicitly requested
    # (nginx usually proxies to 127.0.0.1:8000 via gunicorn — see gunicorn.conf.py)
    host = os.getenv('FAM_HOST', '127.0.0.1' if IS_PRODUCTION else '0.0.0.0')
    port = int(os.getenv('FAM_PORT', '5000'))
    app.run(debug=debug, host=host, port=port, use_reloader=debug)
